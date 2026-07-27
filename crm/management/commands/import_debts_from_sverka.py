"""Import opening balances straight from the reconciled ХОЗМАГЛАР workbook's
АКТ СВЕРКА sheet — a single authoritative source that carries BOTH the net balance
and the date the client last took goods:

  * ОСТАТКА (col E) -> the client's net balance as of go-live
        negative -> the client owes us  -> an opening-balance Sale (is_opening=True,
                    opening_amount=abs). No line items, so it never touches any
                    revenue/profit/sold-kg report — it shows up purely as a receivable.
        positive -> we hold their money  -> an ADVANCE_IN payment (advance credit).
  * ОЛДИ    (col F) -> the date the client last took goods; the debt is dated from it
        so it ages correctly, and the deadline = ОЛДИ + N days (default 14).

Unlike `import_opening_debts` (amounts from ОСТ КОЧА, dates from the sverka), this reads
everything from the one АКТ СВЕРКА sheet. Idempotent: a client that already has an
opening Sale (or an imported advance) is skipped, so re-running never doubles a balance.
Run with --dry-run first to see the totals without writing anything.
"""

import datetime
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from accounts.models import User
from crm.models import Client, Payment, Sale

# Reuse the exact name/phone parsing the sibling importer uses, so a client imported
# by either command dedupes to the same record.
from crm.management.commands.import_opening_debts import (
    ADVANCE_NOTE,
    _norm,
    split_name_phone,
)

DEFAULT_FILE = r"C:\Users\admin\Downloads\Telegram Desktop\ХОЗМАГЛАР 2026 (2) (4) (3).xlsx"
SHEET = "АКТ СВЕРКА"
BALANCE_COL = 4  # ОСТАТКА (0-indexed)
OLDI_COL = 5     # ОЛДИ
# Where a debtor has no usable ОЛДИ date (missing, or an implausible 1900/future typo),
# fall back to this — mostly old balances carried over from before June.
FALLBACK_DATE = datetime.date(2026, 6, 1)
DEFAULT_DEADLINE_DAYS = 14  # to'lov muddati = ОЛДИ + shu kun


class Command(BaseCommand):
    help = "ХОЗМАГЛАР АКТ СВЕРКА varag'idan qarz/avans qoldiqlarini import qiladi (ОСТАТКА + ОЛДИ)."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE, help="ХОЗМАГЛАР .xlsx yo'li")
        parser.add_argument("--seller", default="admin",
                            help="Ega/sotuvchi username (default: admin)")
        parser.add_argument("--seller-email", default=None,
                            help="Ega/sotuvchini EMAIL bo'yicha topadi (username o'rniga).")
        parser.add_argument("--production-debt", default=None,
                            help="Shu sotuvchining ishlab chiqarishga ochilish qarzi (so'm).")
        parser.add_argument("--deadline-days", type=int, default=DEFAULT_DEADLINE_DAYS,
                            help="To'lov muddati = ОЛДИ + shu kun (default: 14)")
        parser.add_argument("--fallback-date", default=None,
                            help="ОЛДИ sanasi yo'q qarzdorlar uchun sana (YYYY-MM-DD). "
                                 f"Berilmasa: {FALLBACK_DATE}")
        parser.add_argument("--dry-run", action="store_true",
                            help="Hech narsa yozmaydi, faqat xulosani ko'rsatadi.")

    def handle(self, *args, **opt):
        import openpyxl

        # Prefer an explicit email (the seller may exist on prod under an unknown
        # username); fall back to the username otherwise.
        if opt["seller_email"]:
            qs = User.objects.filter(email__iexact=opt["seller_email"])
            if not qs.exists():
                raise CommandError(f"Bu email bilan sotuvchi topilmadi: {opt['seller_email']}")
            if qs.count() > 1:
                raise CommandError(f"Bu email bir nechta hisobda: {opt['seller_email']}")
            seller = qs.first()
        else:
            try:
                seller = User.objects.get(username=opt["seller"])
            except User.DoesNotExist:
                raise CommandError(f"Sotuvchi topilmadi: {opt['seller']}")

        today = timezone.localdate()
        deadline_days = opt["deadline_days"]

        fallback_date = FALLBACK_DATE
        if opt["fallback_date"]:
            try:
                fallback_date = datetime.date.fromisoformat(opt["fallback_date"])
            except ValueError:
                raise CommandError(
                    f"--fallback-date noto'g'ri: {opt['fallback_date']} (YYYY-MM-DD kutilgan)"
                )

        wb = openpyxl.load_workbook(opt["file"], read_only=True, data_only=True)
        if SHEET not in wb.sheetnames:
            raise CommandError(f"'{SHEET}' varag'i topilmadi. Bor varaqlar: {wb.sheetnames}")
        ws = wb[SHEET]

        # (clean_name, phone, balance, oldi_date_or_None) for every non-zero-balance row.
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            name = _norm(r[0]) if r and r[0] is not None else ""
            if not name:
                continue
            val = r[BALANCE_COL] if len(r) > BALANCE_COL else None
            if not isinstance(val, (int, float)) or val == 0:
                continue
            oldi = r[OLDI_COL] if len(r) > OLDI_COL else None
            oldi_date = None
            if isinstance(oldi, datetime.datetime):
                d = oldi.date()
                if d.year >= 2020 and d <= today:  # drop 1900 / future typos
                    oldi_date = d
            clean_name, phone = split_name_phone(name)
            rows.append((clean_name, phone, Decimal(str(val)), oldi_date))

        dry = opt["dry_run"]
        created_debt = created_adv = skipped_existing = 0
        dated_from_oldi = dated_from_fallback = 0
        debt_total = adv_total = Decimal("0")

        with transaction.atomic():
            for clean_name, phone, val, oldi_date in rows:
                client = Client.find_duplicate(seller, clean_name) or Client.objects.filter(
                    name__iexact=clean_name
                ).first()
                if client is None and not dry:
                    client = Client.objects.create(
                        name=clean_name, phone=phone, owner=seller
                    )

                if val < 0:  # client owes us -> opening debt
                    amount = -val
                    exists = client is not None and Sale.objects.filter(
                        client=client, is_opening=True
                    ).exists()
                    if exists:
                        skipped_existing += 1
                        continue
                    created_debt += 1
                    debt_total += amount
                    if oldi_date:
                        dated_from_oldi += 1
                    else:
                        dated_from_fallback += 1
                    debt_date = oldi_date or fallback_date
                    deadline = debt_date + datetime.timedelta(days=deadline_days)
                    if not dry:
                        Sale.objects.create(
                            client=client, sales_rep=seller, date=debt_date,
                            debt_deadline=deadline, is_opening=True,
                            opening_amount=amount,
                        )
                else:  # val > 0 -> we hold their money -> advance credit
                    exists = client is not None and Payment.objects.filter(
                        client=client, kind=Payment.Kind.ADVANCE_IN, note=ADVANCE_NOTE
                    ).exists()
                    if exists:
                        skipped_existing += 1
                        continue
                    created_adv += 1
                    adv_total += val
                    if not dry:
                        Payment.objects.create(
                            client=client, created_by=seller, date=today,
                            amount=val, kind=Payment.Kind.ADVANCE_IN, note=ADVANCE_NOTE,
                            is_opening=True,
                        )

            # Opening production debt (what the seller owes production at go-live) — a
            # per-seller figure that flows into the kassa's "Ishlab chiqarishga qarz".
            prod_debt = None
            if opt["production_debt"] is not None:
                prod_debt = Decimal(str(opt["production_debt"]).replace(" ", ""))
                if not dry:
                    seller.opening_production_debt = prod_debt
                    seller.save(update_fields=["opening_production_debt"])

            if dry:
                transaction.set_rollback(True)

        head = "DRY-RUN (hech narsa yozilmadi)" if dry else "IMPORT bajarildi"
        self.stdout.write(self.style.SUCCESS(f"\n=== {head} ==="))
        self.stdout.write(
            f"Manba: {SHEET} | Sotuvchi: {seller.username} ({seller.email or 'email yoq'}) "
            f"| Muddat = ОЛДИ + {deadline_days} kun"
        )
        self.stdout.write(f"Qarz (opening) yozuvlari: {created_debt} ta = {debt_total:,.0f} so'm")
        self.stdout.write(f"  sana ОЛДИ'dan: {dated_from_oldi} ta | fallback ({fallback_date}): {dated_from_fallback} ta")
        self.stdout.write(f"Avans yozuvlari:          {created_adv} ta = {adv_total:,.0f} so'm")
        self.stdout.write(f"Allaqachon bor (o'tkazib): {skipped_existing} ta")
        if prod_debt is not None:
            self.stdout.write(f"Ishlab chiqarishga qarz o'rnatildi: {prod_debt:,.0f} so'm")
