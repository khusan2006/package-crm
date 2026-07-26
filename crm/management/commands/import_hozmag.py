"""Import the FULL detailed ХОЗМАГЛАР workbook (this file's own layout) into the CRM.

Unlike `import_excel` (which reads one consolidated `Лист1`) and
`import_debts_from_sverka` (which reads only the reconciled net balances), this reads
the three detailed sheets so the system derives every figure itself:

  * СОТУВ           — every sale line (САНА/МИЖОЗ/ТОВАР/СОНИ/НАРХИ/ТАННАРХ …). Grouped
                      by client+date into receipts. Negative-СОНИ rows are returns.
  * ТЎЛОВ           — every payment (САНА/МИЖОЗ/СЎМ/ВАЛЮТА/КУРС/ИЗОХ). Allocated FIFO
                      against that client's opening + sales, oldest first.
  * БОШЛАҒИЧ САЛЬДО — each client's net balance as of go-live (1-iyun). Negative → an
                      opening-balance Sale (is_opening, no line items). Positive → an
                      ADVANCE_IN credit.

Because every SaleItem carries its ТАННАРХ, the seller's production debt
(`seller_production_debt`) comes out as the tannarx of everything sold — exactly what
the owner asked for. Client debt = opening + sales − payments per client.

After importing, each client's computed debt is reconciled against the workbook's own
АКТ СВЕРКА (ОСТАТКА) sheet and any mismatch is reported. Run --dry-run first.
"""

import datetime
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import User
from crm.models import (
    AuditLog,
    Client,
    Expense,
    Payment,
    Product,
    ProductionReceipt,
    ProductionReceiptItem,
    ProductionRemittance,
    ProfitPayout,
    Return,
    Sale,
    SaleItem,
    StockEntry,
)
from crm.management.commands.import_opening_debts import split_name_phone, _norm

# Product-name variants in the file → one canonical (spacing only; grades stay distinct).
PRODUCT_ALIASES = {
    "ШПУЛ -": "ШПУЛ-",
    "ШПУЛ–": "ШПУЛ-",
    "ОҚ 2": "ОҚ 2 СОРТ",
    "ОҚ 1": "ОҚ 1 СОРТ",
}
METHOD_BY_NOTE = {"КЛИК": Payment.Method.CARD, "ПЕР": Payment.Method.TRANSFER}
DEADLINE_DAYS = 14
MIN_YEAR = 2020  # rows dated before this (stray 1900 cells) are dropped.
DEFAULT_FILE = str(Path(settings.BASE_DIR) / "data" / "sverka.xlsx")
# Historical products in the file that are NOT in the clean catalogue get this SKU
# prefix. Deliberately NOT "IMP-": `clear_imported` truncates every crm table while an
# IMP-* product exists, so an IMP marker that we keep around would re-wipe on each deploy.
HIST_SKU_PREFIX = "HIST-"
# Marks that this go-live import has already run (idempotency for --once on deploy).
GOLIVE_NOTE = "Go-live: ishlab chiqarishga avval topshirilgan (975 mln qarzga moslash)"

# crm_* models to empty before import, children before parents. Users are kept.
_CLEAR_ORDER = (
    Return, Payment, SaleItem, Sale, Expense,
    ProductionReceiptItem, ProductionReceipt, ProductionRemittance,
    ProfitPayout, StockEntry, AuditLog, Client, Product,
)


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _money(v):
    return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _qty(v):
    return Decimal(str(v)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)


def _prodname(v):
    name = " ".join(str(v).split()).upper()
    return PRODUCT_ALIASES.get(name, name)


class Command(BaseCommand):
    help = "ХОЗМАГЛАР to'liq faylini (СОТУВ+ТЎЛОВ+БОШЛАҒИЧ САЛЬДО) tizimga import qiladi."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE,
                            help=f"ХОЗМАГЛАР .xlsx yo'li (default: {DEFAULT_FILE})")
        parser.add_argument("--seller", default="kamola",
                            help="Ega/sotuvchi username (default: kamola)")
        parser.add_argument("--seller-email", default=None,
                            help="Ega/sotuvchini EMAIL bo'yicha topadi (username o'rniga). "
                                 "Production'da username boshqacha bo'lishi mumkin.")
        parser.add_argument("--production-debt", default=None,
                            help="Import oxirida shu qiymat i.ch. qarzi bo'lib qolishi uchun "
                                 "muvozanatlovchi topshiruv yoziladi (masalan 975956588).")
        parser.add_argument("--once", action="store_true",
                            help="Deploy uchun: allaqachon import qilingan bo'lsa (yoki "
                                 "sotuvchi/fayl topilmasa) jimgina o'tkazib yuboradi.")
        parser.add_argument("--dry-run", action="store_true",
                            help="Hech narsa yozmaydi — parse qilib, xulosa va sverka tekshiruvini ko'rsatadi.")

    # ---------------------------------------------------------------- parsing

    def parse_sales(self, ws):
        """СОТУВ: split into sales (qty>0) and returns (qty<0). Carry the last known
        tannarx forward per product (and backward for the head), like import_excel."""
        sales, returns, skipped = [], [], []
        last_valid_date = None
        for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            dt, client, product, unit, qty, price, cost = (
                r[0], r[1], r[2], r[3], _num(r[4]), _num(r[5]), _num(r[8])
            )
            if not (isinstance(dt, datetime.datetime) and client and product and qty):
                if any(v is not None for v in r[:11]):
                    skipped.append((i, "sana/mijoz/tovar/soni yo'q"))
                continue
            d = dt.date()
            if dt.year < MIN_YEAR:
                # A typo'd date (stray 1900 cell) on an otherwise real sale — the row is
                # counted in the file's own СОТИЛДИ total, so repair it with the previous
                # valid row's date rather than dropping a real sale.
                if last_valid_date is None:
                    skipped.append((i, f"sana xato ({d}), oldingi sana yo'q"))
                    continue
                d = last_valid_date
            else:
                last_valid_date = d
            row = {
                "line": i, "date": d,
                "client": _norm(client), "product": _prodname(product),
                "qty": _qty(abs(qty)), "price": _money(price or 0),
                "cost": _money(cost) if cost and cost > 0 else None,
            }
            (sales if qty > 0 else returns).append(row)

        sales.sort(key=lambda x: (x["date"], x["line"]))
        returns.sort(key=lambda x: (x["date"], x["line"]))
        last = {}
        for row in sales:
            if row["cost"] is not None:
                last[row["product"]] = row["cost"]
            else:
                row["cost"] = last.get(row["product"])
        first = {}
        for row in reversed(sales):
            if row["cost"] is not None:
                first[row["product"]] = row["cost"]
            else:
                row["cost"] = first.get(row["product"], Decimal("0"))
        return sales, returns, skipped

    def parse_payments(self, ws):
        rows, skipped = [], []
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            dt, client, som, usd, rate, jami, note = (
                r[0], r[1], _num(r[2]), _num(r[3]), _num(r[4]), r[5], r[7]
            )
            if not (isinstance(dt, datetime.datetime) and client):
                if any(v is not None for v in r[:8]):
                    skipped.append((i, "sana/mijoz yo'q"))
                continue
            if dt.year < MIN_YEAR:
                skipped.append((i, f"sana xato ({dt.date()})"))
                continue
            if not som and not usd:
                skipped.append((i, "summa yo'q"))
                continue
            # A voided/pending payment: the file zeroes its ЖАМИ (total) column so it is
            # left out of the reconciled ТЎЛАНДИ — respect that and skip it.
            if isinstance(jami, (int, float)) and jami == 0:
                skipped.append((i, "bekor qilingan (ЖАМИ=0)"))
                continue
            if usd and not rate:
                skipped.append((i, "valyuta bor, kurs yo'q"))
                continue
            note = _norm(note) if note else ""
            rows.append({
                "line": i, "date": dt.date(), "client": _norm(client),
                "som": _money(som) if som else Decimal("0"),
                "usd": _money(usd) if usd else Decimal("0"),
                "rate": _money(rate) if rate else Decimal("0"),
                "note": note,
                "method": METHOD_BY_NOTE.get(note.upper(), Payment.Method.CASH),
            })
        rows.sort(key=lambda x: (x["date"], x["line"]))
        return rows, skipped

    def parse_opening(self, ws):
        """БОШЛАҒИЧ САЛЬДО: {clean client -> Decimal saldo} (negative = client owes)."""
        rows, skipped = [], []
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            client, val = r[1], _num(r[2])
            if not client:
                continue
            if not val:
                skipped.append((i, "saldo 0/yo'q"))
                continue
            rows.append({"client": _norm(client), "saldo": Decimal(str(val))})
        return rows, skipped

    def parse_sverka(self, ws):
        """АКТ СВЕРКА: {clean client -> ОСТАТКА} — the reconciled net (negative=owes).
        Names carry a trailing phone here too, so strip it the same way as everywhere
        else; rows that collapse to one client (same base name) are summed."""
        out = defaultdict(Decimal)
        for r in ws.iter_rows(min_row=2, values_only=True):
            name = _norm(r[0]) if r and r[0] is not None else ""
            if not name:
                continue
            val = r[4] if len(r) > 4 else None
            if isinstance(val, (int, float)):
                clean, _ = split_name_phone(name)
                out[clean.upper()] += Decimal(str(val))
        return out

    # ------------------------------------------------------------------- run

    def handle(self, *args, **opt):
        once = opt["once"]

        # -- resolve the seller (by e-mail if given, else username)
        seller = self._resolve_seller(opt["seller_email"], opt["seller"], once)
        if seller is None:
            return  # --once: seller not on this instance yet — nothing to do

        # -- idempotency guard: skip if this go-live import already ran here
        if once and ProductionRemittance.objects.filter(note=GOLIVE_NOTE).exists():
            self.stdout.write("import_hozmag: allaqachon bajarilgan — o'tkazib yuborildi (--once).")
            return

        if once and not Path(opt["file"]).exists():
            self.stdout.write(self.style.WARNING(
                f"import_hozmag: fayl topilmadi ({opt['file']}) — o'tkazib yuborildi (--once)."))
            return

        try:
            wb = openpyxl.load_workbook(opt["file"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Fayl topilmadi: {opt['file']}")
        for sheet in ("СОТУВ", "ТЎЛОВ", "БОШЛАҒИЧ САЛЬДО", "АКТ СВЕРКА"):
            if sheet not in wb.sheetnames:
                raise CommandError(f"'{sheet}' varag'i topilmadi. Bor: {wb.sheetnames}")

        prod_target = None
        if opt["production_debt"] is not None:
            prod_target = Decimal(str(opt["production_debt"]).replace(" ", ""))

        sales, returns, sk_sales = self.parse_sales(wb["СОТУВ"])
        pays, sk_pays = self.parse_payments(wb["ТЎЛОВ"])
        openings, sk_open = self.parse_opening(wb["БОШЛАҒИЧ САЛЬДО"])
        sverka = self.parse_sverka(wb["АКТ СВЕРКА"])
        if not sales:
            raise CommandError("СОТУВ varag'idan bironta yaroqli sotuv topilmadi.")

        from crm.models import seller_production_debt
        with transaction.atomic():
            self.wipe()
            report = self.import_all(sales, returns, pays, openings, seller)
            recon = self.reconcile(sverka, seller)
            # Balance the production debt to the owner-supplied figure by recording the
            # cash already handed to production (the file has no remittance rows). Net
            # debt = sold cost − this remittance = the target.
            report["remittance"] = Decimal("0")
            if prod_target is not None:
                gross = seller_production_debt(seller)
                already_remitted = gross - prod_target
                if already_remitted > 0:
                    ProductionRemittance.objects.create(
                        seller=seller, created_by=seller,
                        date=datetime.date.today(), amount=already_remitted,
                        method=Payment.Method.CASH, note=GOLIVE_NOTE,
                    )
                    report["remittance"] = already_remitted
            # Compute the production debt INSIDE the transaction — on a dry-run the rows
            # are rolled back on block exit, so reading it afterwards would give 0.
            report["prod_debt"] = seller_production_debt(seller)
            if opt["dry_run"]:
                transaction.set_rollback(True)

        self.report(report, recon, sk_sales, sk_pays, sk_open, seller, opt["dry_run"])

    def _resolve_seller(self, email, username, once):
        """Find the owner: by e-mail when given (production, where the username may
        differ), else by username. Under --once a missing seller is not an error — the
        instance just isn't ready for the go-live import yet."""
        if email:
            qs = User.objects.filter(email__iexact=email)
            if qs.count() > 1:
                raise CommandError(f"Bu email bir nechta hisobda: {email}")
            seller = qs.first()
            if seller is None:
                if once:
                    self.stdout.write(
                        f"import_hozmag: '{email}' topilmadi — o'tkazib yuborildi (--once).")
                    return None
                raise CommandError(f"Bu email bilan sotuvchi topilmadi: {email}")
            return seller
        try:
            return User.objects.get(username=username)
        except User.DoesNotExist:
            if once:
                self.stdout.write(
                    f"import_hozmag: '{username}' topilmadi — o'tkazib yuborildi (--once).")
                return None
            raise CommandError(f"Sotuvchi topilmadi: {username}")

    def wipe(self):
        for model in _CLEAR_ORDER:
            model.objects.all().delete()
        # Reset every seller's carried production debt — this import derives it purely
        # from the sold tannarx, so a leftover opening figure would double-count.
        User.objects.exclude(opening_production_debt=0).update(opening_production_debt=0)
        self.stdout.write("Eski crm ma'lumotlari tozalandi; ishlab chiqarish ochilish qarzlari nollandi (userlar qoldi).")

    def import_all(self, sales, returns, pays, openings, seller):
        # -- products: seed the clean catalogue (7 priced products with razmer/mikron)
        #    for ongoing sales, map the file's names onto it, and create the leftover
        #    historical grades (ОҚ 1/2 СОРТ, ШПУЛ-, ҚОП …) as HIST-* products. Each
        #    SaleItem still carries its own per-line tannarx from the file, so mapping a
        #    name to the catalogue never loses the historical cost.
        call_command("seed_client_products")
        catalog = {p.name.upper(): p for p in Product.objects.all()}
        latest = {}
        for row in sales:
            latest[row["product"]] = row
        products = {}
        hist = 0
        for name, row in sorted(latest.items()):
            if name.upper() in catalog:
                products[name] = catalog[name.upper()]
            else:
                hist += 1
                products[name] = Product.objects.create(
                    name=name, sku=f"{HIST_SKU_PREFIX}{hist:02d}",
                    price=row["price"], cost_price=row["cost"] or Decimal("0"),
                )

        # -- clients: union of every sheet's names, deduped on clean name
        raw_names = (
            {r["client"] for r in sales} | {r["client"] for r in returns}
            | {p["client"] for p in pays} | {o["client"] for o in openings}
        )
        clients = {}          # upper(clean) -> Client
        client_for = {}       # raw name     -> Client
        for raw in sorted(raw_names):
            clean, phone = split_name_phone(raw)
            key = clean.upper()
            if key not in clients:
                clients[key] = Client.objects.create(
                    name=clean, phone=phone, owner=seller
                )
            elif phone and not clients[key].phone:
                clients[key].phone = phone
                clients[key].save(update_fields=["phone"])
            client_for[raw] = clients[key]

        first_date = min(r["date"] for r in sales)
        opening_date = first_date - datetime.timedelta(days=1)

        # slots per client pk: [{sale, remaining, products}] — payment-allocation queue
        slots = defaultdict(list)

        # -- opening balances FIRST (so payments pay the oldest debt first)
        n_open_debt = n_open_adv = 0
        open_debt_total = open_adv_total = Decimal("0")
        for o in openings:
            client = client_for[o["client"]]
            saldo = o["saldo"]
            if saldo < 0:                      # client owes us
                amount = _money(-saldo)
                sale = Sale.objects.create(
                    client=client, sales_rep=seller, date=opening_date,
                    debt_deadline=opening_date + datetime.timedelta(days=DEADLINE_DAYS),
                    is_opening=True, opening_amount=amount,
                )
                slots[client.pk].insert(0, {"sale": sale, "remaining": amount, "products": set()})
                n_open_debt += 1
                open_debt_total += amount
            else:                              # we hold their money -> advance credit
                Payment.objects.create(
                    client=client, created_by=seller, date=opening_date,
                    amount=_money(saldo), kind=Payment.Kind.ADVANCE_IN,
                    note="Ochilish avansi (import)", is_opening=True,
                )
                n_open_adv += 1
                open_adv_total += saldo

        # -- sales: group a client's rows on one date into one receipt
        grouped = defaultdict(list)
        for row in sales:
            grouped[(row["client"], row["date"])].append(row)
        item_index = defaultdict(list)  # (client_pk, product) -> [SaleItem] for returns
        for (raw_client, d), rows in sorted(grouped.items(), key=lambda kv: kv[0][1]):
            client = client_for[raw_client]
            sale = Sale.objects.create(date=d, client=client, sales_rep=seller)
            total = Decimal("0")
            for row in rows:
                item = SaleItem.objects.create(
                    sale=sale, product=products[row["product"]],
                    dimension=Sale.Dimension.KG, weight=row["qty"],
                    price=row["price"], cost_price=row["cost"] or Decimal("0"),
                    fulfilled_kg=row["qty"], fulfilled_at=d,
                )
                item_index[(client.pk, row["product"])].append(item)
                total += row["qty"] * row["price"]
            slots[client.pk].append({
                "sale": sale, "remaining": total,
                "products": {row["product"] for row in rows},
            })

        # -- returns (negative-qty rows): attach to the client's latest matching line
        n_returns = 0
        skipped_returns = []
        for row in returns:
            client = client_for[row["client"]]
            items = [
                it for it in item_index.get((client.pk, row["product"]), [])
                if it.sale.date <= row["date"]
            ] or item_index.get((client.pk, row["product"]), [])
            if not items:
                skipped_returns.append((row, "mos sotuv qatori yo'q"))
                continue
            item = items[-1]
            Return.objects.create(
                sale_item=item, weight=row["qty"], date=row["date"],
                restock=True, note="Импорт: возврат", created_by=seller,
            )
            for slot in slots[client.pk]:
                if slot["sale"].pk == item.sale_id:
                    slot["remaining"] -= row["qty"] * item.price
                    break
            n_returns += 1

        # -- payments: FIFO against the client's slots, splitting across sales
        n_payments = 0
        leftover_advance = 0
        for p in pays:
            client = client_for[p["client"]]
            queue = slots[client.pk]
            parts = []
            if p["som"]:
                parts.append(("uzs", p["som"], p["som"]))
            if p["usd"]:
                parts.append(("usd", _money(p["usd"] * p["rate"]), p["usd"]))
            for currency, som_value, original in parts:
                if som_value < 0:  # correction row: book against the first slot
                    target = queue[0]["sale"] if queue else None
                    if target is None:
                        continue
                    self.make_payment(p, target, currency, som_value, original, seller)
                    queue[0]["remaining"] -= som_value
                    n_payments += 1
                    continue
                left = som_value
                for slot in queue:
                    if left <= 0:
                        break
                    if slot["remaining"] <= 0:
                        continue
                    chunk = min(left, slot["remaining"])
                    orig = chunk if currency == "uzs" else _money(chunk / p["rate"])
                    self.make_payment(p, slot["sale"], currency, chunk, orig, seller)
                    slot["remaining"] -= chunk
                    left -= chunk
                    n_payments += 1
                if left > 0:
                    # client overpaid (advance) — park the surplus as client credit
                    orig = left if currency == "uzs" else _money(left / p["rate"])
                    Payment.objects.create(
                        client=client, created_by=seller, date=p["date"],
                        amount=left,
                        currency=(Payment.Currency.USD if currency == "usd" else Payment.Currency.UZS),
                        exchange_rate=p["rate"] if currency == "usd" else 0,
                        amount_original=orig if currency == "usd" else 0,
                        method=p["method"], note="Ortiqcha to'lov (import)",
                        kind=Payment.Kind.ADVANCE_IN,
                    )
                    n_payments += 1
                    leftover_advance += 1

        return {
            "products": Product.objects.count(), "hist_products": hist,
            "clients": len(clients),
            "sales": Sale.objects.filter(is_opening=False).count(),
            "sale_rows": len(sales),
            "open_debt": n_open_debt, "open_debt_total": open_debt_total,
            "open_adv": n_open_adv, "open_adv_total": open_adv_total,
            "returns": n_returns, "skipped_returns": skipped_returns,
            "payments": n_payments, "leftover_advance": leftover_advance,
        }

    def make_payment(self, p, sale, currency, som_value, original, seller):
        Payment.objects.create(
            date=p["date"], amount=som_value,
            currency=(Payment.Currency.USD if currency == "usd" else Payment.Currency.UZS),
            exchange_rate=p["rate"] if currency == "usd" else 0,
            amount_original=original if currency == "usd" else 0,
            method=p["method"], note=p["note"],
            kind=(Payment.Kind.SALE if p["date"] == sale.date else Payment.Kind.DEBT),
            sale=sale, created_by=seller,
        )

    # ---------------------------------------------------------- reconciliation

    def reconcile(self, sverka, seller):
        """Compare each client's computed net debt to the workbook's АКТ СВЕРКА ОСТАТКА.
        A client owes `-ОСТАТКА`; we compare that to the sum of their sale debts minus
        their advance credit. Returns (matches, mismatches, unmatched_names)."""
        # client owes = Σ sale.debt_remaining (opening + sales - allocated payments)
        #               - client-level advance credit (ADVANCE_IN not tied to a sale)
        by_client = defaultdict(Decimal)
        for sale in Sale.objects.filter(sales_rep=seller).prefetch_related("items", "payments", "returns"):
            by_client[sale.client_id] += sale.debt_remaining
        adv = defaultdict(Decimal)
        for pay in Payment.objects.filter(created_by=seller, sale__isnull=True,
                                          kind=Payment.Kind.ADVANCE_IN):
            adv[pay.client_id] += pay.amount
        names = {c.pk: c.name for c in Client.objects.filter(owner=seller)}

        matches, mism = 0, []
        seen = set()
        for pk, name in names.items():
            computed = _money(by_client.get(pk, Decimal("0")) - adv.get(pk, Decimal("0")))
            key = name.upper()
            seen.add(key)
            if key not in sverka:
                mism.append((name, computed, None, "sverkada yo'q"))
                continue
            expected = _money(-sverka[key])  # owes = -ОСТАТКА
            if abs(computed - expected) <= Decimal("1"):
                matches += 1
            else:
                mism.append((name, computed, expected, f"farq {computed - expected:,.0f}"))
        unmatched = [n for k, n in
                     ((k, k) for k in sverka) if k not in seen and sverka[k] != 0]
        return matches, mism, unmatched

    # ---------------------------------------------------------------- report

    def report(self, rep, recon, sk_sales, sk_pays, sk_open, seller, dry):
        w = self.stdout.write
        matches, mism, unmatched = recon
        w("")
        w("=" * 60)
        w(f"Mahsulotlar:      {rep['products']}  (tarixiy qo'shimcha: {rep['hist_products']})")
        w(f"Mijozlar:         {rep['clients']}")
        w(f"Sotuvlar (chek):  {rep['sales']}  (qatorlar: {rep['sale_rows']})")
        w(f"Ochilish qarzi:   {rep['open_debt']} ta = {rep['open_debt_total']:,.0f} so'm")
        w(f"Ochilish avansi:  {rep['open_adv']} ta = {rep['open_adv_total']:,.0f} so'm")
        w(f"Qaytarishlar:     {rep['returns']}")
        w(f"To'lovlar:        {rep['payments']}  (ortiqcha->avans: {rep['leftover_advance']})")
        w("")
        if rep.get("remittance"):
            w(f"Ishlab chiqarishga (avval) topshirilgan: {rep['remittance']:,.0f} so'm")
        w(f"ISHLAB CHIQARISH QARZI ({seller.username}): {rep['prod_debt']:,.0f} so'm")
        w("")
        w(f"SVERKA TEKSHIRUVI:  mos {matches} ta | farqli {len(mism)} ta "
          f"| sverkada bor lekin importda yo'q {len(unmatched)} ta")
        for name, comp, exp, why in mism[:25]:
            exp_s = f"{exp:,.0f}" if exp is not None else "—"
            w(f"   ! {name[:32]:32s} hisob={comp:,.0f} sverka={exp_s} ({why})")
        if len(mism) > 25:
            w(f"   … yana {len(mism) - 25} ta")
        if unmatched:
            w(f"   Importda topilmagan (sverkada balans bor): {', '.join(unmatched[:10])}"
              + (" …" if len(unmatched) > 10 else ""))
        if sk_sales:
            w(f"\nO'tkazib yuborilgan sotuv qatorlari: {len(sk_sales)}")
            for ln, why in sk_sales[:10]:
                w(f"   qator {ln}: {why}")
        if sk_pays:
            w(f"O'tkazib yuborilgan to'lov qatorlari: {len(sk_pays)}")
            for ln, why in sk_pays[:10]:
                w(f"   qator {ln}: {why}")
        if rep["skipped_returns"]:
            w(f"O'tkazib yuborilgan qaytarishlar: {len(rep['skipped_returns'])}")
        w("")
        if dry:
            w(self.style.WARNING("DRY-RUN — hech narsa saqlanmadi."))
        else:
            w(self.style.SUCCESS(f"Import yakunlandi → {seller.username}."))
