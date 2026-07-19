"""Import the client's historical Excel ledger (sales + payments) into the CRM.

The workbook holds two side-by-side tables on one sheet:
  - sales (РЕАЛИЗАЦИЯ):  B:САНА C:МИЖОЗ D:ТОВАР E:ЎЛ.БИР F:СОНИ G:НАРХИ
                         H:ЖАМИ I:ОҒИРЛИГИ J:ТАННАРХ K:ЖАМИ2 L:ФОЙДА
  - payments (ПЛАТЕЖИ):  O:САНА P:МИЖОЗ Q:СЎМ R:ВАЛЮТА S:КУРС T:ЖАМИ
                         U:ТЎЛОВ ТУРИ V:ИЗОХ

The ЖАМИ/ФОЙДА columns are formulas linked to an external workbook whose cached
values are stale, so every total here is recomputed from the raw columns
(qty × price, som + usd × rate).

Payments in the file are per-client, not per-sale; they are allocated FIFO
against that client's sales (oldest first). When a client paid more than their
sales in the file (debt predating the ledger), the surplus is booked against a
synthetic "opening balance" sale so the imported debt matches the file exactly.
"""

import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import User
from crm.models import (
    Client,
    Expense,
    Payment,
    Product,
    ProductionReceipt,
    ProductionReceiptItem,
    ProductionRemittance,
    ProfitPayout,
    Return,
    Sale,
    SaleItem,
)

SHEET = "Лист1"
# Product-name typos/variants in the file → one canonical product.
PRODUCT_ALIASES = {
    "ШПУЛ -": "ШПУЛ",
    "ШПУЛ-": "ШПУЛ",
    "ОҚ 2": "ОҚ 2 СОРТ",
}
OPENING_PRODUCT_NAME = "ЭСКИ ҚАРЗ (импорт)"
OPENING_SKU = "OPENING"
# Trailing phone in client names: "БАХТИЁР ЎЗБ МАХАЛЛА 97 420 20 60"
PHONE_RE = re.compile(r"(\d{2})[\s.]*(\d{3})[\s.]*(\d{2})[\s.]*(\d{2})\s*$")

METHOD_BY_NOTE = {"КЛИК": Payment.Method.CARD, "ПЕР": Payment.Method.TRANSFER}


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _money(v):
    return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _qty(v):
    return Decimal(str(v)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)


def _squash(s):
    return re.sub(r"\s+", " ", str(s)).strip()


class Command(BaseCommand):
    help = "Import the historical Excel ledger (sales and payments) into the CRM."

    def add_arguments(self, parser):
        parser.add_argument("xlsx", help="Path to the Excel file")
        parser.add_argument(
            "--owner", default="admin",
            help="Username that will own imported clients/sales/payments",
        )
        parser.add_argument(
            "--wipe", action="store_true",
            help="Delete ALL existing CRM data (clients, products, sales, payments, "
                 "expenses, production records) before importing. Users are kept.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Parse, validate and report, then roll everything back.",
        )
        parser.add_argument(
            "--once", action="store_true",
            help="Skip silently if the import has already been applied "
                 "(an IMP-* product exists). For automatic runs on deploy.",
        )

    # ------------------------------------------------------------------ parse

    def load_rows(self, path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {path}")
        if SHEET not in wb.sheetnames:
            raise CommandError(f"Sheet {SHEET!r} not found in {path}")
        ws = wb[SHEET]

        sales, payments = [], []
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            left, right = row[1:12], row[14:22]
            if any(v is not None for v in left):
                sales.append((i, left))
            if any(v is not None for v in right):
                payments.append((i, right))
        return sales, payments

    def parse_sales(self, raw):
        """Split rows into sales (qty > 0) and returns (qty < 0 — возврат);
        collect skipped rows. Cost price (ТАННАРХ) is empty on most rows, so
        the last known cost per product is carried forward (and backward for
        rows before the first known one)."""
        rows, returns, skipped = [], [], []
        for lineno, r in raw:
            dt, client, product, unit, qty, price, cost = (
                r[0], r[1], r[2], r[3], _num(r[4]), _num(r[5]), _num(r[8])
            )
            if not (isinstance(dt, datetime) and client and product and qty):
                skipped.append((lineno, "sana/mijoz/tovar/soni yo'q", r))
                continue
            name = _squash(product)
            name = PRODUCT_ALIASES.get(name, name)
            row = {
                "line": lineno,
                "date": dt.date(),
                "client": _squash(client),
                "product": name,
                "qty": _qty(abs(qty)),
                "price": _money(price or 0),
                "cost": _money(cost) if cost and cost > 0 else None,
            }
            (rows if qty > 0 else returns).append(row)
        returns.sort(key=lambda x: (x["date"], x["line"]))

        rows.sort(key=lambda x: (x["date"], x["line"]))
        # forward fill of cost per product, then backward fill for the head
        last_cost = {}
        for row in rows:
            if row["cost"] is not None:
                last_cost[row["product"]] = row["cost"]
            else:
                row["cost"] = last_cost.get(row["product"])
        first_cost = {}
        for row in reversed(rows):
            if row["cost"] is not None:
                first_cost[row["product"]] = row["cost"]
            else:
                row["cost"] = first_cost.get(row["product"], Decimal("0"))
        return rows, returns, skipped

    def parse_payments(self, raw):
        rows, skipped = [], []
        for lineno, r in raw:
            dt, client, som, usd, rate, note = (
                r[0], r[1], _num(r[2]), _num(r[3]), _num(r[4]), r[7]
            )
            if not (isinstance(dt, datetime) and client):
                skipped.append((lineno, "sana/mijoz yo'q", r))
                continue
            if not som and not usd:
                skipped.append((lineno, "summa yo'q", r))
                continue
            if usd and not rate:
                skipped.append((lineno, "valyuta bor, kurs yo'q", r))
                continue
            note = _squash(note) if note else ""
            method = METHOD_BY_NOTE.get(note.upper(), Payment.Method.CASH)
            rows.append({
                "line": lineno,
                "date": dt.date(),
                "client": _squash(client),
                "som": _money(som) if som else Decimal("0"),
                "usd": _money(usd) if usd else Decimal("0"),
                "rate": _money(rate) if rate else Decimal("0"),
                "note": note,
                "method": method,
            })
        rows.sort(key=lambda x: (x["date"], x["line"]))
        return rows, skipped

    # ---------------------------------------------------------------- clients

    def build_clients(self, names):
        """raw name → (display name, phone). The trailing phone is moved to the
        phone field unless stripping it would collide two different clients."""
        parsed = {}
        stripped_names = defaultdict(set)
        for raw in names:
            m = PHONE_RE.search(raw)
            if m:
                base = _squash(raw[: m.start()])
                phone = "+998 " + " ".join(m.groups())
                parsed[raw] = (base or raw, phone)
                stripped_names[(base or raw).upper()].add(phone)
            else:
                parsed[raw] = (raw, "")
                stripped_names[raw.upper()].add("")
        # collision: same base name, different phones → keep the full raw name
        for raw, (base, phone) in list(parsed.items()):
            if phone and len(stripped_names[base.upper()]) > 1:
                parsed[raw] = (raw, phone)
        return parsed

    # ------------------------------------------------------------------- run

    def handle(self, *args, **opts):
        if opts["once"] and Product.objects.filter(sku__startswith="IMP-").exists():
            self.stdout.write("Import allaqachon bajarilgan — o'tkazib yuborildi (--once).")
            return

        raw_sales, raw_pays = self.load_rows(opts["xlsx"])
        sales, returns, skipped_sales = self.parse_sales(raw_sales)
        pays, skipped_pays = self.parse_payments(raw_pays)
        if not sales:
            raise CommandError("No valid sales rows found — wrong file/layout?")

        with transaction.atomic():
            owner = self.get_owner(opts["owner"])
            if opts["wipe"]:
                self.wipe()
            report = self.import_all(sales, returns, pays, owner)
            if opts["dry_run"]:
                transaction.set_rollback(True)

        self.report(report, skipped_sales, skipped_pays, opts["dry_run"])

    def get_owner(self, username):
        """The user who will own imported records. Created if absent (unattended
        deploy run) with an unusable password — set one afterwards with
        `manage.py changepassword`."""
        owner, created = User.objects.get_or_create(
            username=username, defaults={"role": User.Role.ADMIN}
        )
        if created:
            owner.set_unusable_password()
            owner.save(update_fields=["password"])
            self.stdout.write(self.style.WARNING(
                f"Foydalanuvchi {username!r} topilmadi — yaratildi (parolsiz). "
                f"Parol o'rnating: manage.py changepassword {username}"
            ))
        return owner

    def wipe(self):
        for model in (
            Payment, Return, SaleItem, Sale, Expense,
            ProductionReceiptItem, ProductionReceipt,
            ProductionRemittance, ProfitPayout, Client, Product,
        ):
            n, _ = model.objects.all().delete()
            if n:
                self.stdout.write(f"  o'chirildi: {model.__name__} x {n}")

    def import_all(self, sales, returns, pays, owner):
        # -- products: catalog price/cost = the latest row that mentions them
        products = {}
        latest = {}
        for row in sales:
            latest[row["product"]] = row  # rows are date-ordered
        for idx, (name, row) in enumerate(sorted(latest.items()), start=1):
            products[name] = Product.objects.create(
                name=name,
                sku=f"IMP-{idx:02d}",
                price=row["price"],
                cost_price=row["cost"],
            )

        # -- clients
        all_names = {r["client"] for r in sales} | {p["client"] for p in pays}
        parsed = self.build_clients(sorted(all_names))
        clients = {}
        for raw, (display, phone) in parsed.items():
            key = display.upper()
            if key not in clients:
                clients[key] = Client.objects.create(
                    name=display, phone=phone, owner=owner
                )
            elif phone and not clients[key].phone:
                clients[key].phone = phone
                clients[key].save(update_fields=["phone"])
        client_for = {raw: clients[parsed[raw][0].upper()] for raw in parsed}

        # -- sales: group rows of one client on one date into a single receipt
        grouped = defaultdict(list)
        for row in sales:
            grouped[(row["client"], row["date"])].append(row)

        sale_objs = defaultdict(list)   # client pk -> [(sale, total, remaining)]
        for (raw_client, d), rows in sorted(grouped.items(), key=lambda kv: kv[0][1]):
            client = client_for[raw_client]
            sale = Sale.objects.create(date=d, client=client, sales_rep=owner)
            total = Decimal("0")
            for row in rows:
                SaleItem.objects.create(
                    sale=sale,
                    product=products[row["product"]],
                    dimension=Sale.Dimension.KG,
                    weight=row["qty"],
                    price=row["price"],
                    cost_price=row["cost"],
                    # history is already delivered — don't show as pending zakaz
                    fulfilled_kg=row["qty"],
                    fulfilled_at=d,
                )
                total += row["qty"] * row["price"]
            sale_objs[client.pk].append({
                "sale": sale,
                "remaining": total,
                "products": {row["product"] for row in rows},
            })

        # -- returns (negative-qty rows): attach each to the client's latest
        #    sale of that product on/before the return date; the returned value
        #    reduces that sale's debt before payments are allocated
        n_returns = 0
        skipped_returns = []
        for row in returns:
            client = client_for[row["client"]]
            slots = sale_objs.get(client.pk, [])
            candidates = [
                s for s in slots
                if s["sale"].date <= row["date"] and row["product"] in s["products"]
            ] or [s for s in slots if row["product"] in s["products"]] or slots
            if not candidates:
                skipped_returns.append((row, "mijozda sotuv yo'q"))
                continue
            slot = candidates[-1]
            Return.objects.create(
                sale=slot["sale"],
                product=products[row["product"]],
                dimension=Sale.Dimension.KG,
                weight=row["qty"],
                price=row["price"],
                date=row["date"],
                restock=True,
                note="Импорт: возврат",
                created_by=owner,
            )
            slot["remaining"] -= row["qty"] * row["price"]
            n_returns += 1

        # -- opening balances: surplus of payments over sales becomes a synthetic
        #    zero-profit sale dated before the ledger starts
        first_date = min(r["date"] for r in sales) - timedelta(days=1)
        pay_totals = defaultdict(Decimal)
        for p in pays:
            pay_totals[client_for[p["client"]].pk] += p["som"] + p["usd"] * p["rate"]
        opening_product = None
        opening_count = 0
        for client_pk, paid in pay_totals.items():
            sold = sum(s["remaining"] for s in sale_objs[client_pk])
            surplus = _money(paid - sold)
            if surplus <= 0:
                continue
            if opening_product is None:
                opening_product = Product.objects.create(
                    name=OPENING_PRODUCT_NAME, sku=OPENING_SKU,
                    price=0, cost_price=0, is_active=False,
                )
            client = Client.objects.get(pk=client_pk)
            sale = Sale.objects.create(
                date=first_date, client=client, sales_rep=owner
            )
            SaleItem.objects.create(
                sale=sale, product=opening_product,
                dimension=Sale.Dimension.KG, weight=1,
                price=surplus, cost_price=surplus,  # zero profit
                fulfilled_kg=1, fulfilled_at=first_date,
            )
            sale_objs[client_pk].insert(0, {"sale": sale, "remaining": surplus})
            opening_count += 1

        # -- payments: FIFO against the client's sales, splitting when a payment
        #    crosses a sale boundary
        n_payments = 0
        unallocated = []
        for p in pays:
            client = client_for[p["client"]]
            slots = sale_objs[client.pk]
            if not slots:
                unallocated.append((p, "mijozda sotuv yo'q"))
                continue
            parts = []
            if p["som"]:
                parts.append(("uzs", p["som"], p["som"]))
            if p["usd"]:
                parts.append(("usd", _money(p["usd"] * p["rate"]), p["usd"]))
            for currency, som_value, original in parts:
                if som_value < 0:
                    # correction rows: book whole against the first sale
                    self.make_payment(p, slots[0]["sale"], currency,
                                      som_value, original, owner)
                    slots[0]["remaining"] -= som_value
                    n_payments += 1
                    continue
                left = som_value
                for slot in slots:
                    if left <= 0:
                        break
                    if slot["remaining"] <= 0:
                        continue
                    chunk = min(left, slot["remaining"])
                    orig = (
                        chunk if currency == "uzs"
                        else _money(chunk / p["rate"])
                    )
                    self.make_payment(p, slot["sale"], currency, chunk, orig, owner)
                    slot["remaining"] -= chunk
                    left -= chunk
                    n_payments += 1
                if left > 0:
                    # shouldn't happen (opening sale absorbs surplus) except for
                    # rounding after negative corrections — book on the last sale
                    orig = left if currency == "uzs" else _money(left / p["rate"])
                    self.make_payment(p, slots[-1]["sale"], currency,
                                      left, orig, owner)
                    n_payments += 1
                    unallocated.append((p, f"oxirgi sotuvga yozildi: {left}"))

        return {
            "products": len(products) + (1 if opening_product else 0),
            "clients": len(clients),
            "sales": Sale.objects.count(),
            "sale_rows": len(sales),
            "openings": opening_count,
            "returns": n_returns,
            "skipped_returns": skipped_returns,
            "payments": n_payments,
            "unallocated": unallocated,
        }

    def make_payment(self, p, sale, currency, som_value, original, owner):
        Payment.objects.create(
            date=p["date"],
            amount=som_value,
            currency=(
                Payment.Currency.USD if currency == "usd" else Payment.Currency.UZS
            ),
            exchange_rate=p["rate"] if currency == "usd" else 0,
            amount_original=original if currency == "usd" else 0,
            method=p["method"],
            note=p["note"],
            kind=(
                Payment.Kind.SALE if p["date"] == sale.date else Payment.Kind.DEBT
            ),
            sale=sale,
            created_by=owner,
        )

    # ---------------------------------------------------------------- report

    def report(self, rep, skipped_sales, skipped_pays, dry_run):
        w = self.stdout.write
        w("")
        w(f"Mahsulotlar: {rep['products']}")
        w(f"Mijozlar:    {rep['clients']}")
        w(f"Sotuvlar:    {rep['sales']} (qatorlar: {rep['sale_rows']}, "
          f"shu jumladan eski qarz uchun: {rep['openings']})")
        w(f"Qaytarishlar: {rep['returns']}")
        w(f"To'lovlar:   {rep['payments']}")
        if rep["skipped_returns"]:
            w(f"\nO'tkazib yuborilgan qaytarishlar: {len(rep['skipped_returns'])}")
            for row, why in rep["skipped_returns"]:
                w(f"  {row['date']} {row['client']}: {why}")
        if skipped_sales:
            w(f"\nO'tkazib yuborilgan sotuv qatorlari: {len(skipped_sales)}")
            for lineno, why, r in skipped_sales:
                w(f"  qator {lineno}: {why} | {r[1]} | {r[2]} | soni={r[4]}")
        if skipped_pays:
            w(f"\nO'tkazib yuborilgan to'lov qatorlari: {len(skipped_pays)}")
            for lineno, why, r in skipped_pays:
                w(f"  qator {lineno}: {why} | {r[1]}")
        if rep["unallocated"]:
            w(f"\nDiqqat — taqsimlashda muammo: {len(rep['unallocated'])}")
            for p, why in rep["unallocated"]:
                w(f"  {p['date']} {p['client']}: {why}")
        w("")
        if dry_run:
            self.stdout.write(self.style.WARNING(
                "DRY RUN — hech narsa saqlanmadi."))
        else:
            self.stdout.write(self.style.SUCCESS("Import yakunlandi."))
