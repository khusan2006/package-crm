"""Import the pre-CRM opening balances from the client's reconciled ОСТ КОЧА sheet.

Each row is a client and their net balance as of go-live:
  * negative  -> the client owes us  -> an opening-balance Sale (is_opening=True,
                 opening_amount=abs). No line items, so it never touches any
                 revenue/profit/sold-kg report — it shows up purely as a receivable
                 and is paid down by ordinary debt payments.
  * positive  -> we hold their money  -> an ADVANCE_IN payment (client advance credit).

Idempotent: a client that already has an opening Sale (or an imported advance) is
skipped, so re-running never doubles a balance. Run with --dry-run first to see the
totals without writing anything.
"""

import datetime
import re
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from accounts.models import User
from crm.models import Client, Payment, Sale

DEFAULT_FILE = r"C:\Users\User\Downloads\Telegram Desktop\ОСТ КОЧА.xlsx"
# The reconciled workbook whose АКТ СВЕРКА sheet holds ОЛДИ (the date each client last
# took goods) — used as the opening debt's date so it ages from when it was incurred.
DEFAULT_SVERKA = r"C:\Users\User\Downloads\Telegram Desktop\ХОЗМАГЛАР 2026 (2) (4).xlsx"
ADVANCE_NOTE = "Ochilish avansi (import)"
# Where a debtor has no usable ОЛДИ date, fall back to the opening-saldo date — these
# are mostly old balances carried over from before June.
FALLBACK_DATE = datetime.date(2026, 6, 1)
DEADLINE_DAYS = 15  # to'lov muddati = olgan sana + 15 kun

# Names to leave out of this run (normalised-substring match). Empty now — the client
# confirmed every balance, including ОТАБЕК ФОМЛАЙН's large advance, should be imported.
SKIP_SUBSTRINGS = []

# A trailing phone number written into the client name, e.g. "ЎТКИР ТЕКСТИЛ 93 300 30 67".
# A run of digits+spaces at the very end, starting and ending on a digit.
_PHONE_RE = re.compile(r"\s(\d[\d\s]*\d)\s*$")


def _norm(s):
    return " ".join(str(s).split()) if s is not None else ""


def format_phone(raw):
    """Canonical phone the app stores: "+998 " plus the 9 national digits grouped
    XX XXX XX XX (matching the data-phone widget). Leading 998 country code is dropped.
    Returns "" when there aren't 9 usable digits."""
    digits = re.sub(r"\D", "", str(raw))
    # Drop the 998 country code ONLY when it's actually a prefix (more than 9 digits) —
    # a bare national number can itself start with 99…8, which must not be stripped.
    if len(digits) > 9 and digits.startswith("998"):
        digits = digits[3:]
    digits = digits[:9]
    if len(digits) != 9:
        return ""
    return f"+998 {digits[0:2]} {digits[2:5]} {digits[5:7]} {digits[7:9]}"


def split_name_phone(full):
    """Pull a trailing phone number out of a client name, so the name stays clean and
    the number lands in its own field in the app's canonical "+998 …" format. Uzbek
    mobiles are 9 digits; we only split when the trailing run has 7-12 digits, so a
    name that merely ends in a small number is left alone. Returns (name, phone) —
    phone is "" when there is none / not 9 digits."""
    full = _norm(full)
    m = _PHONE_RE.search(full)
    if m:
        raw = m.group(1)
        digits = re.sub(r"\D", "", raw)
        name = full[: m.start()].strip()
        if 7 <= len(digits) <= 12 and name:
            return name, format_phone(raw)
    return full, ""


def build_oldi_map(sverka_file, today):
    """{normalised full client name -> the date they last took goods (ОЛДИ)} from the
    АКТ СВЕРКА sheet. A date is only kept when it's plausible — a real datetime, this
    century, and not in the future (the sheet has stray 1900 cells and a couple of
    Oct/Nov typos). Missing/implausible ones are left out so the caller can fall back."""
    import openpyxl

    wb = openpyxl.load_workbook(sverka_file, read_only=True, data_only=True)
    ws = wb["АКТ СВЕРКА"]
    result = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = _norm(r[0])
        if not name:
            continue
        oldi = r[5] if len(r) > 5 else None  # ОЛДИ column
        if isinstance(oldi, datetime.datetime):
            d = oldi.date()
            if d.year >= 2020 and d <= today:
                result[name.upper()] = d
    return result


class Command(BaseCommand):
    help = "ОСТ КОЧА faylidan eski qarz/avans qoldiqlarini import qiladi."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_FILE, help="ОСТ КОЧА .xlsx yo'li")
        parser.add_argument("--sverka-file", default=DEFAULT_SVERKA,
                            help="АКТ СВЕРКА (ОЛДИ sanalari) bor .xlsx yo'li")
        parser.add_argument("--seller", default="sotuvchi1",
                            help="Ega/sotuvchi username (default: sotuvchi1)")
        parser.add_argument("--dry-run", action="store_true",
                            help="Hech narsa yozmaydi, faqat xulosani ko'rsatadi.")

    def handle(self, *args, **opt):
        import openpyxl

        try:
            seller = User.objects.get(username=opt["seller"])
        except User.DoesNotExist:
            raise CommandError(f"Sotuvchi topilmadi: {opt['seller']}")

        today = timezone.localdate()
        # Per-client "took" dates; a debtor with none falls back to FALLBACK_DATE.
        oldi_map = build_oldi_map(opt["sverka_file"], today)

        wb = openpyxl.load_workbook(opt["file"], read_only=True, data_only=True)
        ws = wb.active

        rows = []
        for r in ws.iter_rows(values_only=True):
            name = _norm(r[0])
            val = r[1] if len(r) > 1 else None
            if not name or not isinstance(val, (int, float)):
                continue
            rows.append((name, Decimal(str(val))))

        dry = opt["dry_run"]
        created_debt = created_adv = skipped_existing = skipped_named = 0
        dated_from_oldi = dated_from_fallback = 0
        debt_total = adv_total = Decimal("0")

        # Wrap the whole run so a mid-way failure leaves the DB untouched.
        with transaction.atomic():
            for name, val in rows:
                if any(s in name.upper() for s in SKIP_SUBSTRINGS):
                    skipped_named += 1
                    self.stdout.write(f"  ⏭  chetda qoldirildi: {name}")
                    continue

                clean_name, phone = split_name_phone(name)
                client = Client.find_duplicate(seller, clean_name) or Client.objects.filter(
                    name__iexact=clean_name
                ).first()
                if client is None and not dry:
                    client = Client.objects.create(
                        name=clean_name, phone=phone, owner=seller
                    )

                if val < 0:  # client owes us -> opening debt
                    amount = -val
                    exists = client is not None and Sale.objects.filter(
                        client=client, is_opening=True
                    ).exists()
                    if exists:
                        skipped_existing += 1
                        continue
                    created_debt += 1
                    debt_total += amount
                    oldi = oldi_map.get(name.upper())
                    if oldi:
                        dated_from_oldi += 1
                    else:
                        dated_from_fallback += 1
                    debt_date = oldi or FALLBACK_DATE
                    deadline = debt_date + datetime.timedelta(days=DEADLINE_DAYS)
                    if not dry:
                        Sale.objects.create(
                            client=client, sales_rep=seller, date=debt_date,
                            debt_deadline=deadline, is_opening=True,
                            opening_amount=amount,
                        )
                elif val > 0:  # we hold their money -> advance credit
                    exists = client is not None and Payment.objects.filter(
                        client=client, kind=Payment.Kind.ADVANCE_IN, note=ADVANCE_NOTE
                    ).exists()
                    if exists:
                        skipped_existing += 1
                        continue
                    created_adv += 1
                    adv_total += val
                    if not dry:
                        Payment.objects.create(
                            client=client, created_by=seller, date=today,
                            amount=val, kind=Payment.Kind.ADVANCE_IN, note=ADVANCE_NOTE,
                            is_opening=True,
                        )

            if dry:
                transaction.set_rollback(True)

        head = "DRY-RUN (hech narsa yozilmadi)" if dry else "IMPORT bajarildi"
        self.stdout.write(self.style.SUCCESS(f"\n=== {head} ==="))
        self.stdout.write(f"Sotuvchi: {seller.username} | Muddat = sana + {DEADLINE_DAYS} kun")
        self.stdout.write(f"Qarz (opening) yozuvlari: {created_debt} ta = {debt_total:,.0f} so'm")
        self.stdout.write(f"  sana ОЛДИ'dan: {dated_from_oldi} ta | fallback ({FALLBACK_DATE}): {dated_from_fallback} ta")
        self.stdout.write(f"Avans yozuvlari:          {created_adv} ta = {adv_total:,.0f} so'm")
        self.stdout.write(f"Chetda qoldirilgan (nom): {skipped_named} ta")
        self.stdout.write(f"Allaqachon bor (o'tkazib): {skipped_existing} ta")
