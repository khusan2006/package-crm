import calendar
import math
import re
from datetime import date, timedelta
from decimal import ROUND_HALF_UP, Decimal

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    BooleanField,
    Case,
    Count,
    F,
    Max,
    ProtectedError,
    Q,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, Replace, TruncMonth
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from accounts.decorators import role_required
from accounts.models import User

from .forms import (
    DEFAULT_DEBT_DAYS,
    AdvanceEditForm,
    AdvanceForm,
    AdvanceRemoveForm,
    ClientForm,
    ClientTransferForm,
    DebtPaymentForm,
    EmployeeForm,
    ExpenseForm,
    OpeningDebtForm,
    PaymentEditForm,
    ProductForm,
    ProductionReceiptForm,
    ProductionReceiptItemFormSet,
    ProductionRefundForm,
    ProductionAdjustForm,
    ProductionRemittanceForm,
    ProfitPayoutForm,
    ReturnForm,
    SaleForm,
    SaleItemFormSet,
    StockAdjustForm,
    StockEntryForm,
)
from .models import (
    ADVANCE_ADJUST_NOTE,
    ADVANCE_DEPOSIT_KINDS,
    ADVANCE_SPENT_KINDS,
    COST,
    ITEM_WEIGHT_KG,
    MICRON_CHOICES,
    PAYING_KINDS,
    PAYMENT_CREDIT,
    PAYMENT_NET,
    PROFIT,
    REFUND_KINDS,
    RETURN_AMOUNT,
    RETURN_COST,
    REVENUE,
    SIZE_CHOICES,
    AuditLog,
    Client,
    Employee,
    Expense,
    Payment,
    client_advance_balance,
    Product,
    ProductionReceipt,
    ProductionAdjustment,
    ProductionRemittance,
    ProfitPayout,
    Return,
    Sale,
    SaleItem,
    SalaryRate,
    StockEntry,
    month_span,
    seller_cash_on_hand,
    seller_production_debt,
)
from .utils import (
    form_changes,
    form_reload,
    form_response,
    form_success,
    is_ajax,
    render_confirm,
    uz_month,
)


def _visible_clients(user):
    qs = Client.objects.select_related("owner")
    return qs if user.can_see_all_records else qs.filter(owner=user)


def _advance_balance_map(client_pks, seller):
    """{client_pk: advance balance} for the given clients, in ONE query. Mirrors
    `client_advance_balance` (deposits minus consumption, gross of bank fees) but
    batched, so a whole list can be sorted/annotated without an N+1. Seller-scoped when
    `seller` is set; otherwise the client's total across every seller's till."""
    rows = Payment.objects.filter(client__in=client_pks)
    if seller is not None:
        rows = rows.filter(created_by=seller)
    agg = rows.values("client").annotate(
        dep=Sum(PAYMENT_CREDIT, filter=Q(kind__in=ADVANCE_DEPOSIT_KINDS)),
        used=Sum(PAYMENT_CREDIT, filter=Q(kind__in=ADVANCE_SPENT_KINDS)),
    )
    return {
        r["client"]: (r["dep"] or Decimal("0")) - (r["used"] or Decimal("0"))
        for r in agg
    }


def _advance_since_map(client_pks, seller):
    """{client_pk: the date the money now sitting in their advance came in} — the
    oldest deposit that has not yet been eaten. Walked in date order with a running
    balance, so a client who spent an old deposit down to nothing and then paid again
    is dated from the new money, not from the deposit that is long gone. Same
    seller-scoping as `_advance_balance_map`, so the date belongs to the balance
    beside it."""
    rows = Payment.objects.filter(
        client__in=client_pks, kind__in=ADVANCE_DEPOSIT_KINDS + ADVANCE_SPENT_KINDS
    )
    if seller is not None:
        rows = rows.filter(created_by=seller)
    moves = (
        rows.annotate(credit=PAYMENT_CREDIT)
        .order_by("client", "date", "pk")
        .values_list("client", "date", "kind", "credit")
    )
    since, balances = {}, {}
    for client_pk, date, kind, credit in moves:
        balance = balances.get(client_pk, Decimal("0"))
        if kind in ADVANCE_DEPOSIT_KINDS:
            if balance <= 0:
                since[client_pk] = date
            balance += credit
        else:
            balance -= credit
            if balance <= 0:
                since.pop(client_pk, None)
        balances[client_pk] = balance
    return since


def _sale_totals(sales):
    """Revenue/cost/profit summed over the line items of the given sales."""
    return SaleItem.objects.filter(sale__in=sales.values("pk")).aggregate(
        revenue=Sum(REVENUE), cost=Sum(COST), profit=Sum(PROFIT)
    )


def _warn_if_negative_stock(request, product):
    """Sales are allowed even without stock, but flag it so it's visible."""
    stock = product.current_stock
    if stock < 0:
        messages.warning(
            request,
            f"Diqqat: “{product.name}” ombori yetarli emas — qoldiq {stock:.3f} kg.",
        )


def _warn_if_negative_stock_items(request, sale):
    """Flag every distinct product on the sale whose stock went negative."""
    seen = set()
    for item in sale.items.select_related("product"):
        if item.product_id not in seen:
            seen.add(item.product_id)
            _warn_if_negative_stock(request, item.product)


def _ombor_shortfall(seller, formset, existing_sale=None):
    """Products on the sale whose requested kg exceed the seller's own ombor.

    Returns a list of (product, requested_kg, available_kg). On edit, this sale's
    current lines are added back to availability — they're already counted as sold
    against the seller, so editing must not double-count them."""
    requested = {}  # product_pk -> {"product": Product, "kg": Decimal}
    for f in formset.forms:
        cd = getattr(f, "cleaned_data", None)
        if not cd or cd.get("DELETE"):
            continue
        product, weight = cd.get("product"), cd.get("weight")
        if not product or weight is None:
            continue
        kg = weight / Decimal("1000") if cd.get("dimension") == Sale.Dimension.G else weight
        row = requested.setdefault(product.pk, {"product": product, "kg": Decimal("0")})
        row["kg"] += kg
    if not requested:
        return []
    on_hand = {
        p.pk: p.stock
        for p in Product.objects.filter(pk__in=requested).with_stock(seller=seller)
    }
    freed = {}
    if existing_sale is not None:
        for item in existing_sale.items.all():
            freed[item.product_id] = freed.get(item.product_id, Decimal("0")) + item.weight_kg
    shortfalls = []
    for pk, row in requested.items():
        available = (on_hand.get(pk) or Decimal("0")) + freed.get(pk, Decimal("0"))
        if row["kg"] > available:
            shortfalls.append((row["product"], row["kg"], available))
    return shortfalls


def _mark_fulfilment(sale, shortfall, only_unset=False):
    """Set each line's fulfilment after a sale saves. In-stock lines are ready on
    the sale date; a line whose product was short is a pending zakaz (fulfilled_at
    stays NULL until stock is bound to it). On edit (`only_unset`) already-fulfilled
    lines are left untouched."""
    short_pks = {p.pk for p, _req, _avail in shortfall}
    items = sale.items.all()
    if only_unset:
        items = items.filter(fulfilled_at__isnull=True)
    for item in items:
        if item.product_id in short_pks:
            item.fulfilled_kg = Decimal("0")
            item.fulfilled_at = None
        else:
            item.fulfilled_kg = item.weight_kg
            item.fulfilled_at = sale.date
        item.save(update_fields=["fulfilled_kg", "fulfilled_at"])


def _zakaz_confirm_response(request, form, formset, title, shortfall):
    """A modal oversell asks the browser to pop a confirm dialog (the X-Zakaz-Confirm
    signal) rather than rejecting. Without JS, fall back to the inline warning
    re-render so the flow still works."""
    if is_ajax(request):
        msg = "; ".join(
            f"{p.name}: qoldiq {available:.0f} kg, so'raldi {requested:.0f} kg"
            for p, requested, available in shortfall
        )
        resp = JsonResponse({"message": msg}, status=409)
        resp["X-Zakaz-Confirm"] = "1"
        return resp
    return _render_sale_form(request, form, formset, title, zakaz_shortfall=shortfall)


def _parse_date(value):
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError):
        return None


def _parse_amount(value):
    """A so'm figure out of a URL, or None. Used only to PREFILL a form field, never
    to move money — whatever comes back still goes through the form."""
    try:
        return Decimal(str(value))
    except (TypeError, ValueError, ArithmeticError):
        return None


# --- Dashboard ---------------------------------------------------------------

UZ_MONTHS_SHORT = ["Yan", "Fev", "Mar", "Apr", "May", "Iyn", "Iyl", "Avg", "Sen", "Okt", "Noy", "Dek"]


def _monthly_series(sales, months=6):
    """Revenue / profit for the last `months` months (oldest first) as SVG line
    points scaled to a fixed viewBox, ready for a line chart."""
    today = timezone.localdate()
    buckets = []
    y, m = today.year, today.month
    for _ in range(months):
        buckets.append((y, m))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    buckets.reverse()
    start = date(buckets[0][0], buckets[0][1], 1)
    rows = (
        SaleItem.objects.filter(sale__in=sales, sale__date__gte=start)
        .annotate(mon=TruncMonth("sale__date"))
        .values("mon")
        .annotate(revenue=Sum(REVENUE), cost=Sum(COST), profit=Sum(PROFIT))
    )
    by_month = {(r["mon"].year, r["mon"].month): r for r in rows}
    data = []
    for yy, mm in buckets:
        row = by_month.get((yy, mm)) or {}
        last_day = (date(yy + (mm == 12), (mm % 12) + 1, 1) - timedelta(days=1)).day
        data.append({
            "label": UZ_MONTHS_SHORT[mm - 1],
            "revenue": row.get("revenue") or Decimal("0"),
            "cost": row.get("cost") or Decimal("0"),
            "profit": row.get("profit") or Decimal("0"),
            "dan": date(yy, mm, 1).isoformat(),
            "gacha": date(yy, mm, last_day).isoformat(),
        })

    # --- line-chart geometry (fixed viewBox, scaled to the tallest revenue) ---
    vb_w, vb_h = 720.0, 240.0
    pad_l, pad_r, pad_t, pad_b = 16.0, 16.0, 18.0, 30.0
    inner_w = vb_w - pad_l - pad_r
    inner_h = vb_h - pad_t - pad_b
    baseline = pad_t + inner_h
    n = len(data)
    peak = max((d["revenue"] for d in data), default=Decimal("0")) or Decimal("1")

    def _y(value):
        return round(baseline - float(value / peak) * inner_h, 2)

    band = inner_w / (n - 1) if n > 1 else inner_w
    for i, d in enumerate(data):
        d["x"] = round(pad_l + (inner_w * i / (n - 1) if n > 1 else inner_w / 2), 2)
        d["y_rev"] = _y(d["revenue"])
        d["y_profit"] = _y(d["profit"])
        # Full-height transparent hit band so a click anywhere in the month's
        # column drills into it — not just on the tiny label/point.
        left = max(d["x"] - band / 2, 0.0)
        right = min(d["x"] + band / 2, vb_w)
        d["hit_x"] = round(left, 2)
        d["hit_w"] = round(right - left, 2)

    rev_line = " ".join(f"{d['x']},{d['y_rev']}" for d in data)
    profit_line = " ".join(f"{d['x']},{d['y_profit']}" for d in data)
    first_x = data[0]["x"] if data else pad_l
    last_x = data[-1]["x"] if data else vb_w - pad_r
    rev_area = f"{rev_line} {last_x},{baseline} {first_x},{baseline}"

    return {
        "rows": data,
        "rev_line": rev_line,
        "profit_line": profit_line,
        "rev_area": rev_area,
        "vb_h": vb_h,
        "viewbox": f"0 0 {vb_w:g} {vb_h:g}",
    }


def _spark_points(values, width=118.0, height=30.0, pad=3.0):
    """Scale a numeric series into an SVG polyline points string for a KPI
    sparkline. A flat or empty series renders as a centred flat line."""
    nums = [float(v or 0) for v in values]
    n = len(nums)
    if n == 0:
        return ""
    lo, hi = min(nums), max(nums)
    span = hi - lo
    inner_w = width - 2 * pad
    inner_h = height - 2 * pad
    pts = []
    for i, v in enumerate(nums):
        x = pad + (inner_w * i / (n - 1) if n > 1 else inner_w / 2)
        frac = (v - lo) / span if span else 0.5   # higher value → higher on screen
        y = pad + (1 - frac) * inner_h
        pts.append(f"{round(x, 1)},{round(y, 1)}")
    return " ".join(pts)


def _kpi_sparklines(flow, scoped, clients_q, months=6):
    """Six-month trend for each dashboard KPI as sparkline point strings. Uses the
    same rep/client/method scoping as the rest of the dashboard (date window aside,
    since a sparkline shows the longer trend, not just the selected period)."""
    today = timezone.localdate()
    buckets = []
    y, m = today.year, today.month
    for _ in range(months):
        buckets.append((y, m))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    buckets.reverse()
    start = date(buckets[0][0], buckets[0][1], 1)

    item_rows = (
        SaleItem.objects.filter(sale__in=flow, sale__date__gte=start)
        .annotate(mon=TruncMonth("sale__date"))
        .values("mon")
        .annotate(revenue=Sum(REVENUE), profit=Sum(PROFIT))
    )
    rev_by = {(r["mon"].year, r["mon"].month): r["revenue"] or 0 for r in item_rows}
    prof_by = {(r["mon"].year, r["mon"].month): r["profit"] or 0 for r in item_rows}

    cnt_rows = (
        flow.filter(date__gte=start).annotate(mon=TruncMonth("date"))
        .values("mon").annotate(c=Count("pk"))
    )
    cnt_by = {(r["mon"].year, r["mon"].month): r["c"] for r in cnt_rows}

    cli_rows = (
        clients_q.filter(created_at__date__gte=start)
        .annotate(mon=TruncMonth("created_at"))
        .values("mon").annotate(c=Count("pk"))
    )
    cli_by = {(r["mon"].year, r["mon"].month): r["c"] for r in cli_rows}

    debt_rows = (
        scoped.outstanding().filter(date__gte=start)
        .annotate(mon=TruncMonth("date"))
        .values("mon").annotate(c=Count("pk"))
    )
    debt_by = {(r["mon"].year, r["mon"].month): r["c"] for r in debt_rows}

    revenue, profit, avg, clients, debt = [], [], [], [], []
    for yy, mm in buckets:
        rev = float(rev_by.get((yy, mm), 0))
        cnt = cnt_by.get((yy, mm), 0)
        revenue.append(rev)
        profit.append(float(prof_by.get((yy, mm), 0)))
        avg.append(rev / cnt if cnt else 0)
        clients.append(cli_by.get((yy, mm), 0))
        debt.append(debt_by.get((yy, mm), 0))

    return {
        "revenue": _spark_points(revenue),
        "profit": _spark_points(profit),
        "avg": _spark_points(avg),
        "clients": _spark_points(clients),
        "debt": _spark_points(debt),
    }


def _short_money(value):
    """Compact so'm label (e.g. 44.4 mln) for tight spaces like a donut centre."""
    v = float(value or 0)
    if v >= 1e9:
        return f"{v / 1e9:.1f} mlrd"
    if v >= 1e6:
        return f"{v / 1e6:.1f} mln"
    if v >= 1e3:
        return f"{v / 1e3:.0f} ming"
    return f"{v:.0f}"


def _donut(items):
    """Build donut-ready arc segments from (key, label, amount, color) tuples: each
    segment carries the stroke-dasharray/offset that draws its slice of the ring.
    `key` is an optional cross-filter handle (e.g. a payment method) — may be None."""
    grand = sum((amount for _, _, amount, _ in items), Decimal("0"))
    radius = 56.0
    circumference = 2 * math.pi * radius
    segments = []
    cursor = 0.0
    for key, label, amount, color in items:
        frac = float(amount / grand) if grand else 0.0
        length = frac * circumference
        segments.append({
            "key": key,
            "label": label,
            "total": amount,
            "color": color,
            "pct": round(frac * 100, 1),
            "dash": round(length, 2),
            "gap": round(circumference - length, 2),
            "offset": round(-cursor, 2),
        })
        cursor += length
    return {
        "segments": segments,
        "grand": grand,
        "grand_short": _short_money(grand),
        "radius": radius,
    }


def _payment_donut(sales):
    """Payment totals split by method, as donut-ready arc segments."""
    rows = Payment.objects.filter(sale__in=sales).values("method").annotate(total=Sum("amount"))
    totals = {r["method"]: r["total"] or Decimal("0") for r in rows}
    palette = [
        ("cash", "Naqd", "var(--accent)"),
        ("card", "Karta", "var(--success)"),
        ("transfer", "Bank o'tkazmasi", "var(--warning)"),
    ]
    return _donut([(key, label, totals.get(key, Decimal("0")), color) for key, label, color in palette])


def _debt_overview(sales, aging_filter=None, top=5):
    """Outstanding receivables split into aging buckets (by days overdue) plus the
    biggest debtors — the dashboard's debt-at-a-glance. A company that sells heavily
    on credit needs to see which money is at risk and who to collect from first.

    `aging_filter` (a bucket key) cross-filters the top-debtors list to that bucket
    while the donut keeps showing the full split, so another bucket stays one click away."""
    today = timezone.localdate()
    open_sales = sales.outstanding().select_related("client")

    # Risk gradient: neutral (not yet due) → gold → orange → red (deeply overdue).
    aging_defs = [
        ("current", "Muddati kelmagan", "color-mix(in srgb, var(--accent) 45%, var(--surface))"),
        ("d1_7", "1–7 kun kechikkan", "var(--warning)"),
        ("d8_30", "8–30 kun kechikkan", "color-mix(in srgb, var(--warning) 45%, var(--danger))"),
        ("d30", "30+ kun kechikkan", "var(--danger)"),
    ]
    aging = [{"key": key, "label": lbl, "color": clr, "amount": Decimal("0"), "count": 0}
             for key, lbl, clr in aging_defs]

    clients = {}
    total = Decimal("0")
    overdue_total = Decimal("0")
    for sale in open_sales:
        rem = sale.remaining or Decimal("0")
        if rem <= 0:
            continue
        total += rem
        deadline = sale.debt_deadline
        if deadline is None or deadline >= today:
            idx = 0
        else:
            overdue_days = (today - deadline).days
            overdue_total += rem
            idx = 1 if overdue_days <= 7 else 2 if overdue_days <= 30 else 3
        aging[idx]["amount"] += rem
        aging[idx]["count"] += 1

        # The donut sees every receipt; the debtor list only the selected bucket.
        if aging_filter and aging[idx]["key"] != aging_filter:
            continue
        debtor = clients.get(sale.client_id)
        if debtor is None:
            debtor = clients[sale.client_id] = {
                "client_id": sale.client_id, "name": sale.client.name,
                "amount": Decimal("0"), "overdue": False,
            }
        debtor["amount"] += rem
        if deadline is not None and deadline < today:
            debtor["overdue"] = True

    grand = total or Decimal("1")
    for bucket in aging:
        bucket["pct"] = round(float(bucket["amount"] / grand) * 100, 2)

    # Draw the aging split as a donut (always the full picture), then fold each
    # bucket's receipt count back onto its arc segment for the legend.
    aging_donut = _donut([(b["key"], b["label"], b["amount"], b["color"]) for b in aging])
    for segment, bucket in zip(aging_donut["segments"], aging):
        segment["count"] = bucket["count"]
        segment["active"] = aging_filter == bucket["key"]

    top_debtors = sorted(clients.values(), key=lambda d: d["amount"], reverse=True)[:top]
    peak = max((d["amount"] for d in top_debtors), default=Decimal("1")) or Decimal("1")
    for debtor in top_debtors:
        debtor["pct"] = round(float(debtor["amount"] / peak) * 100, 2)

    return {
        "aging": aging,
        "aging_donut": aging_donut,
        "top_debtors": top_debtors,
        "selected": aging_filter,
        "total": total,
        "overdue_total": overdue_total,
        "overdue_pct": round(float(overdue_total / grand) * 100, 1),
    }



def _top_clients(sales, limit=5):
    """Top clients by revenue within the given (already-scoped) sales."""
    rows = list(
        SaleItem.objects.filter(sale__in=sales.values("pk"))
        .values("sale__client_id", "sale__client__name")
        .annotate(total=Sum(REVENUE))
        .order_by("-total")[:limit]
    )
    peak = max((r["total"] or Decimal("0") for r in rows), default=Decimal("1")) or Decimal("1")
    for row in rows:
        row["name"] = row["sale__client__name"]
        row["client_id"] = row["sale__client_id"]
        row["pct"] = round(float((row["total"] or Decimal("0")) / peak * 100), 2)
    return rows


def dashboard(request):
    today = timezone.localdate()
    # The period drives every "flow" figure; it defaults to month-to-date.
    date_from = _parse_date(request.GET.get("dan")) or today.replace(day=1)
    date_to = _parse_date(request.GET.get("gacha")) or today
    if date_to < date_from:
        date_from, date_to = date_to, date_from

    # Cross-filters set by clicking a chart element (or the drawer). Rep is
    # admin/manager-only; rep+client scope the whole dashboard, the date window and
    # payment method only the sales "flows" (debt is a method-independent snapshot).
    filters = {key: request.GET.get(key, "") for key in ("rep", "client", "method", "aging")}
    rep_id = filters["rep"] if (filters["rep"].isdigit() and request.user.can_see_all_records) else ""
    client_id = filters["client"] if filters["client"].isdigit() else ""
    method = filters["method"] if filters["method"] in Payment.Method.values else ""
    aging = filters["aging"] if filters["aging"] in ("current", "d1_7", "d8_30", "d30") else ""

    scoped = Sale.objects.visible_to(request.user)
    if rep_id:
        scoped = scoped.filter(sales_rep_id=rep_id)
    if client_id:
        scoped = scoped.filter(client_id=client_id)
    # `flow` narrows the sales set by payment method for the flow figures, but debt
    # keeps using `scoped` — an unpaid receipt has no payment row of any method.
    flow = scoped.filter(payments__method=method).distinct() if method else scoped
    # Opening-balance carry-overs are receivables, not sales — keep them out of the
    # period revenue/count/recent figures (debt below still uses `scoped`, with them in).
    period = flow.filter(date__gte=date_from, date__lte=date_to).real()

    def _margin(t):
        rev = t["revenue"] or 0
        return (t["profit"] or 0) / rev * 100 if rev else 0

    period_totals = _sale_totals(period)
    period_count = period.count()
    period_revenue = period_totals["revenue"] or 0
    avg_check = period_revenue / period_count if period_count else 0

    # Debt is a live snapshot: rep/client scoped, but never date-scoped — an old
    # receipt is still owed today regardless of the selected window.
    open_sales = scoped.outstanding()
    debt_total = _outstanding_balance(open_sales)
    overdue_sales = open_sales.filter(debt_deadline__lt=today)
    overdue_count = overdue_sales.count()
    overdue_total = _outstanding_balance(overdue_sales)
    overdue_clients = overdue_sales.values("client").distinct().count()

    # New clients acquired in the period (owner-scoped when a rep is selected).
    new_clients_q = _visible_clients(request.user)
    if rep_id:
        new_clients_q = new_clients_q.filter(owner_id=rep_id)
    new_clients = new_clients_q.filter(
        created_at__date__gte=date_from, created_at__date__lte=date_to
    ).count()

    recent_sales = (
        period.select_related("client", "sales_rep")
        .prefetch_related("items__product")
        .with_totals()
        .order_by("-date", "-created_at")[:8]
    )

    # Shared toolbar / drawer plumbing (rep + client chips, the date-range picker).
    clients = _visible_clients(request.user).order_by("name")
    reps = (
        User.objects.filter(is_active=True).order_by("first_name", "username")
        if request.user.can_see_all_records
        else None
    )
    client_obj = clients.filter(pk=client_id).first() if client_id else None
    rep_obj = reps.filter(pk=rep_id).first() if reps and rep_id else None
    method_labels = dict(Payment.Method.choices)
    aging_labels = {
        "current": "Muddati kelmagan", "d1_7": "1–7 kun kechikkan",
        "d8_30": "8–30 kun kechikkan", "d30": "30+ kun kechikkan",
    }
    filters["dan"] = date_from.isoformat()
    filters["gacha"] = date_to.isoformat()
    filters["q"] = ""
    active_filters = _filter_chips(request, [
        {"param": "rep", "label": "Sotuvchi", "value": str(rep_obj) if rep_obj else ""},
        {"param": "client", "label": "Mijoz", "value": client_obj.name if client_obj else ""},
        {"param": "method", "label": "To'lov usuli", "value": method_labels.get(method, "")},
        {"param": "aging", "label": "Qarz muddati", "value": aging_labels.get(aging, "")},
    ])

    context = {
        "monthly": _monthly_series(flow),
        "sparks": _kpi_sparklines(flow, scoped, new_clients_q),
        "donut": _payment_donut(period),
        "debt": _debt_overview(scoped, aging_filter=aging),
        "top_clients": _top_clients(period),
        "recent_sales": recent_sales,
        "period_revenue": period_revenue,
        "period_profit": period_totals["profit"] or 0,
        "period_count": period_count,
        "period_margin": _margin(period_totals),
        "avg_check": avg_check,
        "new_clients": new_clients,
        "debt_total": debt_total,
        "overdue_count": overdue_count,
        "overdue_total": overdue_total,
        "overdue_clients": overdue_clients,
        "filters": filters,
        "reps": reps,
        "clients": clients,
        "active_filters": active_filters,
        "has_filters": bool(active_filters),
        "filter_count": len(active_filters),
        "filter_url": reverse("dashboard"),
        "date_from": date_from,
        "date_to": date_to,
        "range_days": (date_to - date_from).days + 1,
        "is_single_day": date_from == date_to,
        "is_today": date_from == today and date_to == today,
        "prev_from": (date_from - timedelta(days=1)).isoformat(),
        "prev_to": (date_to - timedelta(days=1)).isoformat(),
        "next_from": (date_from + timedelta(days=1)).isoformat(),
        "next_to": (date_to + timedelta(days=1)).isoformat(),
        "today_iso": today.isoformat(),
    }
    return render(request, "crm/dashboard.html", context)


# --- Clients ------------------------------------------------------------------

def _client_rows(request):
    """The Mijozlar rows for the current search: every visible client with their
    advance and open debt attached. Shared by the page and its Excel export."""
    clients = (
        _visible_clients(request.user)
        # Real sales only: an opening carry-over is a balance moved into the CRM, not
        # a receipt written here, so it is not one of the client's sales. Its date is
        # a different matter — see `_last_sale_map`.
        .annotate(sale_count=Count("sales", filter=Q(sales__is_opening=False)))
        .order_by("name")
    )
    q = request.GET.get("q", "").strip()
    if q:
        # Broad match: name, company, phone (formatting ignored), location, notes,
        # or responsible employee — so "sergeli" finds every client in that district
        # and "901234567" finds the one whose number is stored as "+998 90 123 45 67".
        clients = _client_search(clients, q, extra=(
            Q(notes__icontains=q)
            | Q(owner__first_name__icontains=q)
            | Q(owner__last_name__icontains=q)
            | Q(owner__username__icontains=q)
        )).distinct()
    # Advance (prepaid) balance per client, batched in one query. A seller sees their own
    # till's balance; an admin/manager sees the client's total across every seller's till.
    scope = None if request.user.can_see_all_records else request.user
    rows = list(clients)
    pks = [c.pk for c in rows]
    adv_map = _advance_balance_map(pks, scope)
    debt_map = _client_debt_map(request.user)
    sale_map = _last_sale_map(request.user, pks)
    pay_map = _last_payment_map(request.user, pks)
    # A deposit fully spent on receipts leaves a zero balance but is still a row that
    # may need fixing, so the Avans cell has to stay clickable for those clients too —
    # a mistyped figure is not unreachable just because sales already ate it.
    with_moves = set(
        _advance_move_qs(request.user)
        .filter(client__in=pks)
        .values_list("client", flat=True)
        .distinct()
    )
    for c in rows:
        c.advance = adv_map.get(c.pk, Decimal("0"))
        c.has_advance_moves = c.pk in with_moves
        c.debt = debt_map.get(c.pk, Decimal("0"))
        c.last_sale = sale_map.get(c.pk)
        c.last_payment = pay_map.get(c.pk)
    # Clients holding an advance float to the top (biggest first), then everyone by name.
    rows.sort(key=lambda c: (0 if c.advance > 0 else 1, -c.advance, c.name.lower()))
    return rows, q


def _client_debt_map(user):
    """{client_pk: open debt} across the sales the user may see, in one query."""
    agg = (
        Sale.objects.visible_to(user).outstanding()
        .values("client").annotate(owed=Sum("remaining"))
    )
    return {r["client"]: r["owed"] or Decimal("0") for r in agg}


def _last_sale_map(user, client_pks):
    """{client_pk: date they last took goods}, in one query.

    Opening carry-overs count here, unlike everywhere else: `redate_opening_debts`
    dates each one on the client's ОЛДИ — the day they last took goods before
    go-live — so for a client whose trading predates the CRM that receipt is the
    answer, and most of today's debtors have no other sale. It costs nothing for the
    rest: an opening balance is always older than their real sales, so Max() still
    lands on the last real shipment."""
    qs = Sale.objects.visible_to(user).filter(client__in=client_pks)
    return {
        r["client"]: r["last"] for r in qs.values("client").annotate(last=Max("date"))
    }


def _last_payment_map(user, client_pks):
    """{client_pk: date they last handed money over}, in two queries.

    Only money coming IN from the client counts: what was paid on a receipt (sale /
    debt) and advance deposits. Spending advance credit (advance_used) is not a new
    payment — that cash arrived on the deposit's own date, which is the one worth
    showing — and return credits / refunds move money the other way. Advances marked
    `is_opening` are dropped too: their cash was taken in earlier and the row carries
    the day it was written up, not the day the client handed anything over, so it says
    nothing about when they last actually paid."""
    on_sales = Payment.objects.filter(
        sale__client__in=client_pks, kind__in=(Payment.Kind.SALE, Payment.Kind.DEBT)
    )
    advances = Payment.objects.filter(
        client__in=client_pks, sale__isnull=True,
        kind=Payment.Kind.ADVANCE_IN, is_opening=False,
    )
    if not user.can_see_all_records:
        # Same scoping as everywhere else: a seller sees their own receipts and the
        # advances they took into their own till.
        on_sales = on_sales.filter(sale__sales_rep=user)
        advances = advances.filter(created_by=user)
    last = {}
    # A payment hangs off either the sale's client or the client directly, so the two
    # sides are aggregated separately and the later date wins.
    for qs, key in ((on_sales, "sale__client"), (advances, "client")):
        for row in qs.values(key).annotate(last=Max("date")):
            pk, when = row[key], row["last"]
            if when and (pk not in last or when > last[pk]):
                last[pk] = when
    return last


def _client_activity(user, client):
    """(last took goods, last paid) for one client — the pair the qarz and tarix
    pages lead with, read through the same maps the list pages use so one client's
    dates never disagree with the row they came from."""
    pks = [client.pk]
    return (
        _last_sale_map(user, pks).get(client.pk),
        _last_payment_map(user, pks).get(client.pk),
    )


def client_list(request):
    rows, q = _client_rows(request)
    # Headline KPIs over the whole (search-filtered) set, not just the current page.
    total_advance = sum((c.advance for c in rows), Decimal("0"))
    advance_clients = sum(1 for c in rows if c.advance > 0)
    page = Paginator(rows, 25).get_page(request.GET.get("page"))
    export_qs = request.GET.urlencode()
    return render(request, "crm/client_list.html", {
        "page": page,
        "q": q,
        "total_clients": len(rows),
        "total_advance": total_advance,
        "advance_clients": advance_clients,
        "export_url": reverse("client_export") + (f"?{export_qs}" if export_qs else ""),
    })


def client_export(request):
    """Excel (.xlsx) of the client list as searched — one row per client."""
    rows, _ = _client_rows(request)
    headers = [
        "Ismi", "Kompaniya", "Telefon", "Manzil", "Mas'ul xodim",
        "Avans (so'm)", "Qarz (so'm)", "Sotuvlar soni",
        "Oxirgi yuk olgan", "Oxirgi to'lov", "Izoh",
    ]
    data = [
        [
            c.name,
            c.company,
            c.phone,
            c.address,
            str(c.owner) if c.owner else "",
            float(c.advance),
            float(c.debt),
            c.sale_count,
            c.last_sale.strftime("%d.%m.%Y") if c.last_sale else "",
            c.last_payment.strftime("%d.%m.%Y") if c.last_payment else "",
            c.notes,
        ]
        for c in rows
    ]
    number_formats = {6: "#,##0.00", 7: "#,##0.00"}
    return _xlsx_response("mijozlar.xlsx", "Mijozlar", headers, data, number_formats)


def client_create(request):
    form = ClientForm(request.POST or None, user=request.user)
    if request.method == "POST":
        if form.is_valid():
            client = form.save(commit=False)
            # Admins/managers pick the responsible employee on the form; sellers'
            # clients are always owned by themselves.
            if not client.owner_id:
                client.owner = request.user
            client.save()
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Mijoz", client.pk,
                f"{client.name} qo'shildi (mas'ul: {client.owner})",
            )
            messages.success(request, f"“{client.name}” mijozi qo'shildi.")
            return form_success(request, reverse("client_list"))
        return form_response(request, form, "Yangi mijoz", invalid=True)
    return form_response(request, form, "Yangi mijoz")


def client_quick_create(request):
    """Create a client inline (from the sale form) and return it as JSON.

    Guards against accidental duplicates: an existing same-name client is
    reported back (409) so the caller can reuse it, unless allow_duplicate is set.
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST kerak"}, status=405)
    name = request.POST.get("name", "").strip()
    if not name:
        return JsonResponse({"error": "Ism kiritilishi shart"}, status=400)
    if not request.POST.get("allow_duplicate"):
        dup = Client.find_duplicate(request.user, name)
        if dup:
            return JsonResponse(
                {
                    "error": f"“{dup.name}” allaqachon bor",
                    "duplicate": True,
                    "existing": {"id": dup.pk, "text": dup.name},
                },
                status=409,
            )
    client = Client.objects.create(
        name=name, phone=request.POST.get("phone", "").strip(), owner=request.user
    )
    AuditLog.record(
        request.user, AuditLog.Action.CREATE, "Mijoz", client.pk,
        f"{client.name} (sotuv oynasidan tez qo'shildi)",
    )
    return JsonResponse({"id": client.pk, "text": client.name})


def _unique_product_sku(name):
    """A short, unique SKU derived from the product name (auto-assigned on a quick
    add; the admin can rename it later)."""
    base = "".join(c for c in name.upper() if c.isalnum())[:8] or "MHS"
    sku, i = base, 1
    while Product.objects.filter(sku=sku).exists():
        i += 1
        sku = f"{base}{i}"
    return sku


def product_quick_create(request):
    """Create a product inline (from the receipt form) so a seller can log goods for
    a product the admin hasn't defined yet. Returns it as JSON. A same-name product
    is reported back (409) so the caller can reuse it, unless allow_duplicate is set."""
    if request.method != "POST":
        return JsonResponse({"error": "POST kerak"}, status=405)
    name = (request.POST.get("name") or "").strip()
    if not name:
        return JsonResponse({"error": "Nom kiriting"}, status=400)
    if not request.POST.get("allow_duplicate") and Product.objects.filter(name__iexact=name).exists():
        return JsonResponse({"duplicate": True, "error": "Bu nomli mahsulot bor"}, status=409)

    def _dec(key):
        raw = (request.POST.get(key) or "").replace(" ", "").replace(",", ".")
        try:
            return Decimal(raw) if raw else Decimal("0")
        except (ArithmeticError, ValueError):
            return Decimal("0")

    product = Product.objects.create(
        name=name, sku=_unique_product_sku(name),
        price=_dec("price"), cost_price=_dec("cost_price"),
    )
    AuditLog.record(
        request.user, AuditLog.Action.CREATE, "Mahsulot", product.pk, f"{name} (tez qo'shildi)"
    )
    return JsonResponse({"id": product.pk, "text": str(product)})


def client_edit(request, pk):
    """The client's own details — and their advance, on the same screen.

    The advance rows have nothing to do with this form and are not saved by it. They
    are here because this is the pencil: the button reached for when something about a
    client is wrong. Sending someone who opened it to fix a mistyped deposit off to
    another screen to find the same pencil again is the kind of detour that gets a
    figure left wrong instead."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    form = ClientForm(
        request.POST or None, instance=client, user=request.user, check_duplicates=False
    )
    title = f"Tahrirlash: {client.name}"
    if request.method == "POST":
        if form.is_valid():
            # What actually changed, not just "yangilandi" — a renamed client or a
            # client moved to another seller is the kind of edit a trail is kept for.
            changes = form_changes(form)
            form.save()
            summary = f"{client.name} yangilandi"
            if changes:
                summary += f" — {changes}"
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Mijoz", client.pk, summary[:255]
            )
            messages.success(request, f"“{client.name}” mijozi yangilandi.")
            return form_reload(request, reverse("client_list"))
        return _render_client_edit(request, client, form, title, invalid=True)
    return _render_client_edit(request, client, form, title)


def _render_client_edit(request, client, form, title, invalid=False):
    context = {
        "form": form,
        "title": title,
        "client": client,
        **_client_advance_context(request, client),
    }
    if is_ajax(request):
        return render(
            request, "crm/_client_edit_modal.html", context,
            status=422 if invalid else 200,
        )
    return render(request, "crm/_client_edit_page.html", context)


def client_delete(request, pk):
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    if request.method == "POST":
        try:
            client.delete()
            AuditLog.record(
                request.user, AuditLog.Action.DELETE, "Mijoz", pk,
                f"{client.name} o'chirildi",
            )
            messages.success(request, f"“{client.name}” mijozi o'chirildi.")
        except ProtectedError:
            messages.error(
                request,
                f"“{client.name}” mijozini o'chirib bo'lmaydi — sotuvlari mavjud.",
            )
        return form_reload(request, reverse("client_list"))
    return render_confirm(
        request,
        "Mijozni o'chirish",
        f"“{client.name}” mijozi o'chiriladi. Bu amalni qaytarib bo'lmaydi.",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


def _render_client_transfer(request, client, form, invalid=False):
    context = {
        "form": form,
        "client": client,
        "sales_count": Sale.objects.filter(client=client).count(),
        "title": f"Mijozni o'tkazish: {client.name}",
    }
    if is_ajax(request):
        return render(
            request, "crm/_client_transfer_modal.html", context,
            status=422 if invalid else 200,
        )
    return render(request, "crm/form.html", context)


def client_transfer(request, pk):
    """Hand a client — and their whole sales history — to another seller.

    Full handover: the client's owner and every one of their sales' sales_rep
    move to the target, atomically. Sellers may transfer only clients they own
    (a non-owned client 404s via the visible-clients scope); admins/managers
    may transfer anyone's."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    if request.method == "POST":
        form = ClientTransferForm(request.POST, client=client)
        if form.is_valid():
            target = form.cleaned_data["new_owner"]
            old_owner = client.owner
            with transaction.atomic():
                moved = Sale.objects.filter(client=client).update(sales_rep=target)
                client.owner = target
                client.save(update_fields=["owner"])
                AuditLog.record(
                    request.user, AuditLog.Action.TRANSFER, "Mijoz", client.pk,
                    f"{client.name}: {old_owner} → {target} ({moved} ta sotuv)",
                )
            messages.success(
                request, f"“{client.name}” {target}ga o'tkazildi ({moved} ta sotuv)."
            )
            return form_reload(request, reverse("client_list"))
        return _render_client_transfer(request, client, form, invalid=True)
    form = ClientTransferForm(client=client)
    return _render_client_transfer(request, client, form)


# Money events on a client's account, in the order they happen. `delta` is how the
# event moved the client's debt: a sale lifts it, a payment or a return brings it
# down, and money handed back on an over-returned sale lifts it again. Advance
# deposits carry no delta — the cash is held as credit and only touches a debt
# later, as an ADVANCE_USED payment.
_PAY_EVENTS = {
    Payment.Kind.SALE: ("Sotuvda to'landi", "badge-ok", "in"),
    Payment.Kind.DEBT: ("Qarz to'lovi", "badge-ok", "in"),
    Payment.Kind.ADVANCE_USED: ("Avansdan yechildi", "badge-info", "in"),
    Payment.Kind.ADVANCE_IN: ("Avans olindi", "badge-ok", "in"),
    Payment.Kind.ADVANCE_OUT: ("Avans qaytarildi", "badge-danger", "out"),
    Payment.Kind.RETURN_CREDIT: ("Qaytarishdan kredit", "badge-shipped", "return"),
    Payment.Kind.REFUND_OUT: ("Naqd qaytarildi", "badge-danger", "out"),
    Payment.Kind.ADJUST_CREDIT: ("Narx tuzatildi — kreditga", "badge-shipped", "return"),
    Payment.Kind.ADJUST_REFUND: ("Narx tuzatildi — naqd berildi", "badge-danger", "out"),
}


def _payment_event(payment):
    """One payment as a history row. Each row counts what it actually credited the
    client with, so the running balance follows their debt: the full sum they sent
    when the seller carried the bank fee, the net when they carried it themselves."""
    label, cls, icon = _PAY_EVENTS[payment.kind]
    paying = payment.kind in PAYING_KINDS
    amount = payment.credited_amount
    if paying:
        delta = -amount
    elif payment.kind in (Payment.Kind.ADVANCE_IN, Payment.Kind.ADVANCE_OUT):
        # Credit moving into or out of the client's own pool. Neither touches what
        # they owe — only an advance actually spent on a receipt does that.
        delta = Decimal("0")
    else:  # a settlement kind — money owed back to the client
        delta = amount
    notes = []
    if payment.commission:
        notes.append(f"bank komissiyasi {payment.commission:,.0f}")
    if payment.currency == Payment.Currency.USD:
        notes.append(
            f"${payment.original_amount:,.2f} × {payment.exchange_rate:,.0f}"
        )
    if payment.note:
        notes.append(payment.note)
    return {
        "date": payment.date,
        "sort": (payment.date, payment.created_at),
        "label": label,
        "cls": cls,
        "icon": icon,
        "desc": payment.get_kind_display(),
        "amount": amount,
        "delta": delta,
        "method": payment.method,
        "method_label": payment.get_method_display(),
        "user": payment.created_by,
        "note": " · ".join(notes),
        "url": reverse("sale_detail", args=[payment.sale_id]) if payment.sale_id else "",
    }


def _client_events(request, client):
    """A client's whole money history, oldest first, with a running debt balance.

    Sales, payments and returns are merged into one timeline; `balance` on each row
    is what the client owed right after that event, so the last row's balance is
    today's debt — the same figure the qarzlar page shows. Seller-scoped: a seller
    sees only the sales they made and the advances they took in themselves."""
    sales = (
        Sale.objects.visible_to(request.user)
        .filter(client=client)
        .select_related("sales_rep")
        .prefetch_related(
            "items__product", "returns__product", "payments__created_by"
        )
    )
    today = timezone.localdate()
    events = []
    for sale in sales:
        # Only a receipt that is BOTH past its deadline and still unpaid is late; a
        # settled one stopped counting the day it was paid. Payments and returns are
        # movements, not obligations, so they never carry the figure.
        deadline = sale.debt_deadline
        events.append({
            "date": sale.date,
            "sort": (sale.date, sale.created_at),
            "label": "Ochilish qoldig'i" if sale.is_opening else "Sotuv",
            "cls": "badge-neutral" if sale.is_opening else "badge-info",
            "icon": "sale",
            "desc": sale.item_summary,
            "overdue_days": (
                (today - deadline).days
                if deadline and deadline < today and sale.debt_remaining > 0
                else None
            ),
            # An opening carry-over has no line items: its debt is opening_amount alone.
            "amount": sale.total_price + sale.opening_amount,
            "delta": sale.total_price + sale.opening_amount,
            "method": "",
            "method_label": "",
            "user": sale.sales_rep,
            "note": "",
            "url": reverse("sale_detail", args=[sale.pk]),
        })
        for payment in sale.payments.all():
            events.append(_payment_event(payment))
        for ret in sale.returns.all():
            events.append({
                "date": ret.date,
                "sort": (ret.date, ret.created_at),
                "label": "Qaytarish",
                "cls": "badge-shipped",
                "icon": "return",
                "desc": f"{ret.product.name} · {_kg(ret.weight_kg)} kg",
                "amount": ret.amount,
                "delta": -ret.amount,
                "method": "",
                "method_label": "",
                "user": ret.created_by,
                "note": ret.note,
                "url": reverse("sale_detail", args=[ret.sale_id]),
            })
    # Advance deposits live on the client, not on any sale, so they are fetched
    # separately — scoped to the seller's own till like every other advance figure.
    advances = Payment.objects.filter(
        client=client, sale__isnull=True
    ).select_related("created_by")
    if not request.user.can_see_all_records:
        advances = advances.filter(created_by=request.user)
    events.extend(_payment_event(p) for p in advances)

    events.sort(key=lambda e: e["sort"])
    balance = Decimal("0")
    for event in events:
        balance += event["delta"]
        event["balance"] = balance
    return events


def _client_history_totals(request, client, events):
    def total(*labels):
        return sum(
            (e["amount"] for e in events if e["label"] in labels), Decimal("0")
        )

    scope = None if request.user.can_see_all_records else request.user
    return {
        "sold": total("Sotuv", "Ochilish qoldig'i"),
        "paid": total("Sotuvda to'landi", "Qarz to'lovi", "Avansdan yechildi"),
        "returned": total("Qaytarish"),
        "debt": events[-1]["balance"] if events else Decimal("0"),
        "advance": client_advance_balance(client, scope),
        "sales_count": sum(1 for e in events if e["label"] == "Sotuv"),
    }


def client_history(request, pk):
    """One client's full account history: every sale, payment and return on a
    single timeline, with what they owed after each step."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    events = _client_events(request, client)
    last_sale, last_payment = _client_activity(request.user, client)
    return render(request, "crm/client_history.html", {
        "client": client,
        "events": events,
        "totals": _client_history_totals(request, client, events),
        "last_sale": last_sale,
        "last_payment": last_payment,
    })


def client_history_export(request, pk):
    """Excel (.xlsx) of one client's history — the same rows as the page."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    events = _client_events(request, client)
    headers = [
        "Sana", "Amal", "Tafsilot", "Summa",
        "Qarz o'zgarishi", "Qarz qoldig'i", "Kechikkan kun", "Usul", "Kim", "Izoh",
        "Oxirgi yuk olgan", "Oxirgi to'lov",
    ]
    # The last two are client-level, so they repeat down every row — the same pair the
    # page shows above the timeline, carried into the file the client is sent.
    last_sale, last_payment = _client_activity(request.user, client)
    stamps = [
        last_sale.strftime("%d.%m.%Y") if last_sale else "",
        last_payment.strftime("%d.%m.%Y") if last_payment else "",
    ]
    rows = [
        [
            e["date"].strftime("%d.%m.%Y"),
            e["label"],
            e["desc"],
            float(e["amount"]),
            float(e["delta"]),
            float(e["balance"]),
            e.get("overdue_days") if e.get("overdue_days") is not None else "",
            e["method_label"],
            str(e["user"]),
            e["note"],
            *stamps,
        ]
        for e in events
    ]
    number_formats = {4: "#,##0.00", 5: "#,##0.00", 6: "#,##0.00"}
    # The client's name can be Cyrillic; keep the filename ASCII so the download
    # header needs no encoding games.
    return _xlsx_response(
        f"mijoz-{client.pk}-tarix.xlsx", "Tarix", headers, rows, number_formats
    )


def _opening_sale(client):
    """The receipt carrying a client's pre-CRM debt, if they have one. There is at most
    one per client (the importers create a single row); the oldest wins if an older
    import ever left two behind."""
    return Sale.objects.filter(client=client, is_opening=True).order_by("pk").first()


def _render_opening_debt(request, client, sale, form, invalid=False):
    context = {
        "form": form,
        "client": client,
        "opening": sale,
        "title": f"Boshlang'ich qarz: {client.name}",
    }
    if is_ajax(request):
        return render(
            request, "crm/_opening_debt_modal.html", context,
            status=422 if invalid else 200,
        )
    return render(request, "crm/_opening_debt_page.html", context)


@transaction.atomic
def client_opening_debt(request, pk):
    """Set (or first record) a client's opening balance — the debt they carried in from
    before the CRM existed.

    These used to arrive only through the import commands, so a client whose old ledger
    was wrong could not be corrected from the app at all: the figure could be read but
    never touched. The seller enters the sum to add (or subtract) and the form does the
    arithmetic — see `OpeningDebtForm` for why it is a delta and not a total.

    An opening balance has no line items, so it never disturbs revenue, profit or sold
    kg — it moves the receivable and nothing else. Lowering it below what has already
    been paid against it is refused, for the same reason `sale_edit` refuses it: the
    receipt would go negative with nobody owed the difference.

    Deliberately not restricted to admins: the seller who owns the client is the one who
    knows their old ledger, and every change is written to the audit log with the before
    and after figures. Visibility is the existing rule — a seller sees only their own
    clients."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    sale = _opening_sale(client)
    paid = (sale.paid_amount - sale.settled_amount) if sale else Decimal("0")
    was = sale.opening_amount if sale else Decimal("0")
    if request.method == "POST":
        form = OpeningDebtForm(request.POST, current=was, paid=paid)
        if form.is_valid():
            amount, on_date = form.new_total, form.cleaned_data["date"]
            if sale is None:
                sale = Sale.objects.create(
                    client=client,
                    sales_rep=client.owner or request.user,
                    date=on_date,
                    debt_deadline=on_date + timedelta(days=DEFAULT_DEBT_DAYS),
                    is_opening=True,
                    opening_amount=amount,
                )
                action = AuditLog.Action.CREATE
            else:
                # Carry the deadline with the date rather than recomputing it: these
                # debts are imported with their own terms, and silently resetting the
                # window would restate how overdue the client is.
                if sale.debt_deadline:
                    sale.debt_deadline += on_date - sale.date
                sale.date = on_date
                sale.opening_amount = amount
                sale.save(update_fields=["opening_amount", "date", "debt_deadline"])
                action = AuditLog.Action.UPDATE
            # A raised balance is a fresh open receipt any credit the client holds
            # should settle onto; a lowered one may free credit that was already spent.
            _reconcile_client_advance(client, sale.sales_rep)
            AuditLog.record(
                request.user, action, "Sotuv", sale.pk,
                f"Mijoz {client.name} boshlang'ich qarzi "
                f"{was:,.0f} → {amount:,.0f} so'm",
            )
            messages.success(
                request,
                f"Boshlang'ich qarz saqlandi: {amount:,.0f} so'm "
                f"(oldin {was:,.0f} so'm).",
            )
            return form_reload(request, reverse("client_history", args=[client.pk]))
        return _render_opening_debt(request, client, sale, form, invalid=True)
    form = OpeningDebtForm(current=was, paid=paid, initial={
        "date": sale.date if sale else timezone.localdate(),
    })
    return _render_opening_debt(request, client, sale, form)


# --- Products -----------------------------------------------------------------

# Paket catalogue facets, mirroring seed_paket_products: a SKU reads
# "{size}-{micron}-{colour}" (e.g. "1,5m-015-oq"), so each facet is a slice of it.
# "-01-" never matches "-015-", so the micron grades stay distinct.
PAKET_COLORS = [("oq", "ОҚ"), ("qora", "ҚОРА"), ("novot", "НОВОТ")]
PAKET_SIZES = [("1,5m", "1,5м"), ("2m", "2м"), ("6m", "6м")]
PAKET_MICRONS = ["015", "01", "08", "06", "05", "04", "03", "02"]


def product_list(request):
    # A plain shared catalog — the reference list sellers pick from when selling.
    # With 56 paket SKUs a text search alone is coarse, so the drawer also filters
    # by the three facets encoded in the SKU.
    products = Product.objects.order_by("name")
    filters = {key: request.GET.get(key, "") for key in ("color", "size", "micron")}
    filters["q"] = request.GET.get("q", "").strip()

    if filters["q"]:
        products = products.filter(
            Q(name__icontains=filters["q"]) | Q(sku__icontains=filters["q"])
        )

    colors, sizes = dict(PAKET_COLORS), dict(PAKET_SIZES)
    # Only known facet values bite; anything else is ignored rather than 0-matching.
    if filters["color"] in colors:
        products = products.filter(sku__endswith=f"-{filters['color']}")
    if filters["size"] in sizes:
        products = products.filter(sku__startswith=f"{filters['size']}-")
    if filters["micron"] in PAKET_MICRONS:
        products = products.filter(sku__contains=f"-{filters['micron']}-")

    active_filters = _filter_chips(request, [
        {"param": "color", "label": "Rang", "value": colors.get(filters["color"], "")},
        {"param": "size", "label": "O'lcham", "value": sizes.get(filters["size"], "")},
        {"param": "micron", "label": "Mikron",
         "value": filters["micron"] if filters["micron"] in PAKET_MICRONS else ""},
    ])

    page = Paginator(products, 25).get_page(request.GET.get("page"))
    return render(request, "crm/product_list.html", {
        "page": page,
        "q": filters["q"],
        "filters": filters,
        "active_filters": active_filters,
        "filter_count": len(active_filters),
        "has_filters": bool(active_filters),
        "filter_url": reverse("product_list"),
        "search_placeholder": "Nomi bo'yicha qidirish…",
        "paket_colors": PAKET_COLORS,
        "paket_sizes": PAKET_SIZES,
        "paket_microns": PAKET_MICRONS,
    })


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    recent_items = product.sale_items.select_related("sale", "sale__client").order_by(
        "-sale__date", "-sale__created_at"
    )
    # The warehouse is shared, so everyone sees the stock-movement log. Sellers
    # still see only their OWN recent sales of the product. Filter before slicing.
    entries = product.stock_entries.select_related("created_by")[:50]
    if not request.user.can_see_all_records:
        recent_items = recent_items.filter(sale__sales_rep=request.user)
    recent_items = recent_items[:10]
    context = {
        "product": product,
        "current_stock": product.current_stock,
        "total_received": product.total_received,
        "total_sold": product.total_sold,
        "entries": entries,
        "recent_items": recent_items,
    }
    return render(request, "crm/product_detail.html", context)


def product_create(request):
    form = ProductForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            product = form.save()
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Mahsulot", product.pk,
                f"{product.name} qo'shildi — narx {product.price:,.0f} so'm, "
                f"tannarx {product.cost_price:,.0f} so'm",
            )
            messages.success(request, f"“{product.name}” mahsuloti qo'shildi.")
            return form_success(request, reverse("product_detail", args=[product.pk]))
        return form_response(request, form, "Yangi mahsulot", invalid=True)
    return form_response(request, form, "Yangi mahsulot")


def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    # Read BEFORE the form is validated: a ModelForm writes the posted values onto its
    # instance during `is_valid()`, so by then `product.price` is already the new one.
    was_price, was_cost = product.price, product.cost_price
    form = ProductForm(request.POST or None, instance=product)
    title = f"Tahrirlash: {product.name}"
    if request.method == "POST":
        if form.is_valid():
            # A price is the one field on this form that changes what money the next
            # sale makes, so it is spelled out rather than left as "yangilandi".
            form.save()
            moved = []
            if product.price != was_price:
                moved.append(f"narx {was_price:,.0f} → {product.price:,.0f} so'm")
            if product.cost_price != was_cost:
                moved.append(f"tannarx {was_cost:,.0f} → {product.cost_price:,.0f} so'm")
            summary = f"{product.name} yangilandi"
            if moved:
                summary += " — " + ", ".join(moved)
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Mahsulot", product.pk, summary[:255]
            )
            messages.success(request, f"“{product.name}” mahsuloti yangilandi.")
            return form_reload(request, reverse("product_detail", args=[product.pk]))
        return form_response(request, form, title, invalid=True)
    return form_response(request, form, title)


def product_delete(request, pk):
    """Remove a product from the shared catalogue. Blocked (ProtectedError) when
    it has any sales or returns — those records must keep pointing at a real
    product. A product with only stock entries deletes cleanly (entries cascade)."""
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        name = product.name
        try:
            product.delete()
        except ProtectedError:
            messages.error(
                request,
                f"“{name}” mahsulotini o'chirib bo'lmaydi — sotuv yoki "
                f"qaytarishlarda ishlatilgan.",
            )
            return form_reload(request, reverse("product_list"))
        AuditLog.record(request.user, AuditLog.Action.DELETE, "Mahsulot", pk, name)
        messages.success(request, f"“{name}” mahsuloti o'chirildi.")
        return form_reload(request, reverse("product_list"))
    return render_confirm(
        request,
        "Mahsulotni o'chirish",
        f"“{product.name}” mahsuloti o'chiriladi. Bu amalni qaytarib bo'lmaydi.",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


def stock_entry_create(request, pk):
    product = get_object_or_404(Product, pk=pk)
    title = f"Kirim qo'shish: {product.name}"
    form = StockEntryForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            entry = form.save(commit=False)
            entry.product = product
            entry.created_by = request.user
            entry.save()
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Ombor", entry.pk,
                f"{product.name} — {entry.quantity_kg:.3f} kg kirim"
                + (f" ({entry.note})" if entry.note else ""),
            )
            messages.success(
                request, f"“{product.name}” omboriga {entry.quantity_kg} kg kirim qilindi."
            )
            return form_success(request, reverse("product_detail", args=[product.pk]))
        return form_response(request, form, title, invalid=True)
    return form_response(request, form, title)


def stock_adjust(request, pk):
    product = get_object_or_404(Product, pk=pk)
    title = f"Miqdorni tuzatish: {product.name}"
    if request.method == "POST":
        form = StockAdjustForm(request.POST)
        if form.is_valid():
            current = product.current_stock
            target = form.cleaned_data["quantity"]
            delta = target - current
            if delta != 0:
                note = form.cleaned_data["note"] or (
                    f"Miqdor tuzatildi: {current:.3f} → {target:.3f} kg"
                )
                entry = StockEntry.objects.create(
                    product=product, quantity_kg=delta, note=note, created_by=request.user
                )
                AuditLog.record(
                    request.user, AuditLog.Action.UPDATE, "Ombor", entry.pk,
                    f"{product.name} — miqdor tuzatildi: {current:.3f} → {target:.3f} kg "
                    f"({delta:+.3f})",
                )
                messages.success(
                    request, f"“{product.name}” ombori {target:.3f} kg qilib belgilandi."
                )
            else:
                messages.info(request, "Miqdor o'zgarmadi.")
            return form_success(request, reverse("product_detail", args=[product.pk]))
        return form_response(request, form, title, invalid=True)
    form = StockAdjustForm(initial={"quantity": product.current_stock})
    return form_response(request, form, title)


# --- Sales --------------------------------------------------------------------

# Formatting a phone number picks up along the way — the digits are what people
# actually search by, so both sides are stripped down to them before comparing.
PHONE_NOISE = (" ", "+", "-", "(", ")", ".")


def _phone_digits_expr(field):
    """SQL expression reducing a stored phone to bare digits, so
    '+998 90 123 45 67' compares as '998901234567'."""
    expr = F(field)
    for char in PHONE_NOISE:
        expr = Replace(expr, Value(char), Value(""))
    return expr


def _client_search(qs, term, base="", extra=None):
    """Filter `qs` by its client's name / company / phone / location, case-insensitive
    — the toolbar's search box. `base` is the lookup path to the Client ("client" on
    a Sale, "sale__client" on a Payment); leave it empty when `qs` IS the clients.
    `extra` is an already-built Q of further lookups to OR in.

    A term with digits in it also matches the phone with its formatting removed, so
    "901234567", "90 123 45 67" and "+998901234567" all find the same client. Short
    digit runs are skipped: a "5" in a name shouldn't drag in every phone."""
    at = f"{base}__" if base else ""
    match = (
        Q(**{f"{at}name__icontains": term})
        | Q(**{f"{at}company__icontains": term})
        | Q(**{f"{at}phone__icontains": term})
        | Q(**{f"{at}address__icontains": term})
    )
    if extra is not None:
        match |= extra
    digits = re.sub(r"\D", "", term)
    if len(digits) >= 3:
        # `base` is a plain FK here, so annotating can't multiply rows.
        qs = qs.annotate(_phone_digits=_phone_digits_expr(f"{at}phone"))
        match |= Q(_phone_digits__contains=digits)
    return qs.filter(match)


def _filter_sales(request, sales):
    """Filter sales by client/product/rep/status and, only when no such filter
    is active, a date window (dan..gacha, default today..today).

    A content filter searches across ALL dates — the date window is the default
    (unfiltered) view's concern, so the two never apply at once.
    Returns (queryset, filters, date_from, date_to, has_filters)."""
    today = timezone.localdate()
    filters = {key: request.GET.get(key, "") for key in ("client", "product", "rep", "status")}
    filters["q"] = request.GET.get("q", "").strip()
    has_filters = bool(
        filters["q"]
        or filters["client"].isdigit()
        or filters["product"].isdigit()
        or filters["status"] in ("paid", "debt", "overdue")
        or (filters["rep"].isdigit() and request.user.can_see_all_records)
    )

    date_from = _parse_date(request.GET.get("dan")) or today
    date_to = _parse_date(request.GET.get("gacha")) or date_from
    if date_to < date_from:
        date_from, date_to = date_to, date_from
    if not has_filters:
        sales = sales.filter(date__gte=date_from, date__lte=date_to)

    filters["dan"] = date_from.isoformat()
    filters["gacha"] = date_to.isoformat()
    if filters["q"]:
        sales = _client_search(sales, filters["q"], "client")
    if filters["client"].isdigit():
        sales = sales.filter(client_id=filters["client"])
    if filters["product"].isdigit():
        sales = sales.filter(items__product_id=filters["product"]).distinct()
    if filters["rep"].isdigit() and request.user.can_see_all_records:
        sales = sales.filter(sales_rep_id=filters["rep"])
    # Status is derived from the running balance (annotated by with_balance)
    if filters["status"] == "paid":
        sales = sales.filter(remaining__lte=0)
    elif filters["status"] == "debt":
        sales = sales.filter(remaining__gt=0)
    elif filters["status"] == "overdue":
        sales = sales.filter(remaining__gt=0, debt_deadline__lt=today)
    return sales, filters, date_from, date_to, has_filters


def _filter_chips(request, specs):
    """Build removable filter chips from a list of specs:
    {"param", "label", "value"}. Only specs with a truthy `value` produce a chip.
    The remove-URL drops that param, the page, and any empty filter params."""
    params = [s["param"] for s in specs]

    def without(param):
        qs = request.GET.copy()
        qs.pop(param, None)
        qs.pop("page", None)
        for key in params:
            if not qs.get(key):
                qs.pop(key, None)
        query = qs.urlencode()
        return f"{request.path}?{query}" if query else request.path

    return [
        {"label": s["label"], "value": s["value"], "remove_url": without(s["param"])}
        for s in specs
        if s.get("value")
    ]


def _segment_url(request, **params):
    """The current query with `params` set — an empty value drops its param, and the
    page number always goes, since a narrower view has different pages."""
    qs = request.GET.copy()
    for key, value in params.items():
        if value in ("", None):
            qs.pop(key, None)
        else:
            qs[key] = value
    qs.pop("page", None)
    query = qs.urlencode()
    return f"{request.path}?{query}" if query else request.path


def _segment(request, param, value, label, count, current):
    """One button of a toolbar switch: the current query with `param` set to `value`
    (or dropped, for the "everything" button). Unlike a chip it is always visible —
    the choice is part of the page, not a filter you applied and can remove."""
    return {
        "label": label,
        "count": count,
        "url": _segment_url(request, **{param: value}),
        "active": current == value,
    }


def _active_filter_chips(request, filters, clients, products, reps):
    """Sotuvlar filter chips (client/product/rep/status)."""
    status_labels = {"paid": "To'langan", "debt": "Qarz", "overdue": "Muddati o'tgan"}
    client = clients.filter(pk=filters["client"]).first() if filters["client"].isdigit() else None
    product = products.filter(pk=filters["product"]).first() if filters["product"].isdigit() else None
    rep = reps.filter(pk=filters["rep"]).first() if reps and filters["rep"].isdigit() else None
    specs = [
        {"param": "client", "label": "Mijoz", "value": client.name if client else ""},
        {"param": "product", "label": "Mahsulot", "value": product.name if product else ""},
        {"param": "rep", "label": "Sotuvchi", "value": str(rep) if rep else ""},
        {"param": "status", "label": "To'lov", "value": status_labels.get(filters["status"], "")},
    ]
    return _filter_chips(request, specs)


def _month_end(day):
    """The last date of the calendar month `day` falls in."""
    return day.replace(day=calendar.monthrange(day.year, day.month)[1])


def _date_range_context(request, default_window="today"):
    """Parse ?dan/?gacha into a today-default window plus the navigation vars
    the shared toolbar's date-range picker needs.

    `default_window` picks what a page opens on when NEITHER date is given:
    "today", "month" for the whole current calendar month (what the ombor report
    wants — it reads as a monthly sverka, and a single day's slice of one is usually
    empty), or "all" for no window at all — what a full history like the audit trail
    opens on, where cutting to one day by default would hide almost everything.
    Once either date is in the URL the old parsing stands unchanged, so every
    existing link, filter and day-step arrow still means exactly what it did.

    `is_month` tells the toolbar the window is one whole calendar month, so it can
    label it by name ("Avgust") instead of spelling out both end dates."""
    today = timezone.localdate()
    dan = _parse_date(request.GET.get("dan"))
    gacha = _parse_date(request.GET.get("gacha"))
    if dan is None and gacha is None and default_window == "all":
        # No window: the picker says "Hammasi" and its arrows step days from today,
        # so the first click lands somewhere meaningful rather than in 1970.
        return {
            "date_from": None,
            "date_to": None,
            "range_days": 0,
            "is_single_day": False,
            "is_month": False,
            "is_today": False,
            "is_all": True,
            "prev_from": (today - timedelta(days=1)).isoformat(),
            "prev_to": (today - timedelta(days=1)).isoformat(),
            "next_from": today.isoformat(),
            "next_to": today.isoformat(),
            "today_iso": today.isoformat(),
        }
    if dan is None and gacha is None and default_window == "month":
        date_from = today.replace(day=1)
        date_to = _month_end(today)
    elif dan is None and gacha is None:
        date_from = date_to = today
    else:
        date_from = dan or today
        date_to = gacha or date_from
    if date_to < date_from:
        date_from, date_to = date_to, date_from
    return {
        "date_from": date_from,
        "date_to": date_to,
        "range_days": (date_to - date_from).days + 1,
        "is_single_day": date_from == date_to,
        "is_month": date_from.day == 1 and date_to == _month_end(date_from),
        "is_today": date_from == today and date_to == today,
        "is_all": False,
        "prev_from": (date_from - timedelta(days=1)).isoformat(),
        "prev_to": (date_to - timedelta(days=1)).isoformat(),
        "next_from": (date_from + timedelta(days=1)).isoformat(),
        "next_to": (date_to + timedelta(days=1)).isoformat(),
        "today_iso": today.isoformat(),
    }


def _outstanding_balance(sales):
    """Total still owed across the given sales: item revenue − returns − payments,
    plus any carried-over opening balance.

    Payments count gross (a bank fee is the seller's cost, not the client's) and
    returned goods are subtracted, matching how each sale's remaining balance is
    computed. `opening_amount`
    (0 on a normal sale) is the pre-CRM debt that has no line items — it must be added
    here or an imported opening debt would total to zero. Mirrors the `remaining`
    annotation in `SaleQuerySet.with_balance`."""
    pks = sales.values("pk")
    revenue = SaleItem.objects.filter(sale__in=pks).aggregate(v=Sum(REVENUE))["v"] or 0
    returned = Return.objects.filter(sale__in=pks).aggregate(v=Sum(RETURN_AMOUNT))["v"] or 0
    paid = Payment.objects.filter(sale__in=pks).aggregate(v=Sum(PAYMENT_CREDIT))["v"] or 0
    opening = Sale.objects.filter(pk__in=pks).aggregate(v=Sum("opening_amount"))["v"] or 0
    return revenue - returned - paid + opening


def sale_list(request):
    base = (
        Sale.objects.visible_to(request.user)
        .real()  # opening-balance carry-overs live on the Qarzlar page, not here
        .select_related("client", "sales_rep")
        .prefetch_related("items__product")
        .with_balance()
    )
    sales, filters, date_from, date_to, has_filters = _filter_sales(request, base)
    sales = sales.order_by("-date", "-created_at")

    totals = _sale_totals(sales)
    outstanding = sales.filter(remaining__gt=0)
    totals["debt"] = _outstanding_balance(outstanding)
    totals["debtors"] = outstanding.values("client").distinct().count()

    # Real ratios for the KPI card meta-lines (no fabricated trends)
    revenue = totals["revenue"] or 0
    total_clients = _visible_clients(request.user).count()
    totals["count"] = sales.count()
    totals["margin"] = (totals["profit"] or 0) / revenue * 100 if revenue else 0
    totals["debt_share"] = (totals["debt"] or 0) / revenue * 100 if revenue else 0
    totals["debtor_pct"] = totals["debtors"] / total_clients * 100 if total_clients else 0

    clients = _visible_clients(request.user).order_by("name")
    products = Product.objects.order_by("name")
    reps = (
        User.objects.filter(is_active=True).order_by("first_name", "username")
        if request.user.can_see_all_records
        else None
    )
    active_filters = _active_filter_chips(request, filters, clients, products, reps)
    page = Paginator(sales, 25).get_page(request.GET.get("page"))
    export_qs = request.GET.urlencode()
    return render(
        request,
        "crm/sale_list.html",
        {
            "page": page,
            "totals": totals,
            "filters": filters,
            "has_filters": has_filters,
            "active_filters": active_filters,
            "filter_count": len(active_filters),
            **_date_range_context(request),
            "clients": clients,
            "products": products,
            "reps": reps,
            "export_qs": export_qs,
            "filter_url": reverse("sale_list"),
            "sale_export_url": reverse("sale_export") + (f"?{export_qs}" if export_qs else ""),
        },
    )


def _debtor_rows(request):
    """One row per debtor client for the current filters: total owed, open receipts,
    earliest deadline. Shared by the Qarzlar page and its Excel export."""
    today = timezone.localdate()
    open_sales = (
        Sale.objects.visible_to(request.user).outstanding().select_related("client")
    )

    filters = {key: request.GET.get(key, "") for key in ("client", "rep", "overdue", "tur")}
    filters["q"] = request.GET.get("q", "").strip()
    if filters["q"]:
        open_sales = _client_search(open_sales, filters["q"], "client")
    if filters["client"].isdigit():
        open_sales = open_sales.filter(client_id=filters["client"])
    if filters["rep"].isdigit() and request.user.can_see_all_records:
        open_sales = open_sales.filter(sales_rep_id=filters["rep"])
    if filters["overdue"] == "1":
        open_sales = open_sales.filter(debt_deadline__lt=today)

    groups = {}
    for sale in open_sales:
        remaining = sale.remaining
        group = groups.get(sale.client_id)
        if group is None:
            group = groups[sale.client_id] = {
                "client": sale.client,
                "remaining": Decimal("0"),
                "count": 0,
                "earliest": sale.debt_deadline,
                "overdue_count": 0,
                "overdue_amount": Decimal("0"),
                "advance": Decimal("0"),
                "advance_since": None,
            }
        group["remaining"] += remaining
        group["count"] += 1
        if sale.debt_deadline and (
            group["earliest"] is None or sale.debt_deadline < group["earliest"]
        ):
            group["earliest"] = sale.debt_deadline
        if sale.debt_deadline and sale.debt_deadline < today:
            group["overdue_count"] += 1
            group["overdue_amount"] += remaining

    # The advance pool belongs on this page too — an advance is the same account read
    # the other way round: money already taken for goods not yet handed over. Most
    # clients holding one owe nothing, so they get a row of their own instead of only a
    # column on the debtors; otherwise their money would sit in the headline figure with
    # no line on the page to explain it. Seller-scoped like everywhere else, and an
    # admin filtering by a seller sees that seller's till.
    clients = _visible_clients(request.user)
    if filters["q"]:
        clients = _client_search(clients, filters["q"])
    if filters["client"].isdigit():
        clients = clients.filter(pk=filters["client"])
    scope = None if request.user.can_see_all_records else request.user
    if request.user.can_see_all_records and filters["rep"].isdigit():
        scope = User.objects.filter(pk=filters["rep"]).first()
    client_map = {c.pk: c for c in clients}
    for pk, advance in _advance_balance_map(list(client_map), scope).items():
        if advance <= 0:
            continue
        group = groups.get(pk)
        if group is not None:
            group["advance"] = advance
        elif filters["overdue"] != "1":
            # Nothing outstanding: no receipts, no deadline — just the balance held.
            groups[pk] = {
                "client": client_map[pk],
                "remaining": Decimal("0"),
                "count": 0,
                "earliest": None,
                "overdue_count": 0,
                "overdue_amount": Decimal("0"),
                "advance": advance,
            }

    # How long that money has been lying there — the advance's own "deadline": nobody
    # is chasing it, so its age is the only thing that says it needs attention.
    advance_pks = [pk for pk, g in groups.items() if g["advance"]]
    for pk, when in _advance_since_map(advance_pks, scope).items():
        groups[pk]["advance_since"] = when

    # The Qarzdorlar / Avans switch. Counted before the switch is applied, so each
    # button can carry how many rows it leads to — including the one you are not on.
    counts = {
        "": len(groups),
        "qarz": sum(1 for g in groups.values() if g["remaining"]),
        "avans": sum(1 for g in groups.values() if g["advance"]),
    }
    if filters["tur"] == "qarz":
        groups = {pk: g for pk, g in groups.items() if g["remaining"]}
    elif filters["tur"] == "avans":
        groups = {pk: g for pk, g in groups.items() if g["advance"]}
    else:
        filters["tur"] = ""

    # When each debtor last took goods and last paid — the two dates you want in front
    # of you before ringing them up.
    pks = list(groups)
    sale_map = _last_sale_map(request.user, pks)
    pay_map = _last_payment_map(request.user, pks)
    for pk, group in groups.items():
        group["last_sale"] = sale_map.get(pk)
        group["last_payment"] = pay_map.get(pk)

    # Most urgent first: overdue (earliest deadlines) at the top, and the advance-only
    # rows — nobody chasing them — after everyone who owes.
    debtors = sorted(
        groups.values(), key=lambda g: (0 if g["remaining"] else 1, g["earliest"] or today)
    )
    # Headline figures over the rows actually on the page, so the KPI cards and the
    # table can never tell two different stories.
    totals = {
        "debt": sum((g["remaining"] for g in debtors), Decimal("0")),
        "overdue": sum((g["overdue_amount"] for g in debtors), Decimal("0")),
        "advance": sum((g["advance"] for g in debtors), Decimal("0")),
        "debtors": sum(1 for g in debtors if g["remaining"]),
        "overdue_debtors": sum(1 for g in debtors if g["overdue_count"]),
        "advance_clients": sum(1 for g in debtors if g["advance"]),
        "counts": counts,
    }
    return debtors, filters, totals


def debt_list(request):
    """One row per debtor client: total owed, open receipts, earliest deadline."""
    debtors, filters, totals = _debtor_rows(request)
    clients = _visible_clients(request.user).order_by("name")
    reps = (
        User.objects.filter(is_active=True).order_by("first_name", "username")
        if request.user.can_see_all_records
        else None
    )
    client_obj = clients.filter(pk=filters["client"]).first() if filters["client"].isdigit() else None
    rep_obj = reps.filter(pk=filters["rep"]).first() if reps and filters["rep"].isdigit() else None
    active_filters = _filter_chips(request, [
        {"param": "client", "label": "Mijoz", "value": client_obj.name if client_obj else ""},
        {"param": "rep", "label": "Sotuvchi", "value": str(rep_obj) if rep_obj else ""},
        {"param": "overdue", "label": "Holat", "value": "Muddati o'tgan" if filters["overdue"] == "1" else ""},
    ])
    # Qarzdorlar / Avans — a switch, not a chip: it sits in the toolbar as its own
    # buttons, each carrying the number of rows behind it.
    counts = totals["counts"]
    segments = [
        _segment(request, "tur", code, label, counts[code], filters["tur"])
        for code, label in (("", "Hammasi"), ("qarz", "Qarzdorlar"), ("avans", "Avans"))
    ]

    return render(
        request,
        "crm/debt_list.html",
        {
            "debtors": debtors,
            "total_debt": totals["debt"],
            "overdue_total": totals["overdue"],
            "total_debtors": totals["debtors"],
            "overdue_debtors": totals["overdue_debtors"],
            "total_advance": totals["advance"],
            "advance_clients": totals["advance_clients"],
            "segments": segments,
            "filters": filters,
            "clients": clients,
            "reps": reps,
            "active_filters": active_filters,
            "filter_count": len(active_filters),
            "has_filters": bool(active_filters),
            "filter_url": reverse("debt_list"),
            "export_url": reverse("debt_export") + (
                f"?{request.GET.urlencode()}" if request.GET.urlencode() else ""
            ),
        },
    )


def debt_export(request):
    """Excel (.xlsx) of the debtor list for the current filters — one row per client."""
    debtors, _, _ = _debtor_rows(request)
    today = timezone.localdate()
    headers = [
        "Mijoz", "Telefon", "Mas'ul xodim", "Ochiq cheklar",
        "Muddati o'tgan cheklar", "Eng yaqin muddat", "Kechikkan kun", "Holat",
        "Oxirgi yuk olgan", "Oxirgi to'lov", "Qarz qoldig'i", "Avans (so'm)",
        "Avans qachondan",
    ]
    rows = []
    for g in debtors:
        earliest = g["earliest"]
        # Counted off the SOONEST deadline, so the figure is how long the oldest
        # unpaid receipt has been sitting — the number people ask for. Left blank
        # rather than zeroed when nothing is late, so sorting by it puts the worst
        # debtors on top and everyone else out of the way.
        overdue_days = (today - earliest).days if earliest and earliest < today else None
        rows.append([
            g["client"].name,
            g["client"].phone,
            str(g["client"].owner) if g["client"].owner else "",
            g["count"],
            g["overdue_count"],
            earliest.strftime("%d.%m.%Y") if earliest else "",
            overdue_days if overdue_days is not None else "",
            # A row that is here only for its advance owes nothing, so it is neither
            # on time nor late.
            "Avans" if not g["remaining"]
            else "Muddati o'tgan" if earliest and earliest < today
            else "Muddatida",
            g["last_sale"].strftime("%d.%m.%Y") if g["last_sale"] else "",
            g["last_payment"].strftime("%d.%m.%Y") if g["last_payment"] else "",
            float(g["remaining"]),
            float(g["advance"]),
            g["advance_since"].strftime("%d.%m.%Y") if g["advance_since"] else "",
        ])
    return _xlsx_response(
        "qarzlar.xlsx", "Qarzlar", headers, rows, {11: "#,##0.00", 12: "#,##0.00"}
    )


def _open_receipts(request, client):
    """One debtor's open receipts, soonest deadline first."""
    return (
        Sale.objects.visible_to(request.user)
        .filter(client=client)
        .outstanding()
        .select_related("client", "sales_rep")
        .prefetch_related("items__product")
        .order_by("debt_deadline")
    )


def debt_client(request, pk):
    """A single debtor's open receipts, with per-receipt balance and deadline."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    sales = _open_receipts(request, client)
    total = sum((s.remaining for s in sales), Decimal("0"))
    # Every movement on this client's advance pool — deposits and money handed back —
    # each editable/voidable right here. The kassa ledger is not enough on its own: a
    # deposit taken in outside the till, or a return that never touched it, appears
    # nowhere on that page, and an entry with no screen to reach it cannot be undone.
    last_sale, last_payment = _client_activity(request.user, client)
    return render(
        request,
        "crm/debt_client.html",
        {
            "client": client,
            "sales": sales,
            "total": total,
            "last_sale": last_sale,
            "last_payment": last_payment,
            **_client_advance_context(request, client),
        },
    )


def debt_client_export(request, pk):
    """Excel (.xlsx) of one debtor's open receipts — the rows of their qarz page."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    today = timezone.localdate()
    headers = [
        "Sana", "Mahsulotlar", "Sotuvchi", "Umumiy", "To'langan",
        "Qoldiq", "Muddat", "Kechikkan kun", "Holat",
        "Oxirgi yuk olgan", "Oxirgi to'lov",
    ]
    # Client-level dates, so they repeat down every row: whoever opens the file reads
    # them off the first line instead of hunting for a summary block.
    last_sale, last_payment = _client_activity(request.user, client)
    rows = []
    for sale in _open_receipts(request, client):
        deadline = sale.debt_deadline
        overdue = deadline and deadline < today
        rows.append([
            sale.date.strftime("%d.%m.%Y"),
            sale.item_summary,
            str(sale.sales_rep),
            float(sale.total),
            float(sale.paid),
            float(sale.remaining),
            deadline.strftime("%d.%m.%Y") if deadline else "",
            # A number of its own rather than "12 kun o'tgan" buried in the status
            # text: the debtor list exports it the same way, and a figure can be
            # sorted and totalled where a sentence cannot.
            (today - deadline).days if overdue else "",
            "Muddati o'tgan" if overdue else "Muddatida",
            last_sale.strftime("%d.%m.%Y") if last_sale else "",
            last_payment.strftime("%d.%m.%Y") if last_payment else "",
        ])
    number_formats = {4: "#,##0.00", 5: "#,##0.00", 6: "#,##0.00"}
    return _xlsx_response(
        f"qarz-mijoz-{client.pk}.xlsx", "Ochiq cheklar", headers, rows, number_formats
    )


def _client_outstanding_fifo(request, client):
    """A client's open receipts ordered oldest debt first (FIFO)."""
    return list(
        Sale.objects.visible_to(request.user)
        .filter(client=client)
        .outstanding()
        .order_by("date", "created_at")
    )


def _slice_figures(credit, percent, payer, currency, exchange_rate):
    """One slice of a lump payment, as (so'm gross, bank fee, figure in its currency).

    `credit` is what this receipt's debt must fall by. Which sum the client actually
    transferred for that depends on who carries the fee:

      seller: the client sent exactly `credit`; the bank's cut is charged on top of
              it, against the seller, so gross == credit.
      client: the client had to send MORE so that `credit` survived the bank — the
              slice is grossed back up, and the difference is the recorded fee.

    The dollar figure (what the dollar till counts) is the gross at the payment's
    rate; a so'm payment's original is simply the so'm."""
    if payer == Payment.Payer.CLIENT and percent < Decimal("100"):
        gross = (credit / (Decimal("1") - percent / Decimal("100"))).quantize(
            Decimal("0.01"), ROUND_HALF_UP
        )
        commission = gross - credit
    else:
        gross = credit
        commission = (gross * percent / Decimal("100")).quantize(
            Decimal("0.01"), ROUND_HALF_UP
        )
    if currency == Payment.Currency.USD and exchange_rate:
        original = (gross / exchange_rate).quantize(Decimal("0.01"), ROUND_HALF_UP)
    else:
        original = gross
    return gross, commission, original


def _distribute_debt_payment(
    sales, amount, method, percent, note, user, currency=None, exchange_rate=Decimal("0"),
    on_date=None, client=None, payer=Payment.Payer.SELLER,
):
    """Spread a lump payment across FIFO-ordered debts, oldest first.

    `amount` is the gross the client handed over. How much of it pays down debt turns
    on `payer`: with the fee on the seller the whole sum counts, with it on the client
    only the net does and they still owe the bank's cut. Each receipt is credited its
    share up to its outstanding balance; the last one reached may receive a partial
    payment.

    A client may hand over MORE than they owe. Whatever is left once every receipt
    is settled becomes their advance (kredit) — an ADVANCE_IN deposit against
    `client`, which is real cash in the till that covers their next purchase.
    Without a `client` there is nobody to hold the credit, so the surplus is
    refused rather than silently dropped.

    Returns (receipts touched, surplus turned into advance).
    """
    is_transfer = method == Payment.Method.TRANSFER
    percent = percent if is_transfer else Decimal("0")
    payer = payer if (is_transfer and percent) else Payment.Payer.SELLER
    currency = currency or Payment.Currency.UZS
    on_date = on_date or timezone.localdate()  # backdated when an old debt is settled
    # What the client's debts can absorb in total — the fee is off the top when they
    # are the ones carrying it.
    if payer == Payment.Payer.CLIENT:
        fee_total = (amount * percent / Decimal("100")).quantize(
            Decimal("0.01"), ROUND_HALF_UP
        )
        left = amount - fee_total
    else:
        left = amount
    touched = 0

    def _create(credit, **extra):
        gross, commission, original = _slice_figures(
            credit, percent, payer, currency, exchange_rate
        )
        Payment.objects.create(
            amount=gross,
            amount_original=original,
            currency=currency,
            exchange_rate=exchange_rate,
            method=method,
            commission=commission,
            commission_percent=percent,
            commission_payer=payer,
            note=note,
            date=on_date,
            created_by=user,
            **extra,
        )

    with transaction.atomic():
        for sale in sales:
            if left <= 0:
                break
            due = sale.remaining or Decimal("0")
            chunk = min(left, due)
            if chunk <= 0:
                continue
            _create(chunk, sale=sale, kind=Payment.Kind.DEBT)
            left -= chunk
            touched += 1
        surplus = Decimal("0")
        if left > 0 and client is not None:
            surplus = left
            _create(left, sale=None, client=client, kind=Payment.Kind.ADVANCE_IN)
    return touched, surplus


def _apply_advance_to_open_sales(client, seller, on_date=None):
    """Spend a client's prepaid advance (seller-bound) on their open receipts, oldest
    first. Each slice becomes an ADVANCE_USED payment that settles part of a sale
    WITHOUT adding new till income — the cash already entered the till as the
    ADVANCE_IN deposit. Idempotent: a sale already covered contributes nothing, so it
    is safe to call after every sale AND after every fresh deposit. Only this seller's
    own sales to the client are touched (their till holds the money). Returns the
    total so'm applied."""
    balance = client_advance_balance(client, seller)
    if balance <= 0:
        return Decimal("0")
    on_date = on_date or timezone.localdate()
    sales = (
        Sale.objects.filter(client=client, sales_rep=seller)
        .with_balance()
        .filter(remaining__gt=0)
        .order_by("date", "created_at")
    )
    applied = Decimal("0")
    with transaction.atomic():
        for sale in sales:
            if balance <= 0:
                break
            due = sale.remaining or Decimal("0")
            use = min(balance, due)
            if use <= 0:
                continue
            Payment.objects.create(
                sale=sale,
                client=client,
                amount=use,
                amount_original=use,
                currency=Payment.Currency.UZS,
                method=Payment.Method.CASH,
                commission=Decimal("0"),
                commission_percent=Decimal("0"),
                note="Avansdan yechildi",
                kind=Payment.Kind.ADVANCE_USED,
                # never before the receipt it settles — a backdated advance still
                # can't have paid a sale that hadn't happened yet
                date=max(on_date, sale.date),
                created_by=seller,
            )
            balance -= use
            applied += use
    return applied


def _clean_amount(value):
    """Trim meaningless trailing zeros so a pre-filled amount reads '579300',
    not '579300,00000'. Quantity (3dp) × price (2dp) leaves up to 5 decimal
    places; real so'm never needs more than 2, and whole amounts need none."""
    value = value.quantize(Decimal("0.01"), ROUND_HALF_UP)
    if value == value.to_integral_value():
        return value.to_integral_value()
    return value.normalize()


def _usd_note(cleaned):
    """A ' · $100.00 × 12 700' suffix for audit/success lines on a dollar payment."""
    if cleaned.get("currency") == Payment.Currency.USD and cleaned.get("exchange_rate"):
        usd = cleaned["amount"] / cleaned["exchange_rate"]
        return f" · ${usd:,.2f} × {cleaned['exchange_rate']:,.0f}"
    return ""


def _method_label(code):
    """The Uzbek display name for a payment-method code (naqd/karta/o'tkazma)."""
    return dict(Payment.Method.choices).get(code, code)


def _kg(value):
    """A kg amount without trailing decimal zeros — '23.000' → '23', '23.5' → '23.5'."""
    return ("{:,.3f}".format(value)).rstrip("0").rstrip(".")


def _render_client_pay(request, client, total, form, invalid=False, debts=None):
    context = {
        "form": form,
        "client": client,
        "remaining": total,
        # The open receipts in FIFO order, so the seller sees WHEN each debt was
        # taken before choosing the payment date.
        "debts": debts or [],
        "title": f"Umumiy to'lov: {client.name}",
    }
    if is_ajax(request):
        return render(
            request, "crm/_client_pay_modal.html", context, status=422 if invalid else 200
        )
    return render(request, "crm/_client_pay_page.html", context)


def client_debt_pay(request, pk):
    """Take one amount and pay down the client's debts oldest-first (FIFO)."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    sales = _client_outstanding_fifo(request, client)
    total = sum((s.remaining for s in sales), Decimal("0")).quantize(
        Decimal("0.01"), ROUND_HALF_UP
    )
    if total <= 0:
        return form_reload(request, reverse("debt_client", args=[client.pk]))
    if request.method == "POST":
        form = DebtPaymentForm(request.POST)
        if form.is_valid():
            touched, surplus = _distribute_debt_payment(
                sales,
                form.cleaned_data["amount"],
                form.cleaned_data["method"],
                form.cleaned_data["commission_percent"],
                form.cleaned_data["note"],
                request.user,
                currency=form.cleaned_data["currency"],
                exchange_rate=form.cleaned_data["exchange_rate"],
                on_date=form.cleaned_data["date"],
                client=client,
                payer=form.cleaned_data["commission_payer"],
            )
            AuditLog.record(
                request.user, AuditLog.Action.PAYMENT, "To'lov", client.pk,
                f"Mijoz {client.name} qarz to'lovi "
                f"({_method_label(form.cleaned_data['method'])}){_usd_note(form.cleaned_data)} "
                f"— {form.cleaned_data['amount']:,.0f} so'm",
            )
            msg = f"{form.cleaned_data['amount']:,.0f} so'm {touched} ta chekka taqsimlandi."
            if surplus > 0:
                msg += f" Ortiqcha {surplus:,.0f} so'm avans balansiga qo'shildi."
            messages.success(request, msg)
            return form_reload(request, reverse("debt_client", args=[client.pk]))
        return _render_client_pay(request, client, total, form, invalid=True, debts=sales)
    form = DebtPaymentForm(
        initial={
            "amount": _clean_amount(total),
            "method": Payment.Method.CASH,
            "date": timezone.localdate(),
        },
    )
    return _render_client_pay(request, client, total, form, debts=sales)


def _render_client_advance(request, client, balance, form, invalid=False):
    # The earlier deposits ride along under the form: the moment someone is about to
    # write a new one is exactly when a wrong old one is spotted ("it's already in,
    # just wrong"), and without them on screen the fix is a new deposit on top.
    context = {
        "form": form,
        "client": client,
        "advance_balance": balance,
        "title": f"Avans qabul qilish: {client.name}",
        **_client_advance_context(request, client),
    }
    if is_ajax(request):
        return render(
            request, "crm/_client_advance_modal.html", context,
            status=422 if invalid else 200,
        )
    return render(request, "crm/_client_advance_page.html", context)


def client_advance_pay(request, pk):
    """Take an advance (oldindan to'lov) from a client into the seller's till.

    The cash normally enters the kassa now (ADVANCE_IN) — it is real income the moment
    it's received. It is then spent oldest-debt-first on the client's open receipts;
    any surplus stays as their advance balance to cover future sales (it is NOT
    refunded). Advance is seller-bound: it sits in the till of whoever took it.

    The form's one extra question is whether the cash is arriving now. Answer no —
    money taken in long ago and only being written up today, which is most of what the
    old sverkas turn out to hold — and the deposit is recorded exactly the same but
    kept out of the till, so today's kirim isn't inflated by cash nobody handed over
    (see `AdvanceForm`)."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    balance = client_advance_balance(client, request.user)
    if request.method == "POST":
        form = AdvanceForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            Payment.objects.create(
                client=client,
                sale=None,
                amount=cd["amount"],
                amount_original=cd["amount_original"],
                currency=cd["currency"],
                exchange_rate=cd["exchange_rate"],
                method=cd["method"],
                commission=cd["commission"],
                commission_percent=cd["commission_percent"],
                commission_payer=cd["commission_payer"],
                note=cd["note"],
                kind=Payment.Kind.ADVANCE_IN,
                is_opening=cd["is_opening"],
                date=cd["date"],
                created_by=request.user,
            )
            applied = _apply_advance_to_open_sales(client, request.user, on_date=cd["date"])
            AuditLog.record(
                request.user, AuditLog.Action.PAYMENT, "To'lov", client.pk,
                f"Mijoz {client.name} avans to'lovi "
                f"({_method_label(cd['method'])}){_usd_note(cd)} "
                f"— {cd['amount']:,.0f} so'm{_kassa_note(cd['is_opening'])}",
            )
            left = client_advance_balance(client, request.user)
            msg = f"Avans qabul qilindi: {cd['amount']:,.0f} so'm."
            if cd["is_opening"]:
                msg += " Kassaga kirim qilinmadi."
            if applied > 0:
                msg += f" {applied:,.0f} so'm ochiq qarzga taqsimlandi."
            if left > 0:
                msg += f" Balansda: {left:,.0f} so'm."
            messages.success(request, msg)
            return form_reload(request, reverse("client_list"))
        return _render_client_advance(request, client, balance, form, invalid=True)
    form = AdvanceForm(initial={
        "method": Payment.Method.CASH, "date": timezone.localdate(),
    })
    return _render_client_advance(request, client, balance, form)


def _kassa_note(is_opening):
    """The audit-log tail that says whether an advance touched the till. Silent when
    it did — that is the ordinary case and needs no remark."""
    return " (kassaga kirim qilinmadi)" if is_opening else ""


def _advance_in_qs(user):
    """Advance deposits this user is allowed to touch — their own, or all for an
    admin/manager."""
    qs = Payment.objects.filter(kind=Payment.Kind.ADVANCE_IN).select_related("client")
    if not user.can_see_all_records:
        qs = qs.filter(created_by=user)
    return qs


def _advance_move_qs(user):
    """Both sides of the advance pool — deposits and money handed back — under the
    same visibility rule."""
    qs = Payment.objects.filter(
        kind__in=(Payment.Kind.ADVANCE_IN, Payment.Kind.ADVANCE_OUT)
    ).select_related("client")
    if not user.can_see_all_records:
        qs = qs.filter(created_by=user)
    return qs


def _client_advance_context(request, client):
    """The advance block shown on any screen that offers to fix one: the balance and
    every movement behind it, under the caller's own visibility."""
    scope = None if request.user.can_see_all_records else request.user
    return {
        "advance": client_advance_balance(client, scope),
        "advance_moves": (
            _advance_move_qs(request.user)
            .filter(client=client)
            # A correction's difference is written as an ordinary deposit or advance
            # return, so the row alone can't say it apart — and money leaving on a
            # correction must not be read as the client taking their cash back.
            .annotate(is_adjust=Case(
                When(note__startswith=ADVANCE_ADJUST_NOTE, then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            ))
            .order_by("-date", "-created_at")
        ),
    }


def client_advance_moves(request, pk):
    """One client's advance movements, opened straight from the Mijozlar list.

    The rows and their buttons already existed on the client's qarz card, but getting
    there to fix a mistyped figure meant leaving the list, and for a client with no
    open receipts that card is mostly empty — a long walk to reach two icons. A wrong
    advance is noticed while scanning the list, so the fix belongs where it is noticed.

    Nothing is decided here: the same `advance_edit` / `advance_delete` views do the
    work, loaded into the same dialog. This screen only makes them reachable, which is
    why it takes no POST."""
    client = get_object_or_404(_visible_clients(request.user), pk=pk)
    context = {
        "client": client,
        "title": f"Avans harakatlari: {client.name}",
        **_client_advance_context(request, client),
    }
    if is_ajax(request):
        return render(request, "crm/_advance_moves_modal.html", context)
    return render(request, "crm/_advance_moves_page.html", context)


def _reconcile_client_advance(client, seller):
    """Bring a client's advance (for one seller) back into balance after a deposit
    was changed or removed. If deposits have shrunk below what sales already drew
    (balance < 0), the newest ADVANCE_USED allocations are peeled back — reverting
    those sales to debt — until the pool is non-negative. Then any advance still left
    is re-applied to open receipts, oldest first. Idempotent."""
    with transaction.atomic():
        balance = client_advance_balance(client, seller)
        if balance < 0:
            used = Payment.objects.filter(
                client=client, created_by=seller, kind=Payment.Kind.ADVANCE_USED
            ).order_by("-date", "-created_at")
            for u in used:
                if balance >= 0:
                    break
                balance += u.credited_amount  # freeing this returns it to the pool
                u.delete()               # ...and the sale it covered owes again
        _apply_advance_to_open_sales(client, seller)


def advance_edit(request, pk):
    """Fix a mistaken advance deposit (amount / method / note / whether it hit the
    till). If the new amount is smaller than what sales already drew, the excess is
    clawed back automatically (those sales revert to debt) — see
    `_reconcile_client_advance`.

    The till question is editable after the fact on purpose: a deposit booked as a
    kirim which turns out to be money collected weeks ago can be taken back out of the
    day's income here, instead of the seller ringing someone to fix the kassa by hand.
    The client's credit is untouched either way — only where the cash is said to be
    changes.

    A changed amount asks one more question, because rewriting the figure in place also
    rewrites the kirim of the day it was taken. That is right for a typo from an hour
    ago and wrong for a deposit from three weeks back, so `AdvanceEditForm.diff_mode`
    decides: correct the old day, or leave it alone and write the difference as its own
    row dated today (a deposit if the figure went up, an advance return if it went
    down), with or without touching the till. The client's credit ends up the same
    either way; only the day the money is said to have moved differs.

    The difference row is a plain so'm figure and carries no bank fee: a fee that also
    needs correcting belongs to the original deposit, so that case goes the RETRO
    route."""
    payment = get_object_or_404(_advance_in_qs(request.user), pk=pk)
    client, seller = payment.client, payment.created_by
    if request.method == "POST":
        form = AdvanceEditForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            mode = cd["diff_mode"]
            was = payment.amount
            delta = cd["amount"] - was
            retro = mode == AdvanceEditForm.RETRO or not delta
            payment.is_opening = cd["is_opening"]
            payment.method = cd["method"]
            payment.note = cd["note"]
            payment.date = cd["date"]
            # Off the RETRO route the deposit keeps the figure it was written with —
            # its day stays exactly as it was counted — and only the difference moves.
            if retro:
                payment.amount = cd["amount"]
                payment.amount_original = cd["amount_original"]
                payment.currency = cd["currency"]
                payment.exchange_rate = cd["exchange_rate"]
                payment.commission = cd["commission"]
                payment.commission_percent = cd["commission_percent"]
                payment.commission_payer = cd["commission_payer"]
            payment.save()
            if not retro:
                _write_advance_difference(payment, delta, mode, was, cd["amount"])
            # Re-apply a bigger deposit, or claw back a smaller one, then settle debts.
            _reconcile_client_advance(client, seller)
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "To'lov", client.pk,
                f"Mijoz {client.name} avansi o'zgartirildi — "
                f"{was:,.0f} → {cd['amount']:,.0f} so'm"
                f"{'' if retro else _diff_note(mode)}"
                f"{_kassa_note(payment.is_opening)}",
            )
            messages.success(request, _advance_edit_message(retro, delta, mode))
            # A deposit kept out of the till isn't on the kassa page at all, so the
            # client's own card is where it goes back to being visible.
            fallback = (
                reverse("debt_client", args=[client.pk]) if payment.is_opening
                else reverse("kassa")
            )
            return form_reload(request, fallback)
        return _render_advance_edit(request, payment, form, invalid=True)
    form = AdvanceEditForm(initial={
        "date": payment.date,
        "amount": _clean_amount(payment.original_amount),
        "method": payment.method,
        "currency": payment.currency,
        "exchange_rate": payment.exchange_rate or "",
        "commission_percent": payment.commission_percent or "",
        "commission_payer": payment.commission_payer,
        "note": payment.note,
        "to_kassa": (
            AdvanceForm.OUT_OF_KASSA if payment.is_opening else AdvanceForm.IN_KASSA
        ),
    })
    return _render_advance_edit(request, payment, form)


def _write_advance_difference(payment, delta, mode, was, now):
    """Book the change in a deposit's figure as its own dated row, leaving the deposit
    (and the day it was counted on) untouched.

    Money-wise this is nothing new: more money is a deposit, less money is an advance
    return, and every till and balance figure already knows how to read both. What the
    row carries that they don't need is why it exists — the note says so, so a
    correction is never read off the client's card as "they took their money back"."""
    Payment.objects.create(
        client=payment.client,
        sale=None,
        amount=abs(delta),
        amount_original=abs(delta),
        method=payment.method,
        note=f"{ADVANCE_ADJUST_NOTE}: {was:,.0f} → {now:,.0f}",
        kind=(
            Payment.Kind.ADVANCE_IN if delta > 0 else Payment.Kind.ADVANCE_OUT
        ),
        is_opening=mode == AdvanceEditForm.NO_KASSA,
        date=timezone.localdate(),
        created_by=payment.created_by,
    )


def _diff_note(mode):
    """The audit-log tail naming where a correction's difference was put."""
    if mode == AdvanceEditForm.NO_KASSA:
        return " (farq alohida yozildi, kassaga tegmadi)"
    return " (farq bugungi kassaga yozildi)"


def _advance_edit_message(retro, delta, mode):
    if retro:
        return "Avans yangilandi."
    where = "kassaga tegilmadi" if mode == AdvanceEditForm.NO_KASSA else (
        "kassaga kirim qilindi" if delta > 0 else "kassadan chiqim qilindi"
    )
    direction = "qo'shildi" if delta > 0 else "ayirildi"
    return (
        f"Avans tuzatildi: {abs(delta):,.0f} so'm {direction} — bugungi sanada "
        f"alohida qator yozildi, {where}. Eski kun o'zgarmadi."
    )


def _render_advance_edit(request, payment, form, invalid=False):
    balance = client_advance_balance(payment.client, payment.created_by)
    context = {
        "form": form,
        "client": payment.client,
        "advance_balance": balance,
        "title": f"Avansni tahrirlash: {payment.client.name}",
    }
    if is_ajax(request):
        return render(
            request, "crm/_advance_edit_modal.html", context,
            status=422 if invalid else 200,
        )
    return render(request, "crm/_client_advance_page.html", context)


def advance_delete(request, pk):
    """Take an advance deposit off a client — and ask which of the two things that
    actually means, because they are not the same event.

    "Xato yozilgan" removes the record outright. If sales had already drawn on it those
    allocations are peeled back and the sales revert to debt (see
    `_reconcile_client_advance`), so the money trail stays consistent. The till is left
    as though the deposit had never been written.

    "Pul qaytarildi" is the opposite: the deposit stands, because the cash really did
    come in on its own date, and the money going back out is written as its own
    ADVANCE_OUT row on the day it left. That way a day already counted and reconciled
    doesn't silently change its kirim weeks later because a client asked for their
    credit back. Only credit the client still holds can be handed over — the form caps
    it at their balance, so settled sales are never dragged back into debt to fund a
    refund (see `AdvanceRemoveForm`)."""
    payment = get_object_or_404(_advance_in_qs(request.user), pk=pk)
    client, seller = payment.client, payment.created_by
    balance = client_advance_balance(client, seller)
    if request.method == "POST":
        form = AdvanceRemoveForm(
            request.POST, deposit_amount=payment.credited_amount, balance=balance
        )
        if form.is_valid():
            cd = form.cleaned_data
            if cd["mode"] == AdvanceRemoveForm.CASH_OUT:
                Payment.objects.create(
                    client=client,
                    sale=None,
                    amount=cd["amount"],
                    amount_original=cd["amount"],
                    method=payment.method,
                    note=cd["note"] or "Avans mijozga qaytarildi",
                    kind=Payment.Kind.ADVANCE_OUT,
                    is_opening=cd["is_opening"],
                    date=cd["date"],
                    created_by=seller,
                )
                where = (
                    "kassaga tegmadi" if cd["is_opening"] else "kassadan chiqim"
                )
                AuditLog.record(
                    request.user, AuditLog.Action.PAYMENT, "To'lov", client.pk,
                    f"Mijoz {client.name} avansi qaytarildi — "
                    f"{cd['amount']:,.0f} so'm ({where})",
                )
                messages.success(
                    request,
                    f"Avans qaytarildi: {cd['amount']:,.0f} so'm — "
                    f"{'kassaga tegilmadi' if cd['is_opening'] else 'kassadan chiqim qilindi'}.",
                )
            else:
                AuditLog.record(
                    request.user, AuditLog.Action.VOID, "To'lov", client.pk,
                    f"{client.name} — avans {payment.amount:,.0f} so'm o'chirildi "
                    f"(xato yozuv)",
                )
                payment.delete()
                messages.success(request, "Avans o'chirildi.")
            # Either route can leave the pool out of step — a deletion may have taken
            # away credit sales were already living on, a cash return shrinks it — so
            # both settle through the same reconcile.
            _reconcile_client_advance(client, seller)
            return form_reload(request, reverse("debt_client", args=[client.pk]))
        return _render_advance_remove(request, payment, balance, form, invalid=True)
    # Offer back what this deposit put in, but never more than the client still holds:
    # the rest is already sitting on their receipts. Nothing left means nothing to
    # pre-fill, and the form says so if that route is picked anyway.
    returnable = min(payment.credited_amount, balance)
    form = AdvanceRemoveForm(
        deposit_amount=payment.credited_amount,
        balance=balance,
        initial={
            "amount": _clean_amount(returnable) if returnable > 0 else "",
            "date": timezone.localdate(),
        },
    )
    return _render_advance_remove(request, payment, balance, form)


def _render_advance_remove(request, payment, balance, form, invalid=False):
    context = {
        "form": form,
        "payment": payment,
        "client": payment.client,
        "advance_balance": balance,
        # What sales have already eaten. Named on the page because it is the number
        # that explains why a cash return may be capped below the deposit.
        "spent": max(payment.credited_amount - balance, Decimal("0")),
        "title": f"Avansni o'chirish: {payment.client.name}",
    }
    if is_ajax(request):
        return render(
            request, "crm/_advance_remove_modal.html", context,
            status=422 if invalid else 200,
        )
    return render(request, "crm/_advance_remove_page.html", context)


def advance_out_delete(request, pk):
    """Undo an advance return — the client's credit comes back and the till stops
    showing the money as gone. The way out of a mis-clicked refund, and the reason the
    cash-out route is safe to offer at all."""
    qs = Payment.objects.filter(kind=Payment.Kind.ADVANCE_OUT).select_related("client")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    payment = get_object_or_404(qs, pk=pk)
    client, seller = payment.client, payment.created_by
    if request.method == "POST":
        summary = f"{client.name} — qaytarilgan avans {payment.amount:,.0f} so'm bekor qilindi"
        payment.delete()
        # The freed credit lands back on whatever the client still owes.
        _reconcile_client_advance(client, seller)
        AuditLog.record(request.user, AuditLog.Action.VOID, "To'lov", client.pk, summary)
        messages.success(request, "Avans qaytarilishi bekor qilindi.")
        return form_reload(request, reverse("kassa"))
    return render_confirm(
        request,
        "Avans qaytarilishini bekor qilish",
        f"“{client.name}” — {payment.amount:,.0f} so'm avans qaytarilishi bekor "
        f"qilinadi: pul kassadan chiqmagan hisoblanadi va mijozning avansi tiklanadi. "
        f"Davom etasizmi?",
        "Ha, bekor qilish",
        confirm_class="btn-danger",
    )


# Advance movements are money on the client's credit pool, not on a receipt, so the
# ordinary payment edit/void refuses them: each has its own view, which also puts the
# pool back in balance afterwards.
ADVANCE_KINDS = (
    Payment.Kind.ADVANCE_IN, Payment.Kind.ADVANCE_USED, Payment.Kind.ADVANCE_OUT,
)


def payment_delete(request, pk):
    """Void a mistaken payment by removing it. The debt it covered is restored
    automatically (remaining is derived). Admins/managers may void any payment;
    a seller may void only payments they took in themselves."""
    qs = Payment.objects.select_related("sale", "sale__client")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    payment = get_object_or_404(qs, pk=pk)
    if payment.kind in ADVANCE_KINDS:
        messages.error(request, "Avans to'lovi bu yerdan o'chirilmaydi.")
        return form_reload(request, reverse("kassa"))
    if request.method == "POST":
        sale_pk = payment.sale_id
        summary = f"{payment.sale.client.name} — {payment.amount:,.0f} so'm ({payment.get_method_display()})"
        payment.delete()
        AuditLog.record(request.user, AuditLog.Action.VOID, "To'lov", sale_pk, summary)
        messages.success(request, "To'lov o'chirildi — qarz qayta tiklandi.")
        return form_reload(request, reverse("sale_detail", args=[sale_pk]))
    return render_confirm(
        request,
        "To'lovni bekor qilish",
        f"“{payment.sale.client.name}” — {payment.amount:,.0f} so'm to'lov "
        f"({payment.get_method_display()}) o'chiriladi va qarz qayta tiklanadi. "
        f"Davom etasizmi?",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


def payment_edit(request, pk):
    """Fix a mistaken payment (Kirim) — amount, currency, method, commission, note.
    The sale's remaining debt re-derives from the new amount automatically. Admins/
    managers may edit any payment; a seller may edit only payments they took in
    themselves. The amount is capped so the sale can't become over-paid."""
    qs = Payment.objects.select_related("sale", "sale__client")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    payment = get_object_or_404(qs, pk=pk)
    if payment.kind in ADVANCE_KINDS:
        messages.error(request, "Avans to'lovi bu yerdan tahrirlanmaydi.")
        return form_reload(request, reverse("kassa"))
    # This receipt's ceiling: the sale's remaining already excludes what this payment
    # credited, so add that back to get how much this one may cover.
    max_amount = payment.sale.debt_remaining + payment.credited_amount
    title = "To'lovni tahrirlash"
    form = PaymentEditForm(request.POST or None, instance=payment, max_amount=max_amount)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "To'lov", payment.sale_id,
                f"Mijoz {payment.sale.client.name} to'lovi "
                f"({payment.get_method_display()}){_usd_note(form.cleaned_data)} "
                f"— {payment.amount:,.0f} so'm",
            )
            messages.success(request, "To'lov yangilandi.")
            return form_success(request, reverse("kassa"))
        return form_response(
            request, form, title, invalid=True,
            modal_template="crm/_payment_edit_modal.html",
        )
    return form_response(
        request, form, title, modal_template="crm/_payment_edit_modal.html"
    )


# --- Mijozlar to'lovlari (client payments) ------------------------------------

# What this page counts as "the client paid us": money that actually changed hands.
# `is_opening` advances are dropped — they carry a balance from before the CRM (or
# cash taken into the drawer earlier), not money handed over on their date.
RECEIVED_PAYMENT_KINDS = (
    Payment.Kind.SALE,
    Payment.Kind.DEBT,
    Payment.Kind.ADVANCE_IN,
)

# Listed, but never counted. Spending held credit on a receipt moves no new money:
# those so'm already arrived — and were already shown — as the ADVANCE_IN deposit,
# so adding them to a total would report money the client never handed over. The row
# still earns its place, because without it a receipt looks unpaid on the page that
# is supposed to say how it got paid.
CLIENT_PAYMENT_KINDS = RECEIVED_PAYMENT_KINDS + (Payment.Kind.ADVANCE_USED,)

# Every money figure on this page carries this filter, so an advance being spent can
# never leak into a sum.
RECEIVED = ~Q(kind=Payment.Kind.ADVANCE_USED)

# How each kind reads in the list: (label, badge class).
_CLIENT_PAYMENT_LABELS = {
    Payment.Kind.SALE: ("Sotuvda to'landi", "badge-neutral"),
    Payment.Kind.DEBT: ("Qarz to'lovi", "badge-ok"),
    Payment.Kind.ADVANCE_IN: ("Avans (oldindan)", "badge-info"),
    Payment.Kind.ADVANCE_USED: ("Avansdan yechildi", "badge-shipped"),
}


def _payment_scope(user):
    """The client payments a user may see: everything for an admin/manager, and for
    a seller the payments on their own sales plus the advances they took in
    themselves — the same rule the client history timeline uses.

    `client_pk` is annotated because the client hangs off two different places: a
    per-sale payment reaches it through the sale, an advance deposit carries it
    directly. Grouping and filtering by client need the one column."""
    qs = (
        Payment.objects.filter(kind__in=CLIENT_PAYMENT_KINDS, is_opening=False)
        .select_related("sale", "sale__client", "client", "created_by")
        .annotate(client_pk=Coalesce("sale__client", "client"))
    )
    if user.can_see_all_records:
        return qs
    return qs.filter(Q(sale__sales_rep=user) | Q(sale__isnull=True, created_by=user))


def _filter_payments(request):
    """Client payments for the current filters, newest first.

    Dates work exactly as they do on the sales list: the dan..gacha window (today by
    default) is what the unfiltered page shows, but as soon as a content filter is
    set — a client, a method, a search term — the window steps aside and the search
    runs over every date. Picking a client from the drawer is how you ask "what has
    this person ever paid us", and answering for today alone would make it useless.

    Returns (queryset, filters, has_filters)."""
    qs = _payment_scope(request.user)
    keys = ("client", "rep", "method", "currency", "kind")
    filters = {key: request.GET.get(key, "") for key in keys}
    filters["q"] = request.GET.get("q", "").strip()

    kinds = {str(k) for k in CLIENT_PAYMENT_KINDS}
    if filters["kind"] not in kinds:
        filters["kind"] = ""
    if filters["method"] not in dict(Payment.Method.choices):
        filters["method"] = ""
    if filters["currency"] not in dict(Payment.Currency.choices):
        filters["currency"] = ""

    has_filters = bool(
        filters["q"]
        or filters["client"].isdigit()
        or filters["kind"]
        or filters["method"]
        or filters["currency"]
        or (filters["rep"].isdigit() and request.user.can_see_all_records)
    )

    dates = _date_range_context(request)
    filters["dan"] = dates["date_from"].isoformat()
    filters["gacha"] = dates["date_to"].isoformat()
    if not has_filters:
        qs = qs.filter(date__gte=dates["date_from"], date__lte=dates["date_to"])

    if filters["q"]:
        # The client sits behind two different paths, so match the clients first and
        # then filter the payments by the pks that came back.
        pks = _client_search(_visible_clients(request.user), filters["q"]).values("pk")
        qs = qs.filter(client_pk__in=pks)
    if filters["client"].isdigit():
        qs = qs.filter(client_pk=filters["client"])
    if filters["kind"]:
        qs = qs.filter(kind=filters["kind"])
    if filters["method"]:
        qs = qs.filter(method=filters["method"])
    if filters["currency"]:
        qs = qs.filter(currency=filters["currency"])
    if filters["rep"].isdigit() and request.user.can_see_all_records:
        qs = qs.filter(created_by_id=filters["rep"])
    return qs.order_by("-date", "-created_at"), filters, has_filters


# Sums the KPI cards and the per-client rows are built from. `amount` is the so'm
# value the client handed over — gross, like the kassa kirim ledger: the bank's cut is
# shown beside it rather than quietly netted off. Everything here is scoped to
# RECEIVED; `from_advance` is the one figure that counts the other side, and it is
# reported apart from the totals precisely so it can't be mistaken for income.
_PAYMENT_AGGREGATES = {
    "total": Sum("amount", filter=RECEIVED),
    "cash": Sum("amount", filter=RECEIVED & Q(method=Payment.Method.CASH)),
    "card": Sum("amount", filter=RECEIVED & Q(method=Payment.Method.CARD)),
    "transfer": Sum("amount", filter=RECEIVED & Q(method=Payment.Method.TRANSFER)),
    "commission": Sum("commission", filter=RECEIVED),
    "count": Count("pk", filter=RECEIVED),
    "from_advance": Sum("amount", filter=Q(kind=Payment.Kind.ADVANCE_USED)),
}


def _payment_totals(payments):
    """Period totals for the payments shown, with the method split. Empty sums come
    back as None from the database; they are zeroed here so the template can do
    arithmetic on them."""
    totals = payments.aggregate(**_PAYMENT_AGGREGATES)
    for key, value in totals.items():
        if value is None:
            totals[key] = Decimal("0")
    # Clients who actually paid — a client whose only row is an advance being spent
    # handed nothing over in this window. `client_pk` is an annotation, so the default
    # -date ordering has to go: Django would otherwise carry the ordering columns into
    # the SELECT and make every row distinct on its own.
    totals["clients"] = (
        payments.filter(RECEIVED).order_by().values("client_pk").distinct().count()
    )
    total = totals["total"] or Decimal("0")
    for key in ("cash", "card", "transfer"):
        totals[f"{key}_pct"] = (totals[key] / total * 100) if total else 0
    return totals


def _payment_rows(payments):
    """One display row per payment. The client is resolved here (a sale payment
    reaches it through the sale, an advance carries it directly) so the template
    doesn't have to ask twice on every line."""
    rows = []
    for p in payments:
        client = p.sale.client if p.sale else p.client
        label, badge = _CLIENT_PAYMENT_LABELS[p.kind]
        rows.append({
            "pk": p.pk,
            "date": p.date,
            "client": client,
            "label": label,
            "badge": badge,
            "is_advance": p.kind == Payment.Kind.ADVANCE_IN,
            # False on an advance being spent: the row is shown, but no total counts
            # it, and the page says so on the line itself.
            "counted": p.kind != Payment.Kind.ADVANCE_USED,
            "sale_pk": p.sale_id,
            "method": p.get_method_display(),
            "method_code": p.method,
            "currency": p.currency,
            "amount": p.amount,
            "amount_original": p.original_amount,
            "exchange_rate": p.exchange_rate,
            "commission": p.commission,
            # The reconciliation stamps its rows with a note that repeats the label
            # word for word ("Avansdan yechildi"); printing both just says it twice.
            "note": "" if p.note == label else p.note,
            "created_by": p.created_by,
        })
    return rows


def _payment_client_rows(payments):
    """One row per client for the same payments: what they paid in the period, split
    by method, with how many payments and when the last one landed. Biggest payer
    first — the question this view answers is who has been paying.

    `from_advance` sits outside the total, as everywhere else on this page. `last` is
    the last payment they MADE, so an advance being spent does not pass for one."""
    agg = (
        payments.order_by()
        .values("client_pk")
        .annotate(**_PAYMENT_AGGREGATES, last=Max("date", filter=RECEIVED))
        .order_by("-total")
    )
    agg = list(agg)
    clients = Client.objects.in_bulk([r["client_pk"] for r in agg if r["client_pk"]])
    rows = []
    for r in agg:
        client = clients.get(r["client_pk"])
        if client is None:  # a payment whose client was deleted — nothing to show
            continue
        rows.append({
            "client": client,
            "total": r["total"] or Decimal("0"),
            "cash": r["cash"] or Decimal("0"),
            "card": r["card"] or Decimal("0"),
            "transfer": r["transfer"] or Decimal("0"),
            "from_advance": r["from_advance"] or Decimal("0"),
            "count": r["count"],
            "last": r["last"],
        })
    return rows


def _payment_filter_chips(request, filters, clients, reps):
    kind_labels = dict(Payment.Kind.choices)
    method_labels = dict(Payment.Method.choices)
    currency_labels = dict(Payment.Currency.choices)
    client = (
        clients.filter(pk=filters["client"]).first() if filters["client"].isdigit() else None
    )
    rep = reps.filter(pk=filters["rep"]).first() if reps and filters["rep"].isdigit() else None
    return _filter_chips(request, [
        {"param": "client", "label": "Mijoz", "value": client.name if client else ""},
        {"param": "rep", "label": "Qabul qildi", "value": str(rep) if rep else ""},
        {"param": "kind", "label": "Turi", "value": kind_labels.get(filters["kind"], "")},
        {"param": "method", "label": "Usul", "value": method_labels.get(filters["method"], "")},
        {"param": "currency", "label": "Valyuta",
         "value": currency_labels.get(filters["currency"], "")},
    ])


def payment_list(request):
    """Mijozlar to'lovlari — the money clients have handed over: every payment on a
    sale, every debt repayment and every advance deposit.

    Two ways to read the same set: `?korinish=mijoz` groups it by client (who paid
    how much), anything else lists the payments themselves. Both share one filter
    set, so a chosen client, method or period carries between them."""
    payments, filters, has_filters = _filter_payments(request)
    by_client = request.GET.get("korinish") == "mijoz"
    # Carried in `filters` so the search box and the filter drawer (plain GET forms)
    # keep the chosen view when they submit.
    filters["korinish"] = "mijoz" if by_client else ""
    totals = _payment_totals(payments)

    clients = _visible_clients(request.user).order_by("name")
    reps = (
        User.objects.filter(is_active=True).order_by("first_name", "username")
        if request.user.can_see_all_records
        else None
    )
    active_filters = _payment_filter_chips(request, filters, clients, reps)
    if by_client:
        page = Paginator(_payment_client_rows(payments), 25).get_page(
            request.GET.get("page")
        )
    else:
        page = Paginator(payments, 25).get_page(request.GET.get("page"))
        page.object_list = _payment_rows(page.object_list)
    export_qs = request.GET.urlencode()
    return render(request, "crm/payment_list.html", {
        "page": page,
        "by_client": by_client,
        "totals": totals,
        "filters": filters,
        "clients": clients,
        "reps": reps,
        "rep_label": "Kim qabul qildi",
        "payment_kinds": [
            (str(k), _CLIENT_PAYMENT_LABELS[k][0]) for k in CLIENT_PAYMENT_KINDS
        ],
        # Which of the two readings to show. It lives in the filter drawer with
        # everything else that shapes the page, rather than as its own control.
        "view_options": [("", "To'lovlar"), ("mijoz", "Mijozlar bo'yicha")],
        "active_filters": active_filters,
        "filter_count": len(active_filters),
        "has_filters": has_filters,
        "filter_url": reverse("payment_list"),
        "search_placeholder": "Mijoz ismi yoki telefoni bo'yicha qidirish…",
        "export_url": reverse("payment_export") + (f"?{export_qs}" if export_qs else ""),
        **_date_range_context(request),
    })


def payment_export(request):
    """Excel (.xlsx) of the client payments for the current filters — two tabs: every
    payment, and the same money totalled per client.

    Money received and money drawn from an advance go in SEPARATE columns, so that
    selecting the Summa column in Excel gives the same figure the page shows. Putting
    both in one column would hand whoever opens the file a total the CRM never claims."""
    payments, _, _ = _filter_payments(request)
    detail_headers = [
        "Sana", "Mijoz", "Turi", "Chek", "Usul", "Valyuta", "Asl summa", "Kurs",
        "Summa (so'm)", "Avansdan (jamiga kirmaydi)", "Bank komissiyasi",
        "Kim qabul qildi", "Izoh",
    ]
    detail_rows = [
        [
            r["date"].strftime("%d.%m.%Y"),
            r["client"].name if r["client"] else "",
            r["label"],
            f"#{r['sale_pk']}" if r["sale_pk"] else "",
            r["method"],
            dict(Payment.Currency.choices).get(r["currency"], ""),
            float(r["amount_original"]),
            float(r["exchange_rate"]) if r["currency"] == Payment.Currency.USD else "",
            float(r["amount"]) if r["counted"] else "",
            "" if r["counted"] else float(r["amount"]),
            float(r["commission"]),
            str(r["created_by"]),
            r["note"],
        ]
        for r in _payment_rows(payments)
    ]
    client_headers = [
        "Mijoz", "Telefon", "To'lovlar soni", "Oxirgi to'lov",
        "Naqd", "Karta", "O'tkazma", "Jami to'langan",
        "Avansdan yopilgan (jamiga kirmaydi)",
    ]
    client_rows = [
        [
            r["client"].name,
            r["client"].phone,
            r["count"],
            r["last"].strftime("%d.%m.%Y") if r["last"] else "",
            float(r["cash"]),
            float(r["card"]),
            float(r["transfer"]),
            float(r["total"]),
            float(r["from_advance"]),
        ]
        for r in _payment_client_rows(payments)
    ]
    money = "#,##0.00"
    return _xlsx_book_response("mijozlar-tolovlari.xlsx", [
        ("To'lovlar", detail_headers, detail_rows,
         {7: money, 9: money, 10: money, 11: money}),
        ("Mijozlar bo'yicha", client_headers, client_rows,
         {5: money, 6: money, 7: money, 8: money, 9: money}),
    ])


def _audit_rows(request):
    """The audit trail for the current filters. Admins/managers see every action; a
    seller sees only their own. The trail only grows, so it is filterable by who
    acted, which action, a date window and free text — otherwise a single entry
    becomes unfindable past the first few pages. Shared by the page and its export."""
    logs = AuditLog.objects.select_related("user")
    if not request.user.can_see_all_records:
        logs = logs.filter(user=request.user)

    filters = {key: request.GET.get(key, "") for key in ("rep", "action", "dan", "gacha")}
    filters["q"] = request.GET.get("q", "").strip()

    if filters["q"]:
        logs = logs.filter(
            Q(summary__icontains=filters["q"]) | Q(target_type__icontains=filters["q"])
        )

    reps = rep_obj = None
    if request.user.can_see_all_records:
        reps = User.objects.filter(is_active=True).order_by(
            "first_name", "last_name", "username"
        )
        if filters["rep"].isdigit():
            rep_obj = reps.filter(pk=filters["rep"]).first()
            if rep_obj:
                logs = logs.filter(user=rep_obj)

    actions = AuditLog.Action.choices
    action_label = dict(actions).get(filters["action"], "")
    if action_label:
        logs = logs.filter(action=filters["action"])

    # A full history, so dates only bite once the user actually sets them.
    date_from = _parse_date(filters["dan"])
    date_to = _parse_date(filters["gacha"])
    if date_from and date_to and date_to < date_from:
        date_from, date_to = date_to, date_from
        filters["dan"], filters["gacha"] = date_from.isoformat(), date_to.isoformat()
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    active_filters = _filter_chips(request, [
        {"param": "rep", "label": "Kim", "value": str(rep_obj) if rep_obj else ""},
        {"param": "action", "label": "Amal", "value": action_label},
        {"param": "dan", "label": "Sanadan",
         "value": date_from.strftime("%d.%m.%Y") if date_from else ""},
        {"param": "gacha", "label": "Sanagacha",
         "value": date_to.strftime("%d.%m.%Y") if date_to else ""},
    ])
    return logs, filters, reps, actions, active_filters


def _audit_links(logs):
    """{log.pk: url} — where each trail line leads, for the lines that still have
    somewhere to go.

    Deleted and voided records are left out on purpose: the line survives, the thing it
    happened to does not. Everything else is resolved in batches, one query per kind of
    target on the page rather than one per row.

    "To'lov" and "Qaytarish" are filed under what the money was ABOUT rather than the
    payment row itself — and not always under the same thing: a debt payment carries its
    client, a voided receipt payment carries its sale. Both are tried, and the one whose
    client's name appears in the line's own text wins. A link that opens some unrelated
    client's page would be worse than no link at all."""
    dead = {AuditLog.Action.DELETE, AuditLog.Action.VOID}
    live = [log for log in logs if log.action not in dead and log.target_id]

    def pks(*types):
        return {log.target_id for log in live if log.target_type in types}

    sales = Sale.objects.select_related("client").in_bulk(pks("Sotuv", "To'lov", "Qaytarish"))
    clients = Client.objects.in_bulk(pks("Mijoz", "To'lov"))
    products = Product.objects.in_bulk(pks("Mahsulot"))
    entries = StockEntry.objects.in_bulk(pks("Ombor"))
    employees = Employee.objects.in_bulk(pks("Xodim"))

    def existing(model, *types):
        return set(model.objects.filter(pk__in=pks(*types)).values_list("pk", flat=True))

    expenses = existing(Expense, "Chiqim")
    remittances = existing(ProductionRemittance, "Topshiruv")
    payouts = existing(ProfitPayout, "Foyda")
    receipts = existing(ProductionReceipt, "Qabul", "Zakaz")
    users = existing(User, "Foydalanuvchi")

    links = {}
    for log in live:
        kind, pk, url = log.target_type, log.target_id, None
        if kind == "Sotuv" and pk in sales:
            url = reverse("sale_detail", args=[pk])
        elif kind in ("To'lov", "Qaytarish"):
            # The client is tried first: that is what almost every payment line is
            # filed under, and where a debt payment spread over several receipts
            # actually shows up. A sale id only wins when no client of that id is
            # named in the line.
            sale, client = sales.get(pk), clients.get(pk)
            if client and client.name in log.summary:
                url = reverse("client_history", args=[pk])
            elif sale and sale.client.name in log.summary:
                url = reverse("sale_detail", args=[pk])
        elif kind == "Mijoz" and pk in clients:
            url = reverse("client_history", args=[pk])
        elif kind == "Mahsulot" and pk in products:
            url = reverse("product_detail", args=[pk])
        elif kind == "Ombor" and pk in entries:
            # The entry itself has no screen; the product's page is where it shows up.
            url = reverse("product_detail", args=[entries[pk].product_id])
        elif kind == "Xodim" and pk in employees:
            url = reverse("employee_detail", args=[pk])
        elif kind == "Chiqim" and pk in expenses:
            url = reverse("kassa_entry_detail", args=["expense", pk])
        elif kind == "Topshiruv" and pk in remittances:
            url = reverse("kassa_entry_detail", args=["remittance", pk])
        elif kind == "Foyda" and pk in payouts:
            url = reverse("kassa_entry_detail", args=["profit", pk])
        elif kind in ("Qabul", "Zakaz") and pk in receipts:
            url = reverse("receipt_edit", args=[pk])
        elif kind == "Foydalanuvchi" and pk in users:
            url = reverse("user_edit", args=[pk])
        if url:
            links[log.pk] = url
    return links


def audit_list(request):
    """Amallar tarixi: who did what, and when — every recorded action, newest first."""
    logs, filters, reps, actions, active_filters = _audit_rows(request)
    export_qs = request.GET.urlencode()
    page = Paginator(logs, 50).get_page(request.GET.get("page"))
    # Each line's own destination, hung on the row it belongs to. The queryset caches
    # its rows, so these are the very objects the template will iterate.
    links = _audit_links(list(page.object_list))
    for log in page.object_list:
        log.link = links.get(log.pk)
    return render(request, "crm/audit_list.html", {
        "page": page,
        # The same date control the kassa carries, opening on the whole history:
        # one calendar, one pair of arrows, in the place the eye already looks for it.
        **_date_range_context(request, default_window="all"),
        "show_daterange_picker": True,
        "keep_daterange": True,
        "allow_all_window": True,
        "filters": filters,
        "reps": reps,
        "rep_label": "Kim",
        "actions": actions,
        "active_filters": active_filters,
        "filter_count": len(active_filters),
        "has_filters": bool(active_filters),
        "filter_url": reverse("audit_list"),
        "export_url": reverse("audit_export") + (f"?{export_qs}" if export_qs else ""),
        "search_placeholder": "Tafsilot bo'yicha qidirish…",
    })


def audit_export(request):
    """Excel (.xlsx) of the trail as filtered — the rows on screen, all pages of them."""
    logs, *_ = _audit_rows(request)
    headers = ["Sana", "Soat", "Kim", "Amal", "Obyekt", "ID", "Tafsilot"]
    rows = [
        [
            timezone.localtime(log.created_at).strftime("%d.%m.%Y"),
            timezone.localtime(log.created_at).strftime("%H:%M"),
            str(log.user) if log.user else "",
            log.event["label"],
            log.target_type,
            log.target_id,
            log.summary,
        ]
        # Capped: the trail grows without limit and an export is read, not archived.
        for log in logs[:20000]
    ]
    return _xlsx_response("amallar-tarixi.xlsx", "Amallar tarixi", headers, rows)


# --- Kassa (cash register) ----------------------------------------------------

def _currency_till(payments, expenses, refunds, date_from, date_to, *, som):
    """Income (by method), expense, refunds and the running balance for ONE currency
    drawer.

    The so'm drawer counts net so'm (amount − bank fee); the dollar drawer counts the
    physical dollars handed over (`amount_original`). The balance is cumulative:
    opening = everything strictly before date_from, closing = opening + in − out.

    Refunds are a third flow, kept apart from expenses: money handed back on an
    over-returned sale leaves the drawer just the same, but it is the client's money
    coming back, not a cost the business bore."""
    field = PAYMENT_NET if som else F("amount_original")
    exp_field = "amount" if som else "amount_original"

    def income(**flt):
        return payments.filter(**flt).aggregate(s=Sum(field))["s"] or Decimal("0")

    def outflow(**flt):
        return expenses.filter(**flt).aggregate(s=Sum(exp_field))["s"] or Decimal("0")

    def refund(**flt):
        return refunds.filter(**flt).aggregate(s=Sum(exp_field))["s"] or Decimal("0")

    window = {"date__gte": date_from, "date__lte": date_to}
    cash = income(method=Payment.Method.CASH, **window)
    card = income(method=Payment.Method.CARD, **window)
    bank = income(method=Payment.Method.TRANSFER, **window)
    total_in = cash + card + bank
    total_out = outflow(**window)
    total_refund = refund(**window)
    opening = (
        income(date__lt=date_from)
        - outflow(date__lt=date_from)
        - refund(date__lt=date_from)
    )
    return {
        "cash": cash, "card": card, "bank": bank, "income": total_in,
        "expense": total_out, "refund": total_refund, "opening": opening,
        "closing": opening + total_in - total_out - total_refund,
    }


def _kassa_supplier_cost(date_from, date_to, rep=None):
    """Total supplier cost (Tannarx / asl narx) of goods sold in the window — what
    the business owes suppliers for the goods it moved this period. Scoped to one
    employee when `rep` is given. Always so'm (cost prices are stored in so'm).
    `date_from=None` drops the lower bound, giving the cumulative (as-of date_to)
    figure a standing balance needs."""
    items = SaleItem.objects.filter(sale__date__lte=date_to)
    returns = Return.objects.filter(sale__date__lte=date_to, restock=True)
    if date_from is not None:
        items = items.filter(sale__date__gte=date_from)
        returns = returns.filter(sale__date__gte=date_from)
    if rep is not None:
        items = items.filter(sale__sales_rep=rep)
        returns = returns.filter(sale__sales_rep=rep)
    sold = items.aggregate(s=Sum(COST))["s"] or Decimal("0")
    # Restocked goods are back in the warehouse, so their tannarx is no longer owed.
    # Written-off returns stay in the figure — see `seller_production_debt`.
    given_back = returns.aggregate(s=Sum(RETURN_COST))["s"] or Decimal("0")
    return sold - given_back


def _kassa_remitted(date_from, date_to, rep=None):
    """Total cash handed back to production (Ishlab chiqarishga topshirilgan) in the
    window. Scoped to one seller when `rep` is given. Always so'm. `date_from=None`
    drops the lower bound for the cumulative (as-of date_to) total."""
    qs = ProductionRemittance.objects.filter(date__lte=date_to)
    if date_from is not None:
        qs = qs.filter(date__gte=date_from)
    if rep is not None:
        qs = qs.filter(seller=rep)
    return qs.aggregate(s=Sum("amount"))["s"] or Decimal("0")


def _kassa_paid_profit(date_from, date_to, rep=None):
    """Total profit handed up to the boss (Foyda topshirilgan) in the window. Scoped
    to one seller when `rep` is given. Always so'm. `date_from=None` drops the lower
    bound for the cumulative (as-of date_to) total."""
    qs = ProfitPayout.objects.filter(date__lte=date_to)
    if date_from is not None:
        qs = qs.filter(date__gte=date_from)
    if rep is not None:
        qs = qs.filter(seller=rep)
    return qs.aggregate(s=Sum("amount"))["s"] or Decimal("0")


def _realized_profit_by_seller(date_from, date_to, rep=None):
    """Cost-first realized profit per seller, for sales dated in the window.

    Profit is recognised only as money is collected, and a sale's collections cover
    its tannarx FIRST — only the surplus above cost counts as profit. So an unpaid
    debt sale earns nothing yet, and a part-paid one earns only what's collected
    beyond its cost:  realized = max(0, min(paid, revenue) − cost).

    Returns {seller_pk: realized_profit}. Everything is measured AFTER returns:
    net_revenue/net_cost_total drop the goods that came back, and net_paid drops money
    already handed back to the client, so an over-returned sale can't keep earning
    profit on cash it no longer holds.

    Bank fees the SELLER agreed to carry are then subtracted per seller: the client's
    debt was cleared by the full amount they transferred, so the bank's cut comes
    straight out of the seller's earnings (as it already does out of their till). A
    fee charged to the client is left alone — it stays on their balance and is still
    owed, so it costs the seller nothing. Fees follow the PAYMENT's date, the day the
    money and the fee actually moved.
    `date_from=None` drops the lower bound for the cumulative (as-of date_to) total."""
    sales = Sale.objects.filter(date__lte=date_to)
    fees = Payment.objects.filter(
        date__lte=date_to, commission__gt=0, commission_payer=Payment.Payer.SELLER
    )
    if date_from is not None:
        sales = sales.filter(date__gte=date_from)
        fees = fees.filter(date__gte=date_from)
    if rep is not None:
        sales = sales.filter(sales_rep=rep)
        fees = fees.filter(created_by=rep)
    by_seller = {}
    for s in sales.with_balance().values(
        "sales_rep", "net_revenue", "net_cost_total", "net_paid"
    ):
        realized = max(
            Decimal("0"), min(s["net_paid"], s["net_revenue"]) - s["net_cost_total"]
        )
        by_seller[s["sales_rep"]] = by_seller.get(s["sales_rep"], Decimal("0")) + realized
    for f in fees.values("created_by").annotate(s=Sum("commission")):
        uid = f["created_by"]
        by_seller[uid] = by_seller.get(uid, Decimal("0")) - (f["s"] or Decimal("0"))
    return by_seller


def _kassa_profit(date_from, date_to, rep=None):
    """Total realized (cost-first) profit in the window. Scoped to one seller when
    `rep` is given. See `_realized_profit_by_seller` for how it's recognised."""
    return sum(
        _realized_profit_by_seller(date_from, date_to, rep).values(), Decimal("0")
    )


def _kassa_summary(date_from, date_to, rep=None):
    """Two side-by-side till drawers — so'm and dollar — each with its income by
    method, expense and running balance, plus the period's supplier cost. Also the
    production-debt view: cash on hand, tannarx sold, remitted, and remaining debt.
    Scoped to one employee when `rep` is given."""
    # till_income() drops ADVANCE_USED: an advance's cash already counted as income
    # when it was deposited (ADVANCE_IN), so its consumption must not count again.
    payments = Payment.objects.till_income()
    expenses = Expense.objects.all()
    # Cash handed back to clients — on over-returned or price-corrected sales, and
    # advances returned to them. till_income() already drops these, so they have to be
    # brought in separately as an outflow.
    refunds = Payment.objects.till_outflow()
    if rep is not None:
        payments = payments.filter(created_by=rep)
        expenses = expenses.filter(created_by=rep)
        refunds = refunds.filter(created_by=rep)
    uzs, usd = Payment.Currency.UZS, Payment.Currency.USD
    som = _currency_till(
        payments.filter(currency=uzs), expenses.filter(currency=uzs),
        refunds.filter(currency=uzs), date_from, date_to, som=True,
    )
    cost = _kassa_supplier_cost(date_from, date_to, rep)          # period flow
    remitted = _kassa_remitted(date_from, date_to, rep)           # period flow
    paid_profit = _kassa_paid_profit(date_from, date_to, rep)     # period flow
    profit = _kassa_profit(date_from, date_to, rep)
    # Every expense's so'm value, both currencies — the same figure the per-seller
    # rows sum, so the Jami row equals the sum of its columns.
    expense_total = (
        expenses.filter(date__gte=date_from, date__lte=date_to)
        .aggregate(s=Sum("amount"))["s"] or Decimal("0")
    )
    # Standing balances (as of date_to). Cash on hand and production debt don't reset
    # with the day filter — they carry every movement up to the window's end, the way
    # the till's closing balance already does. Only date_to bounds them.
    cost_cum = _kassa_supplier_cost(None, date_to, rep)
    remitted_cum = _kassa_remitted(None, date_to, rep)
    paid_profit_cum = _kassa_paid_profit(None, date_to, rep)
    # Carried-over pre-CRM production debt: one seller's when scoped, else everyone's.
    # It lifts the debt without ever touching income/cash — just like a client opening debt.
    if rep is not None:
        opening_debt = rep.opening_production_debt or Decimal("0")
    else:
        opening_debt = User.objects.aggregate(s=Sum("opening_production_debt"))["s"] or Decimal("0")
    # Admin corrections to the debt. Signed, and cumulative like every other standing
    # balance here. They never reach `cash_on_hand` — no money moved.
    adjust_qs = ProductionAdjustment.objects.filter(date__lte=date_to)
    if rep is not None:
        adjust_qs = adjust_qs.filter(seller=rep)
    adjusted_cum = adjust_qs.aggregate(s=Sum("amount"))["s"] or Decimal("0")
    # Cash on hand combines every method AND currency: Payment.amount is always the
    # so'm value (a dollar payment is converted at entry), so PAYMENT_NET nets a
    # dollar payment to its so'm too — no currency filter here.
    income_all_cum = (
        payments.filter(date__lte=date_to)
        .aggregate(s=Sum(PAYMENT_NET))["s"] or Decimal("0")
    )
    expense_cum = (
        expenses.filter(date__lte=date_to).aggregate(s=Sum("amount"))["s"] or Decimal("0")
    )
    refund_cum = (
        refunds.filter(date__lte=date_to).aggregate(s=Sum("amount"))["s"] or Decimal("0")
    )
    refunded = (
        refunds.filter(date__gte=date_from, date__lte=date_to)
        .aggregate(s=Sum("amount"))["s"] or Decimal("0")
    )
    cash_on_hand = income_all_cum - refund_cum - expense_cum - remitted_cum - paid_profit_cum
    return {
        "som": som,
        "usd": _currency_till(
            payments.filter(currency=usd), expenses.filter(currency=usd),
            refunds.filter(currency=usd), date_from, date_to, som=False,
        ),
        "cost": cost,
        # Production-debt block (so'm). Cash on hand = so'm income net of fees, less
        # every expense, less what's already been handed to production, less profit
        # already handed to the boss — all cumulative.
        "remitted": remitted,
        "paid_profit": paid_profit,
        "production_debt": opening_debt + cost_cum - remitted_cum + adjusted_cum,
        "cash": cash_on_hand,
        # Profit still sitting in the till, free to hand up: cash beyond the debt.
        "withdrawable_profit": cash_on_hand
        - (opening_debt + cost_cum - remitted_cum + adjusted_cum),
        "profit": profit,
        "expense_total": expense_total,
        "refunded": refunded,
        "net_profit": profit - expense_total,
    }


def _per_employee_kassa(date_from, date_to, rep=None):
    """Per-seller kassa control for the window. Each row carries: money taken in
    (so'm gross / dollars), the bank fees withheld from them, money paid out, the
    profit their sales earned, the tannarx of what they sold (= their production debt
    before handovers), how much they've handed back to production, and two derived
    figures —

      cash            = so'm income − bank fees − expenses − remitted  (naqd qo'lida)
      production_debt = sold tannarx − remitted             (ishlab chiqarishga qarz)
      net             = realized profit − expenses          (samaradorlik)

    `profit` is realized cost-first (see `_realized_profit_by_seller`): an unpaid
    debt sale earns nothing until its tannarx is collected. Scoped to one seller
    when `rep` is given (a seller sees only their own row).

    This is a standing control snapshot: every column is cumulative as of date_to
    (the day filter's lower bound is dropped), so cash on hand and production debt
    read as the true outstanding totals and each row reconciles
    (debt = sold − remitted, cash = income − fees − expenses − remitted)."""
    window = {"date__lte": date_to}
    sale_window = {"sale__date__lte": date_to}
    # Exclude ADVANCE_USED — already counted as income at deposit time (ADVANCE_IN).
    payments = Payment.objects.till_income().filter(**window)
    expenses = Expense.objects.filter(**window)
    sale_items = SaleItem.objects.filter(**sale_window)
    restocked = Return.objects.filter(restock=True, **sale_window)
    refunds = Payment.objects.till_outflow().filter(**window)
    remittances = ProductionRemittance.objects.filter(**window)
    payouts = ProfitPayout.objects.filter(**window)
    adjustments = ProductionAdjustment.objects.filter(**window)
    if rep is not None:
        payments = payments.filter(created_by=rep)
        expenses = expenses.filter(created_by=rep)
        sale_items = sale_items.filter(sale__sales_rep=rep)
        restocked = restocked.filter(sale__sales_rep=rep)
        refunds = refunds.filter(created_by=rep)
        remittances = remittances.filter(seller=rep)
        payouts = payouts.filter(seller=rep)
        adjustments = adjustments.filter(seller=rep)

    users = {u.pk: u for u in User.objects.all()}
    usd = Payment.Currency.USD

    def blank(uid):
        u = users.get(uid)
        return {
            "uid": uid,
            "employee": str(u) if u else "—",
            "in_som": Decimal("0"), "in_usd": Decimal("0"),
            "out_som": Decimal("0"), "out_usd": Decimal("0"),
            "expense_total": Decimal("0"), "profit": Decimal("0"),
            "commission": Decimal("0"),
            "sold_cost": Decimal("0"), "remitted": Decimal("0"),
            "paid_profit": Decimal("0"), "refunded": Decimal("0"),
            # Signed admin corrections to the production debt — no money behind them.
            "adjusted": Decimal("0"),
            # Carried-over pre-CRM production debt, part of this seller's debt from day one.
            "opening_debt": (u.opening_production_debt if u else Decimal("0")) or Decimal("0"),
        }

    rows = {}

    def row(uid):
        return rows.setdefault(uid, blank(uid))

    # Seed a row for every seller carrying an opening production debt, so their debt shows
    # (and feeds the Jami total) even when they've had no other movement in the window.
    for u in users.values():
        if u.opening_production_debt:
            if rep is None or (rep is not None and u.pk == rep.pk):
                row(u.pk)
    # Same for a seller whose only movement is an admin correction.
    for uid in adjustments.values_list("seller", flat=True).distinct():
        row(uid)

    for r in (
        payments.values("created_by", "currency")
        .annotate(som=Sum("amount"), fee=Sum("commission"), usd_amt=Sum("amount_original"))
    ):
        rr = row(r["created_by"])
        # Cash combines currencies: `amount` is the so'm value of every payment, so a
        # dollar payment counts toward in_som at its so'm value too. Income is the
        # gross taken in; the bank's cut is subtracted separately below, which keeps
        # the fee visible as its own figure instead of hiding inside the income.
        rr["in_som"] += r["som"] or Decimal("0")
        rr["commission"] += r["fee"] or Decimal("0")
        if r["currency"] == usd:
            rr["in_usd"] += r["usd_amt"] or Decimal("0")

    for r in (
        expenses.values("created_by", "currency")
        .annotate(som=Sum("amount"), usd_amt=Sum("amount_original"))
    ):
        rr = row(r["created_by"])
        rr["expense_total"] += r["som"] or Decimal("0")  # so'm value of every expense
        if r["currency"] == usd:
            rr["out_usd"] += r["usd_amt"] or Decimal("0")
        else:
            rr["out_som"] += r["som"] or Decimal("0")

    for r in sale_items.values("sale__sales_rep").annotate(cost=Sum(COST)):
        row(r["sale__sales_rep"])["sold_cost"] += r["cost"] or Decimal("0")  # tannarx = debt

    # Restocked goods went back to the warehouse, so their tannarx stops being owed.
    for r in restocked.values("sale__sales_rep").annotate(cost=Sum(RETURN_COST)):
        row(r["sale__sales_rep"])["sold_cost"] -= r["cost"] or Decimal("0")

    # Profit is realized cost-first (only collections above a sale's tannarx count),
    # so it can't be a flat SaleItem sum — pull the per-seller figure instead.
    for uid, realized in _realized_profit_by_seller(None, date_to, rep).items():
        row(uid)["profit"] += realized

    for r in remittances.values("seller").annotate(s=Sum("amount")):
        row(r["seller"])["remitted"] += r["s"] or Decimal("0")

    for r in payouts.values("seller").annotate(s=Sum("amount")):
        row(r["seller"])["paid_profit"] += r["s"] or Decimal("0")

    for r in adjustments.values("seller").annotate(s=Sum("amount")):
        row(r["seller"])["adjusted"] += r["s"] or Decimal("0")

    for r in refunds.values("created_by").annotate(s=Sum("amount")):
        row(r["created_by"])["refunded"] += r["s"] or Decimal("0")

    result = []
    for rr in rows.values():
        # Cash left: gross income, less the bank fees that never reached the till, less
        # money refunded to clients, less expenses, less handed to production, less
        # profit handed to the boss.
        rr["cash"] = (
            rr["in_som"] - rr["commission"] - rr["refunded"] - rr["expense_total"]
            - rr["remitted"] - rr["paid_profit"]
        )
        rr["production_debt"] = (
            rr["opening_debt"] + rr["sold_cost"] - rr["remitted"] + rr["adjusted"]
        )
        rr["net"] = rr["profit"] - rr["expense_total"]  # samaradorlik: foyda − rasxot
        result.append(rr)
    result.sort(key=lambda r: (r["in_som"] + r["profit"]), reverse=True)
    return result


def _kassa_expenses(request):
    """The kassa expense queryset for the window, narrowed by the drawer filters
    (employee, turkum, usul, valyuta). Shared by the page and its CSV export.
    Returns (expenses, dates, filters, rep, reps)."""
    dates = _date_range_context(request)
    filters = {
        key: request.GET.get(key, "")
        for key in ("method", "currency", "rep", "q")
    }
    # Turkum is the one filter that takes more than one answer: the drawer offers a
    # tick box per category, so it arrives as repeated `category=` params and is kept
    # as a list. Blanks are dropped so an empty tick submits as "no filter".
    filters["category"] = [c for c in request.GET.getlist("category") if c.strip()]
    filters["dan"] = dates["date_from"].isoformat()
    filters["gacha"] = dates["date_to"].isoformat()
    # Admins/managers may filter by any employee; a seller is locked to their own
    # till, so the employee filter is never offered to them.
    if request.user.can_see_all_records:
        reps = User.objects.filter(is_active=True).order_by("first_name", "username")
        rep = reps.filter(pk=filters["rep"]).first() if filters["rep"].isdigit() else None
    else:
        reps = None
        rep = request.user
    expenses = Expense.objects.select_related("created_by", "employee").filter(
        date__gte=dates["date_from"], date__lte=dates["date_to"]
    )
    if rep is not None:
        expenses = expenses.filter(created_by=rep)
    if filters["method"] in dict(Payment.Method.choices):
        expenses = expenses.filter(method=filters["method"])
    if filters["category"]:
        expenses = expenses.filter(category__in=filters["category"])
    if filters["currency"] in dict(Payment.Currency.choices):
        expenses = expenses.filter(currency=filters["currency"])
    return expenses.order_by("-date", "-created_at"), dates, filters, rep, reps


def _kassa_transactions(expenses, dates, filters, rep):
    """A unified chronological ledger for the window: every incoming payment (kirim)
    and every expense (chiqim), newest first. Shares the drawer filters (xodim, usul,
    valyuta). A `category` filter is expense-only, so when one is active the incoming
    payments are omitted — the list then reads as a pure expense view."""
    rows = []
    # A category filter is expense-only; when one is active the incoming payments and
    # production handovers are omitted, so the list reads as a pure expense view.
    if not filters["category"]:
        # till_income() drops ADVANCE_USED: it's an internal transfer of already-
        # counted advance money, not a new kirim, so it must not show as income here.
        payments = Payment.objects.till_income().select_related(
            "sale", "sale__client", "client", "created_by"
        ).filter(date__gte=dates["date_from"], date__lte=dates["date_to"])
        if rep is not None:
            payments = payments.filter(created_by=rep)
        if filters["method"] in dict(Payment.Method.choices):
            payments = payments.filter(method=filters["method"])
        if filters["currency"] in dict(Payment.Currency.choices):
            payments = payments.filter(currency=filters["currency"])
        for p in payments:
            # An advance deposit has no sale — its client lives on `client` instead.
            client = p.sale.client if p.sale else p.client
            rows.append({
                "date": p.date, "created_at": p.created_at, "direction": "in",
                "title": client.name if client else "—", "subtitle": p.get_kind_display(),
                "method": p.get_method_display(), "method_code": p.method,
                "currency": p.currency,
                # Gross: the whole sum the client handed over came in. The bank's cut
                # goes out again as its own `commission` row below, so the ledger shows
                # both sides instead of quietly netting them off.
                "amount_som": p.amount, "amount_original": p.original_amount,
                "exchange_rate": p.exchange_rate, "commission_percent": p.commission_percent,
                "created_by": p.created_by,
                "sale_pk": p.sale_id, "client_pk": client.pk if client else None,
                "pk": p.pk, "kind": "payment",
                "is_advance": p.kind == Payment.Kind.ADVANCE_IN,
            })
            # The bank fee on a transfer: money that never reached the till, charged to
            # the seller (the client's debt fell by the full amount). Always so'm — the
            # `commission` field is stored in so'm even on a dollar payment.
            if p.commission:
                rows.append({
                    "date": p.date, "created_at": p.created_at, "direction": "commission",
                    "title": client.name if client else "—",
                    # No subtitle: the "Komissiya" badge on the row already says it.
                    "subtitle": "",
                    "method": p.get_method_display(), "method_code": p.method,
                    "currency": Payment.Currency.UZS,
                    "amount_som": p.commission, "amount_original": p.commission,
                    "exchange_rate": Decimal("0"),
                    "commission_percent": p.commission_percent,
                    "created_by": p.created_by,
                    "sale_pk": p.sale_id, "client_pk": client.pk if client else None,
                    "pk": p.pk, "kind": "commission",
                })
        # Production handovers — so'm only, so a dollar-currency filter hides them.
        if filters["currency"] != Payment.Currency.USD:
            remittances = ProductionRemittance.objects.select_related(
                "seller", "created_by"
            ).filter(date__gte=dates["date_from"], date__lte=dates["date_to"])
            if rep is not None:
                remittances = remittances.filter(seller=rep)
            if filters["method"] in dict(Payment.Method.choices):
                remittances = remittances.filter(method=filters["method"])
            for m in remittances:
                # A negative handover is production giving the cash back. It stays in
                # the chiqim ledger (same stream, same row actions) but carries its own
                # kind so the page can print it as an inflow — and `amount_som` keeps
                # the sign, so the ledger total nets it off on its own.
                rows.append({
                    "date": m.date, "created_at": m.created_at, "direction": "remit",
                    "title": str(m.seller),
                    "subtitle": (
                        "Ishlab chiqarishdan qaytarildi" if m.is_refund
                        else "Ishlab chiqarishga topshiruv"
                    ),
                    "method": m.get_method_display(), "method_code": m.method,
                    "currency": Payment.Currency.UZS,
                    "amount_som": m.amount, "amount_original": m.amount,
                    "amount_abs": m.abs_amount,
                    "exchange_rate": Decimal("0"), "created_by": m.created_by,
                    "pk": m.pk,
                    "kind": "remittance_back" if m.is_refund else "remittance",
                })
        # Profit handovers to the boss — so'm only, so a dollar-currency filter hides them.
        if filters["currency"] != Payment.Currency.USD:
            payouts = ProfitPayout.objects.select_related(
                "seller", "created_by"
            ).filter(date__gte=dates["date_from"], date__lte=dates["date_to"])
            if rep is not None:
                payouts = payouts.filter(seller=rep)
            if filters["method"] in dict(Payment.Method.choices):
                payouts = payouts.filter(method=filters["method"])
            for pp in payouts:
                rows.append({
                    "date": pp.date, "created_at": pp.created_at, "direction": "profit",
                    "title": str(pp.seller), "subtitle": "Foyda topshiruvi",
                    "method": pp.get_method_display(), "method_code": pp.method,
                    "currency": Payment.Currency.UZS,
                    "amount_som": pp.amount, "amount_original": pp.amount,
                    "exchange_rate": Decimal("0"), "created_by": pp.created_by,
                    "pk": pp.pk, "kind": "profit",
                })
    # Cash handed back to clients: settling an over-returned or price-corrected sale,
    # or giving a client the advance they were holding. Not an expense (the business
    # bore no cost — it is the client's own money going back), but it leaves the till
    # all the same, so it belongs in the outflow ledger or the drawer won't reconcile.
    refunds = Payment.objects.till_outflow().filter(
        date__gte=dates["date_from"], date__lte=dates["date_to"],
    ).select_related("client", "created_by").prefetch_related("settled_return")
    if rep is not None:
        refunds = refunds.filter(created_by=rep)
    if filters["method"] in dict(Payment.Method.choices):
        refunds = refunds.filter(method=filters["method"])
    if filters["currency"] in dict(Payment.Currency.choices):
        refunds = refunds.filter(currency=filters["currency"])
    for rf in refunds:
        # A return's refund is only ever corrected by undoing that return, so carry its
        # pk onto the row — the kassa action links straight to it. A price correction
        # has no return behind it; there the action links to the sale (see sale_pk).
        linked = list(rf.settled_return.all())
        is_adjust = rf.kind == Payment.Kind.ADJUST_REFUND
        # An advance return hangs off no sale and no return, so it gets its own row
        # kind: the panel behind it is the one that can undo it.
        is_advance_out = rf.kind == Payment.Kind.ADVANCE_OUT
        rows.append({
            "date": rf.date, "created_at": rf.created_at, "direction": "refund",
            "title": str(rf.client) if rf.client else "—",
            "subtitle": rf.note or (
                "Avans mijozga qaytarildi" if is_advance_out
                else "Narx tuzatildi — naqd qaytarildi" if is_adjust
                else "Qaytarish uchun naqd berildi"
            ),
            "method": rf.get_method_display(), "method_code": rf.method,
            "currency": rf.currency,
            "amount_som": rf.amount, "amount_original": rf.original_amount,
            "exchange_rate": rf.exchange_rate, "created_by": rf.created_by,
            "pk": rf.pk, "kind": "advance_out" if is_advance_out else "refund",
            "sale_pk": rf.sale_id,
            "return_pk": linked[0].pk if linked else None,
            "is_adjust": is_adjust,
            "client_pk": rf.client_id,
        })
    for e in expenses:
        rows.append({
            "date": e.date, "created_at": e.created_at, "direction": "out",
            "title": e.category, "subtitle": e.note,
            "method": e.get_method_display(), "method_code": e.method,
            "currency": e.currency,
            "amount_som": e.amount, "amount_original": e.original_amount,
            "exchange_rate": e.exchange_rate, "created_by": e.created_by,
            "pk": e.pk, "kind": "expense", "employee_pk": e.employee_id,
            # Whose money it was: "Oylik / xodim" on its own leaves the one question
            # the row is asked — who got it — unanswered.
            "employee": e.employee.name if e.employee_id else "",
        })
    rows.sort(key=lambda r: (r["date"], r["created_at"]), reverse=True)
    return _kassa_search(rows, filters.get("q", ""))


def _url_without(request, url_name, *drop):
    """The current URL with some query parameters removed — for a "clear this one"
    link that leaves every other filter standing."""
    params = request.GET.copy()
    for key in drop:
        params.pop(key, None)
    base = reverse(url_name)
    return f"{base}?{params.urlencode()}" if params else base


# Every apostrophe an Uzbek keyboard or a paste from Word can produce, mapped to the
# plain one the database happens to hold.
_APOSTROPHES = str.maketrans({c: "'" for c in "`´‘’ʻʼ‛"})


def _flatten_apostrophes(text):
    return text.translate(_APOSTROPHES)


def _kassa_search(rows, query):
    """Free-text filter over the built ledger rows — one box for both drawers.

    Deliberately NOT five database queries. The ledger is already one merged list of
    payments, bank fees, handovers, profit payouts, refunds and expenses, and each
    built row carries the words a person would actually type: the client or seller,
    the category, the note, the method, who entered it. Matching the rows means the
    same box searches every stream; matching per model would need five near-identical
    clauses and would still miss the rows that have no text column at all.

    Digits are matched separately so "39200", "39 200" and "39,200" all find the same
    row — people read the amount off the screen and type it back with the spaces. An
    amount matches from the START, never mid-number: typing 39200 must not drag in
    10 639 200, because on a money page a wrong row that merely contains the digits
    is worse than no row at all.

    Apostrophes are flattened on both sides. "O'tkazma" is written with ' on one
    keyboard, ’ or ` on the next, and a seller who types the wrong one would otherwise
    get an empty page for a word that is plainly on screen."""
    needle = _flatten_apostrophes((query or "").strip().lower())
    if not needle:
        return rows
    digits = needle.replace(" ", "").replace(",", "").replace("'", "")
    hits = []
    for row in rows:
        haystack = _flatten_apostrophes(" ".join(str(row.get(key) or "") for key in (
            "title", "subtitle", "method", "employee", "created_by",
        )).lower())
        if needle in haystack:
            hits.append(row)
            continue
        if digits.isdigit():
            amount = f"{abs(row.get('amount_som') or 0):.0f}"
            if amount.startswith(digits):
                hits.append(row)
    return hits


def _last_kassa_activity(rep):
    """The most recent day that has anything on the kassa page, for the empty-day
    notice. Rows are placed by the date written ON them, not the day they were typed,
    and a seller normally enters yesterday's takings this morning — so a kassa that
    opens on today reads as empty and the whole day gets keyed in a second time. The
    notice points at the day the money is actually on."""
    days = [
        Payment.objects.till_income().filter(created_by=rep) if rep else
        Payment.objects.till_income(),
        Expense.objects.filter(created_by=rep) if rep else Expense.objects.all(),
        ProductionRemittance.objects.filter(seller=rep) if rep else
        ProductionRemittance.objects.all(),
        ProfitPayout.objects.filter(seller=rep) if rep else ProfitPayout.objects.all(),
    ]
    found = [qs.aggregate(d=Max("date"))["d"] for qs in days]
    return max([d for d in found if d], default=None)


def kassa_view(request):
    """The cash register (Kassa): two till drawers (so'm + dollar) with income by
    method and running balance, per-employee kassa & performance, and the expense
    list. Visible to everyone — the shared company till. Any filter (employee, turkum,
    usul, valyuta) scopes the figures so a supervisor can drill into one employee."""
    expenses, dates, filters, rep, reps = _kassa_expenses(request)
    date_from, date_to = dates["date_from"], dates["date_to"]
    summary = _kassa_summary(date_from, date_to, rep=rep)
    transactions = _kassa_transactions(expenses, dates, filters, rep)
    # Two side-by-side ledgers: kirim (client payments) on the left, chiqim
    # (expenses, production handovers and bank fees) on the right. Totals use
    # amount_som so a USD payment counts at its so'm value. Newest-first order is
    # inherited.
    income_rows = [t for t in transactions if t["direction"] == "in"]
    outflow_rows = [
        t for t in transactions
        if t["direction"] in ("out", "remit", "profit", "refund", "commission")
    ]
    income_total = sum((t["amount_som"] for t in income_rows), Decimal("0"))
    outflow_total = sum((t["amount_som"] for t in outflow_rows), Decimal("0"))

    # Nothing on the chosen day? Say where the money actually is, and offer one click
    # to go there — see `_last_kassa_activity` for why this happens every morning.
    empty_hint = None
    if not income_rows and not outflow_rows:
        last_day = _last_kassa_activity(rep)
        in_window = last_day and date_from and date_from <= last_day <= date_to
        if last_day and not in_window:
            params = request.GET.copy()
            params["dan"] = params["gacha"] = last_day.isoformat()
            empty_hint = {"date": last_day, "url": f"?{params.urlencode()}"}

    method_labels = dict(Payment.Method.choices)
    currency_labels = dict(Payment.Currency.choices)
    # Only the company view exposes a rep chip; a seller's own scope isn't a filter.
    rep_chip = str(rep) if (reps is not None and rep) else ""
    active_filters = _filter_chips(request, [
        {"param": "rep", "label": "Xodim", "value": rep_chip},
        {"param": "category", "label": "Turkum", "value": ", ".join(filters["category"])},
        {"param": "method", "label": "Usul", "value": method_labels.get(filters["method"], "")},
        {"param": "currency", "label": "Valyuta", "value": currency_labels.get(filters["currency"], "")},
    ])
    # Per-seller control rows. Admins/managers see everyone (or one, if they filtered
    # by a rep); a seller sees only their own row.
    seller_rows = _per_employee_kassa(date_from, date_to, rep=rep)
    # Column-wise totals for the Jami row — sums the cumulative rows so the footer
    # always equals the sum of what's shown, whatever the filter.
    seller_totals = {
        key: sum((r[key] for r in seller_rows), Decimal("0"))
        for key in ("cash", "production_debt", "sold_cost", "remitted", "paid_profit",
                    "expense_total", "commission", "net")
    }
    my_row = None
    if not request.user.can_see_all_records:
        my_row = seller_rows[0] if seller_rows else None

    # Debt corrections for the window. Listed on their own — they move no cash, so
    # putting them in the outflow ledger would stop the drawer reconciling. A seller
    # sees only the ones filed against them.
    adjustments = ProductionAdjustment.objects.filter(
        date__gte=date_from, date__lte=date_to
    ).select_related("seller", "created_by")
    if rep is not None:
        adjustments = adjustments.filter(seller=rep)
    elif not request.user.can_see_all_records:
        adjustments = adjustments.filter(seller=request.user)
    export_qs = request.GET.urlencode()
    return render(request, "crm/kassa.html", {
        "summary": summary,
        "empty_hint": empty_hint,
        "debt_adjustments": adjustments,
        "income_rows": income_rows,
        "outflow_rows": outflow_rows,
        "income_total": income_total,
        "outflow_total": outflow_total,
        "expenses": expenses,
        "per_employee": seller_rows if request.user.can_see_all_records else None,
        "seller_totals": seller_totals,
        "my_row": my_row,
        "filters": filters,
        "reps": reps,
        "active_filters": active_filters,
        "filter_count": len(active_filters),
        "has_filters": bool(active_filters),
        "filter_url": reverse("kassa"),
        # One box over both drawers. The window and the drawer filters ride along as
        # hidden inputs, so searching narrows what is already on screen instead of
        # throwing the page back to today.
        "show_search": True,
        "search_placeholder": "Mijoz, turkum, izoh, summa…",
        # Turkum can hold several answers, so it contributes one hidden field per
        # ticked category rather than a single value like the others.
        "search_keep": [
            {"name": name, "value": value} for name, value in (
                ("dan", filters["dan"]), ("gacha", filters["gacha"]),
                ("method", filters["method"]),
                ("currency", filters["currency"]), ("rep", filters["rep"]),
            )
        ] + [{"name": "category", "value": c} for c in filters["category"]],
        "search_clear_url": _url_without(request, "kassa", "q"),
        "rep_label": "Xodim",
        "show_daterange_picker": True,
        "keep_daterange": True,
        "show_method": True,
        "show_category": True,
        "category_options": Expense.used_categories(),
        "show_currency": True,
        # One Excel button in the toolbar; it opens a chooser (hammasi / kirim /
        # chiqim) because the kassa page holds two ledgers, not one list.
        "export_url": reverse("kassa_export") + (f"?{export_qs}" if export_qs else ""),
        "export_modal": True,
        **dates,
    })


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _fill_sheet(ws, headers, rows, number_formats=None):
    """Write one worksheet: bold frozen header, one row per record, columns sized
    to their widest value. `number_formats` maps a 1-based column index to an Excel
    format string, applied to that column's data cells."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    number_formats = number_formats or {}
    for i, header in enumerate(headers, start=1):
        letter = get_column_letter(i)
        longest = max([len(str(header))] + [len(str(r[i - 1])) for r in rows])
        ws.column_dimensions[letter].width = min(longest + 2, 40)
        fmt = number_formats.get(i)
        if fmt:
            for cell in ws[letter][1:]:  # data cells only, skip the header
                cell.number_format = fmt


def _xlsx_download(workbook, filename):
    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def _xlsx_response(filename, sheet_title, headers, rows, number_formats=None):
    """Build a one-sheet .xlsx download."""
    wb = Workbook()
    wb.active.title = sheet_title
    _fill_sheet(wb.active, headers, rows, number_formats)
    return _xlsx_download(wb, filename)


def _xlsx_book_response(filename, sheets):
    """Build a multi-sheet .xlsx download. `sheets` is a list of
    (title, headers, rows, number_formats) — one tab each."""
    wb = Workbook()
    for i, (title, headers, rows, formats) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        _fill_sheet(ws, headers, rows, formats)
    return _xlsx_download(wb, filename)


# How an outflow row reads in the Chiqim export — the same four kinds the Chiqim
# panel shows side by side.
_OUTFLOW_LABELS = {
    "expense": "Rasxot",
    "remittance": "Ishlab chiqarishga topshiruv",
    "remittance_back": "Ishlab chiqarishdan qaytarildi",
    "profit": "Foyda topshiruvi",
    "refund": "Qaytarish (naqd berildi)",
    "advance_out": "Avans qaytarildi",
    "commission": "Bank komissiyasi",
}


def _kassa_ledger(request):
    """The kassa ledger for the current window and drawer filters, split exactly as
    the page splits it: (kirim rows, chiqim rows)."""
    expenses, dates, filters, rep, _ = _kassa_expenses(request)
    rows = _kassa_transactions(expenses, dates, filters, rep)
    income = [r for r in rows if r["direction"] == "in"]
    outflow = [
        r for r in rows
        if r["direction"] in ("out", "remit", "profit", "refund", "commission")
    ]
    return income, outflow


def _currency_label(row):
    return dict(Payment.Currency.choices).get(row["currency"], "")


def _rate_cell(row):
    """The exchange rate, but only where there is one — a so'm row leaves it blank."""
    return float(row["exchange_rate"]) if row["currency"] == Payment.Currency.USD else ""


def _income_sheet(income):
    """The Kirim tab: money taken in over the window."""
    headers = [
        "Sana", "Kimdan", "Kirim turi", "Usul", "Valyuta",
        "Kirim summa (so'm)", "Asl summa", "Kurs", "Qarzga ta'sir", "Kim qabul qildi",
    ]
    rows = [
        [
            r["date"].strftime("%d.%m.%Y"),
            r["title"],
            r["subtitle"],
            r["method"],
            _currency_label(r),
            float(r["amount_som"]),
            float(r["amount_original"]),
            _rate_cell(r),
            # An advance deposit is credit held for the client, not a debt payment,
            # so it moves no debt — the page shows a dash in this column.
            0 if r.get("is_advance") else float(r["amount_som"]),
            str(r["created_by"]),
        ]
        for r in income
    ]
    number_formats = {6: "#,##0.00", 7: "#,##0.00", 8: "#,##0.00", 9: "#,##0.00"}
    return ("Kirimlar", headers, rows, number_formats)


def _outflow_sheet(outflow):
    """The Chiqim tab: rasxot, ishlab chiqarishga topshiruv, foyda topshiruvi, bank
    fees and cash refunded to clients — everything that left the till."""
    headers = [
        "Sana", "Turi", "Tavsif", "Xodim", "Izoh", "Usul", "Valyuta",
        "Chiqim summa (so'm)", "Asl summa", "Kurs", "Kim kiritdi",
    ]
    rows = [
        [
            r["date"].strftime("%d.%m.%Y"),
            _OUTFLOW_LABELS.get(r["kind"], r["kind"]),
            r["title"],
            r.get("employee", ""),
            r["subtitle"],
            r["method"],
            _currency_label(r),
            float(r["amount_som"]),
            float(r["amount_original"]),
            _rate_cell(r),
            str(r["created_by"]),
        ]
        for r in outflow
    ]
    number_formats = {8: "#,##0.00", 9: "#,##0.00", 10: "#,##0.00"}
    return ("Chiqimlar", headers, rows, number_formats)


def _kassa_export_presets(today):
    """Quick windows offered in the Excel dialog. "Hammasi" starts at the oldest
    money movement on record, so it really does mean everything."""
    firsts = [
        Payment.objects.order_by("date").values_list("date", flat=True).first(),
        Expense.objects.order_by("date").values_list("date", flat=True).first(),
    ]
    earliest = min([d for d in firsts if d] or [today])
    yesterday = today - timedelta(days=1)
    return [
        ("Bugun", today, today),
        ("Kecha", yesterday, yesterday),
        ("7 kun", today - timedelta(days=6), today),
        ("Shu oy", today.replace(day=1), today),
        ("Hammasi", earliest, today),
    ]


def kassa_export(request):
    """The Excel chooser: the kassa holds two ledgers, so the button asks which one
    (or both) before downloading. The period is picked here too — the dialog carries
    its own window, so a report can be pulled without disturbing the page's view."""
    dates = _date_range_context(request)
    # Everything except the window rides along untouched (xodim, turkum, usul,
    # valyuta); the window itself is whatever the dialog currently shows.
    rest = request.GET.copy()
    for key in ("dan", "gacha", "page"):
        rest.pop(key, None)

    def url(name, date_from=None, date_to=None):
        params = rest.copy()
        params["dan"] = (date_from or dates["date_from"]).isoformat()
        params["gacha"] = (date_to or date_from or dates["date_to"]).isoformat()
        return f"{reverse(name)}?{params.urlencode()}"

    presets = []
    for label, start, end in _kassa_export_presets(timezone.localdate()):
        # Two presets can describe the same window (a business one day old has
        # "Bugun" == "Hammasi"); only the first is highlighted, so exactly one
        # chip ever reads as the current choice.
        matches = start == dates["date_from"] and end == dates["date_to"]
        already = any(p["active"] for p in presets)
        presets.append({
            "label": label,
            "url": url("kassa_export", start, end),
            "active": matches and not already,
        })
    rest_qs = rest.urlencode()
    return render(request, "crm/_kassa_export_modal.html", {
        "title": "Excelga yuklash",
        "all_url": url("kassa_export_all"),
        "income_url": url("kassa_income_export"),
        "outflow_url": url("kassa_outflow_export"),
        # Base for the date inputs: the dialog re-opens on this URL with the dates
        # the user typed (see the [data-export-range] handler in base.html).
        "range_url": reverse("kassa_export") + (f"?{rest_qs}" if rest_qs else ""),
        "presets": presets,
        "other_filters": bool(rest_qs),
        **dates,
    })


def kassa_export_all(request):
    """Both ledgers in one workbook — Kirimlar and Chiqimlar as separate tabs."""
    income, outflow = _kassa_ledger(request)
    return _xlsx_book_response(
        "kassa.xlsx", [_income_sheet(income), _outflow_sheet(outflow)]
    )


def kassa_income_export(request):
    income, _ = _kassa_ledger(request)
    title, headers, rows, formats = _income_sheet(income)
    return _xlsx_response("kirimlar.xlsx", title, headers, rows, formats)


def kassa_outflow_export(request):
    _, outflow = _kassa_ledger(request)
    title, headers, rows, formats = _outflow_sheet(outflow)
    return _xlsx_response("chiqimlar.xlsx", title, headers, rows, formats)


def _expense_response(request, form, title, invalid=False):
    """The expense modal, with the categories already in use as datalist
    suggestions — Turkum is free text, the list is only a shortcut."""
    return form_response(
        request, form, title, invalid=invalid,
        modal_template="crm/_expense_modal.html",
        category_suggestions=Expense.used_categories(),
    )


def _expense_back(request):
    """Where an expense form returns to. The kassa by default; the payroll pages when
    they are the ones that sent you here — a wage fixed from a worker's own page should
    land back on that page, not somewhere else entirely. `next` arrives in the URL, so
    only these known routes are honoured, never a raw address."""
    target = request.GET.get("next", "")
    if target == "xodimlar":
        return reverse("employee_list")
    if target.startswith("xodim-"):
        pk = target[len("xodim-"):]
        if pk.isdigit() and Employee.objects.filter(pk=pk).exists():
            return reverse("employee_detail", args=[pk])
    return reverse("kassa")


def expense_create(request):
    """Record a payout from the till. Any logged-in user may add one — staff come to
    the cashier and the expense is written against the kassa (logged for audit).

    The Xodimlar page links in with ?employee=<pk>, which preselects the worker and
    the wage category so paying someone is one click from their row; it may add
    ?summa= to fill in what is still owed, for the common case of settling in full."""
    initial = {}
    employee_pk = request.GET.get("employee", "")
    if request.method == "GET" and employee_pk.isdigit():
        if Employee.objects.filter(pk=employee_pk, is_active=True).exists():
            initial["employee"] = employee_pk
            initial["category"] = SALARY_CATEGORY
            # Arriving from a worker's row means a wage or an advance, so the answer
            # is preselected here and nowhere else: picking someone by hand on a
            # plain chiqim leaves it blank on purpose, and the form insists on one.
            initial["counts_against_salary"] = True
            # Only ever a suggestion in the amount box — it is still typed over and
            # still validated like any other; nothing is paid until the form is saved.
            owed = _parse_amount(request.GET.get("summa"))
            if owed and owed > 0:
                initial["amount"] = owed
    form = ExpenseForm(request.POST or None, initial=initial, user=request.user)
    title = "Chiqim qo'shish"
    if request.method == "POST":
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
            usd = (
                f" · ${expense.original_amount:,.2f} × {expense.exchange_rate:,.0f}"
                if expense.currency == Payment.Currency.USD else ""
            )
            # Whether it came off the wage is the part an audit actually needs — the
            # same worker and sum mean two different things depending on the flag.
            who = (
                f" · {expense.employee}"
                f" ({'oyligidan' if expense.counts_against_salary else 'oyligidan emas'})"
                if expense.employee_id else ""
            )
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Chiqim", expense.pk,
                f"{expense.category} chiqimi "
                f"({expense.get_method_display()}){usd}{who} "
                f"— {expense.amount:,.0f} so'm",
            )
            messages.success(request, f"Chiqim qo'shildi: {expense.amount:,.0f} so'm.")
            return form_success(request, _expense_back(request))
        return _expense_response(request, form, title, invalid=True)
    return _expense_response(request, form, title)


def expense_edit(request, pk):
    """Fix a mistaken expense. Admins/managers may edit any; a seller may edit
    only expenses they entered themselves."""
    qs = Expense.objects.all() if request.user.can_see_all_records \
        else Expense.objects.filter(created_by=request.user)
    expense = get_object_or_404(qs, pk=pk)
    title = "Chiqimni tahrirlash"
    form = ExpenseForm(request.POST or None, instance=expense, user=request.user)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Chiqim", expense.pk,
                f"{expense.category} chiqimi — {expense.amount:,.0f} so'm",
            )
            messages.success(request, "Chiqim yangilandi.")
            return form_success(request, _expense_back(request))
        return _expense_response(request, form, title, invalid=True)
    return _expense_response(request, form, title)


def expense_delete(request, pk):
    """Remove a mistaken expense. Admins/managers may erase any; a seller may
    erase only expenses they entered themselves."""
    qs = Expense.objects.select_related("created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    expense = get_object_or_404(qs, pk=pk)
    if request.method == "POST":
        summary = f"{expense.category} — {expense.amount:,.0f} so'm"
        expense.delete()
        AuditLog.record(request.user, AuditLog.Action.DELETE, "Chiqim", pk, summary)
        messages.success(request, "Chiqim o'chirildi.")
        return form_reload(request, _expense_back(request))
    return render_confirm(
        request,
        "Chiqimni o'chirish",
        f"{expense.category} — {expense.amount:,.0f} so'm chiqim "
        f"o'chiriladi. Davom etasizmi?",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


# --- Kassa amalining ichi ----------------------------------------------------

def _money(value):
    """Space-grouped so'm — the way every template prints money. Non-breaking spaces,
    so a figure never splits across two lines however narrow the panel gets."""
    return f"{value:,.0f}".replace(",", " ") + " so'm"


def _percent(value):
    """A rate without its trailing zeros: 4.00 reads as 4, 1.50 as 1.5."""
    return f"{float(value):g}"


def _entry_payment(request, pk, fee_only=False):
    """A client payment (or the bank fee taken out of one — same row, read two ways)."""
    qs = Payment.objects.select_related("sale", "sale__client", "client", "created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    p = get_object_or_404(qs, pk=pk)
    advance = p.kind == Payment.Kind.ADVANCE_IN
    client = p.sale.client if p.sale else p.client
    rows = [
        ("Sana", p.date.strftime("%d.%m.%Y")),
        ("Mijoz", client.name if client else "—"),
        ("Turi", p.get_kind_display()),
        ("To'lov usuli", p.get_method_display()),
        ("Mijoz yubordi", _money(p.amount)),
    ]
    if p.currency == Payment.Currency.USD:
        rows.append((
            "Valyuta",
            f"${p.original_amount:,.2f} × {p.exchange_rate:,.0f}".replace(",", " "),
        ))
    if p.commission:
        rows += [
            (f"Bank komissiyasi ({_percent(p.commission_percent)}%)",
             f"−{_money(p.commission)}"),
            ("Kassaga tushdi", _money(p.net_amount)),
            ("Komissiyani kim ko'tardi", p.get_commission_payer_display()),
        ]
    if not advance:
        rows.append(("Mijoz qarzidan yechildi", _money(p.credited_amount)))
        if p.fee_on_client:
            rows.append(("Mijozda qolgan komissiya", _money(p.commission)))
    else:
        rows.append(("Qarzga ta'siri", "Yo'q — avans sifatida saqlanadi"))
        rows.append((
            "Kassaga kirim",
            "Yo'q — pul avvalroq olingan" if p.is_opening else "Ha",
        ))
    if p.note:
        rows.append(("Izoh", p.note))
    rows.append(("Kim qabul qildi", str(p.created_by)))

    if advance:
        edit, delete = "advance_edit", "advance_delete"
    else:
        edit, delete = "payment_edit", "payment_delete"
    context = {
        "title": "Bank komissiyasi" if fee_only else p.get_kind_display(),
        "rows": rows,
        "edit_url": reverse(edit, args=[p.pk]),
        "edit_label": "To'lovni tahrirlash",
        "delete_url": reverse(delete, args=[p.pk]),
        "delete_label": "To'lovni o'chirish",
    }
    if fee_only:
        # The fee has no life of its own: both buttons act on the payment it came out
        # of, so the labels say so plainly rather than pretending the fee is a record.
        context["note"] = (
            "Komissiya alohida yozuv emas — u shu to'lovning bir qismi. Foizni "
            "o'zgartirish uchun to'lovni tahrirlang; to'lov o'chirilsa komissiya ham "
            "yo'qoladi."
        )
    if p.sale_id:
        context["open_url"] = reverse("sale_detail", args=[p.sale_id])
        context["open_label"] = "Sotuvni ochish"
    elif client:
        context["open_url"] = reverse("client_history", args=[client.pk])
        context["open_label"] = "Mijoz tarixi"
    return context


def _entry_advance_used(request, pk):
    """An open receipt closed from credit the client was already holding.

    Read-only on purpose, unlike every other kassa row. Nobody types this record: the
    advance reconciliation writes it (`_apply_advance_to_open_sales`) and rewrites or
    deletes it whenever the pool moves, so a figure edited by hand here would be
    recalculated away behind the user's back. The two records that DO decide it — the
    deposit and the receipt — are each one click from this panel."""
    qs = Payment.objects.select_related("sale", "sale__client", "client", "created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    p = get_object_or_404(qs, pk=pk, kind=Payment.Kind.ADVANCE_USED)
    client = p.sale.client if p.sale else p.client
    rows = [
        ("Sana", p.date.strftime("%d.%m.%Y")),
        ("Mijoz", client.name if client else "—"),
        ("Turi", "Avansdan yechildi"),
        ("Chekka o'tkazilgan summa", _money(p.amount)),
        ("Mijoz qarzidan yechildi", _money(p.credited_amount)),
        ("Kassaga kirim", "Yo'q — pul avans olingan kuni kirim bo'lgan"),
        ("Kim qabul qilgan", str(p.created_by)),
    ]
    context = {
        "title": "Avansdan yechildi",
        "rows": rows,
        "note": (
            "Bu yozuv qo'lda kiritilmaydi va shu yerdan tahrirlanmaydi — mijozning "
            "avansi ochiq chekka o'tganda tizim o'zi yozadi. Summa noto'g'ri bo'lsa "
            "avansning o'zini yoki chekni tuzating: avans qaytadan taqsimlanadi."
        ),
    }
    if client:
        context["open_url"] = reverse("client_advance_moves", args=[client.pk])
        context["open_label"] = "Mijoz avanslari"
    return context


def _entry_advance_out(request, pk):
    """A client's prepaid credit handed back to them in cash. The deposit it came out
    of is untouched — undoing this row is what puts the credit back, so that is the
    only action offered here."""
    qs = Payment.objects.select_related("client", "created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    p = get_object_or_404(qs, pk=pk, kind=Payment.Kind.ADVANCE_OUT)
    rows = [
        ("Sana", p.date.strftime("%d.%m.%Y")),
        ("Mijoz", str(p.client) if p.client else "—"),
        ("Sababi", "Avans mijozga qaytarildi"),
        ("Berilgan summa", _money(p.amount)),
        ("To'lov usuli", p.get_method_display()),
        ("Kassadan chiqim", "Yo'q — pul kassadan chiqmagan" if p.is_opening else "Ha"),
        ("Kim berdi", str(p.created_by)),
    ]
    if p.note:
        rows.append(("Izoh", p.note))
    context = {
        "title": "Avans qaytarildi",
        "rows": rows,
        "note": (
            "Avans depoziti o'z o'rnida qoldi — pul o'sha kuni haqiqatan kirim "
            "bo'lgan. Bu yozuv esa pulning qaytib chiqqanini ko'rsatadi."
        ),
        "delete_url": reverse("advance_out_delete", args=[p.pk]),
        "delete_label": "Qaytarishni bekor qilish",
    }
    if p.client_id:
        context["open_url"] = reverse("client_history", args=[p.client_id])
        context["open_label"] = "Mijoz tarixi"
    return context


def _entry_refund(request, pk):
    """Cash handed back to a client — either settling a return or correcting a price
    that was too high. Neither is edited here: the return is undone through the return
    itself, the correction by re-editing the sale."""
    qs = Payment.objects.select_related("client", "created_by").prefetch_related(
        "settled_return"
    )
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    p = get_object_or_404(qs, pk=pk, kind__in=REFUND_KINDS)
    is_adjust = p.kind == Payment.Kind.ADJUST_REFUND
    linked = list(p.settled_return.all())
    rows = [
        ("Sana", p.date.strftime("%d.%m.%Y")),
        ("Mijoz", str(p.client) if p.client else "—"),
        ("Sababi", "Narx tuzatildi" if is_adjust else "Tovar qaytdi"),
        ("Berilgan summa", _money(p.amount)),
        ("To'lov usuli", p.get_method_display()),
        ("Kim berdi", str(p.created_by)),
    ]
    if p.note:
        rows.append(("Izoh", p.note))
    context = {
        "title": (
            "Narx tuzatildi (naqd berildi)" if is_adjust
            else "Qaytarish (naqd berildi)"
        ),
        "rows": rows,
        "note": (
            "Bu yozuv sotuv narxi tuzatilganda paydo bo'lgan — o'zgartirish uchun "
            "sotuvni qayta tahrirlang."
            if is_adjust else
            "Bu yozuv qaytarishdan kelib chiqadi — bekor qilish uchun qaytarishning "
            "o'zini tahrirlang."
        ),
    }
    if is_adjust and p.sale_id:
        context["edit_url"] = reverse("sale_edit", args=[p.sale_id])
        context["edit_label"] = "Sotuvni tahrirlash"
    elif linked:
        context["edit_url"] = reverse("return_edit", args=[linked[0].pk])
        context["edit_label"] = "Qaytarishni tahrirlash"
    if p.sale_id:
        context["open_url"] = reverse("sale_detail", args=[p.sale_id])
        context["open_label"] = "Sotuvni ochish"
    return context


def _entry_expense(request, pk):
    qs = Expense.objects.select_related("created_by", "employee")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    e = get_object_or_404(qs, pk=pk)
    rows = [
        ("Sana", e.date.strftime("%d.%m.%Y")),
        ("Turkum", e.category),
        ("Summa", _money(e.amount)),
        ("To'lov usuli", e.get_method_display()),
    ]
    if e.currency == Payment.Currency.USD:
        rows.append((
            "Valyuta",
            f"${e.original_amount:,.2f} × {e.exchange_rate:,.0f}".replace(",", " "),
        ))
    if e.employee_id:
        held = "oyligidan ayrildi" if e.counts_against_salary \
            else "oyligiga tegmadi — faqat kassadan chiqim"
        rows.append(("Xodim", f"{e.employee.name} — {held}"))
    if e.note:
        rows.append(("Izoh", e.note))
    rows.append(("Kim kiritdi", str(e.created_by)))
    context = {
        "title": "Chiqim (rasxot)",
        "rows": rows,
        "edit_url": reverse("expense_edit", args=[e.pk]),
        "edit_label": "Chiqimni tahrirlash",
        "delete_url": reverse("expense_delete", args=[e.pk]),
        "delete_label": "Chiqimni o'chirish",
    }
    if e.employee_id and request.user.can_see_all_records:
        context["open_url"] = reverse("employee_list")
        context["open_label"] = "Xodimlar bo'limi"
    return context


def _entry_remittance(request, pk):
    qs = ProductionRemittance.objects.select_related("seller", "created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(seller=request.user)
    r = get_object_or_404(qs, pk=pk)
    rows = [
        ("Sana", r.date.strftime("%d.%m.%Y")),
        ("Sotuvchi", str(r.seller)),
        ("Summa", _money(r.abs_amount)),
        ("Yo'nalishi",
         "Ishlab chiqarish sotuvchiga qaytardi" if r.is_refund
         else "Sotuvchi ishlab chiqarishga topshirdi"),
        ("Kassaga ta'siri",
         f"+{_money(r.abs_amount)}" if r.is_refund else f"−{_money(r.abs_amount)}"),
        ("I.ch. qarziga ta'siri",
         f"+{_money(r.abs_amount)}" if r.is_refund else f"−{_money(r.abs_amount)}"),
        ("To'lov usuli", r.get_method_display()),
    ]
    if r.note:
        rows.append(("Izoh", r.note))
    rows.append(("Kim kiritdi", str(r.created_by)))
    return {
        "title": "Ishlab chiqarishdan qaytarish" if r.is_refund
                 else "Ishlab chiqarishga topshiruv",
        "rows": rows,
        "edit_url": reverse("remittance_edit", args=[r.pk]),
        "edit_label": "Tahrirlash",
        "delete_url": reverse("remittance_delete", args=[r.pk]),
        "delete_label": "O'chirish",
    }


def _entry_profit(request, pk):
    qs = ProfitPayout.objects.select_related("seller", "created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(seller=request.user)
    x = get_object_or_404(qs, pk=pk)
    rows = [
        ("Sana", x.date.strftime("%d.%m.%Y")),
        ("Sotuvchi", str(x.seller)),
        ("Summa", _money(x.amount)),
        ("To'lov usuli", x.get_method_display()),
        ("Kassaga ta'siri", f"−{_money(x.amount)}"),
        ("I.ch. qarziga ta'siri", "Yo'q — qarz allaqachon yopilgan"),
    ]
    if x.note:
        rows.append(("Izoh", x.note))
    rows.append(("Kim kiritdi", str(x.created_by)))
    return {
        "title": "Foyda topshiruvi",
        "rows": rows,
        "edit_url": reverse("profit_payout_edit", args=[x.pk]),
        "edit_label": "Tahrirlash",
        "delete_url": reverse("profit_payout_delete", args=[x.pk]),
        "delete_label": "O'chirish",
    }


def kassa_entry_detail(request, kind, pk):
    """The inside of one kassa row: what it is, what it moved, and the buttons that
    change or remove it.

    The ledger rows themselves carry only [tahrirlash] and [ko'rish] — deleting lives
    one step in, behind this panel, so a mis-click in a dense table can't erase money.
    Every builder scopes its queryset the same way the edit/delete views do, so a
    seller can only open their own rows."""
    builders = {
        "payment": lambda: _entry_payment(request, pk),
        "advance": lambda: _entry_payment(request, pk),
        # Read-only: the row is derived, so it gets its own panel with no buttons.
        "advance_used": lambda: _entry_advance_used(request, pk),
        "commission": lambda: _entry_payment(request, pk, fee_only=True),
        "refund": lambda: _entry_refund(request, pk),
        "advance_out": lambda: _entry_advance_out(request, pk),
        "expense": lambda: _entry_expense(request, pk),
        "remittance": lambda: _entry_remittance(request, pk),
        "remittance_back": lambda: _entry_remittance(request, pk),
        "profit": lambda: _entry_profit(request, pk),
    }
    build = builders.get(kind)
    if build is None:
        raise Http404("Noma'lum amal turi")
    context = build()
    template = (
        "crm/_entry_detail_modal.html" if is_ajax(request) else "crm/entry_detail.html"
    )
    return render(request, template, context)


# --- Xodimlar (payroll) ------------------------------------------------------

# The category a wage payout is filed under by default. It is only a suggestion —
# `Expense.category` stays free text; what actually ties money to a worker is the
# `employee` field, so an advance filed as "Boshqa" still counts against their pay.
SALARY_CATEGORY = "Oylik / xodim"


def _month_shift(year, month, step):
    """The month `step` places away, rolling the year over in either direction."""
    index = year * 12 + (month - 1) + step
    return index // 12, index % 12 + 1


def _month_options(year, month, today, back=12):
    """Months offered in the picker: a rolling year back from today, newest first.
    A month reached by hand-editing ?oy= is folded in so the select never renders
    blank on a URL that the view itself accepts."""
    months = [_month_shift(today.year, today.month, -step) for step in range(back)]
    if (year, month) not in months:
        months.append((year, month))
        months.sort(reverse=True)
    return [{"value": f"{y:04d}-{m:02d}", "label": uz_month(y, m)} for y, m in months]


def _payroll_month(request):
    """The month the Xodimlar page is showing, from ?oy=YYYY-MM (this month by
    default). A wage is a monthly figure, so every total on the page is scoped to one
    calendar month rather than the shared day-range filter."""
    today = timezone.localdate()
    raw = request.GET.get("oy", "")
    try:
        year, month = (int(part) for part in raw.split("-", 1))
        date(year, month, 1)  # rejects month 0/13 and any junk
    except (ValueError, TypeError):
        year, month = today.year, today.month
    return year, month


def _rate_at(rates, first_day, fallback):
    """The wage in force on `first_day`, from a worker's (effective_from, amount) list
    given oldest first. Falls back to their current wage for months before the first
    rate row — which is exactly what a start month moved backwards leaves behind."""
    amount = fallback
    for effective_from, value in rates:
        if effective_from > first_day:
            break
        amount = value
    return amount


def _payroll_rows(employees, year, month):
    """One row per worker for the chosen month: what rode in unpaid from earlier
    months, what this month adds, what the till has paid against it, and what rides on.

    A month does not close itself. An unpaid remainder is still owed in September, and
    an advance drawn beyond the wage is still drawn — so the figure that matters is
    cumulative: the opening balance, plus every month's wage, minus everything paid,
    from `start_month` through the month on screen.

    Two queries carry the whole payroll however deep the history runs: wage rates and
    payouts are each fetched once and folded together in Python. Payouts dated BEFORE a
    worker's start month are deliberately left out — that stretch is what the opening
    balance already summarises, and counting it again would credit the firm twice for
    money it settled before the CRM was watching."""
    ids = [e.pk for e in employees]
    target = date(year, month, 1)
    next_year, next_month = _month_shift(year, month, 1)
    drawn_rows = (
        Expense.objects.filter(
            employee_id__in=ids,
            counts_against_salary=True,
            date__lt=date(next_year, next_month, 1),
        )
        .annotate(m=TruncMonth("date"))
        .values("employee_id", "m")
        .annotate(total=Sum("amount"))
    )
    drawn = {
        (r["employee_id"], r["m"].year, r["m"].month): r["total"] for r in drawn_rows
    }
    rates = {}
    for rate in SalaryRate.objects.filter(employee_id__in=ids).order_by("effective_from"):
        rates.setdefault(rate.employee_id, []).append((rate.effective_from, rate.amount))

    rows = []
    for e in employees:
        own_rates = rates.get(e.pk, [])
        # What the till paid in the month on screen — read straight from the payouts,
        # never from the accumulation. A month before the account opened still shows
        # what was handed over in it, or the figure here would contradict the payout
        # list printed directly underneath.
        paid = drawn.get((e.pk, year, month), Decimal("0"))
        tracked = target >= e.start_month.replace(day=1)
        if not tracked:
            rows.append({
                "employee": e, "carried": None, "salary": None, "due": None,
                "paid": paid, "remaining": None, "accrues": False, "tracked": False,
            })
            continue
        carried = e.opening_balance
        for y, m in month_span(e.start_month, target):
            if (y, m) == (year, month):
                continue                      # the month on screen is not "carried in"
            first = date(y, m, 1)
            wage = (
                _rate_at(own_rates, first, e.salary)
                if e.accrues_in(y, m) else Decimal("0")
            )
            carried += wage - drawn.get((e.pk, y, m), Decimal("0"))
        accrues = e.accrues_in(year, month)
        salary = _rate_at(own_rates, target, e.salary) if accrues else Decimal("0")
        rows.append({
            "employee": e,
            "carried": carried,                     # o'tgan oylardan qolgan
            "salary": salary,                       # shu oy oyligi
            "due": carried + salary,                # jami olishi kerak
            "paid": paid,                           # shu oy berilgan
            "remaining": carried + salary - paid,   # kelasi oyga o'tadi
            "accrues": accrues,
            "tracked": True,
        })
    return rows


def _payroll_employees(request):
    """The workers the Xodimlar page is showing: the name search and the
    Hammasi / Faol / Faol emas switch. Shared by the page and its Excel export so the
    file always holds exactly the rows on screen."""
    employees = Employee.objects.all()
    q = request.GET.get("q", "").strip()
    if q:
        employees = employees.filter(Q(name__icontains=q) | Q(note__icontains=q))
    status = request.GET.get("holat", "")
    if status == "faol":
        employees = employees.filter(is_active=True)
    elif status == "nofaol":
        employees = employees.filter(is_active=False)
    else:
        status = ""
    return list(employees), q, status


def employee_list(request):
    """Payroll (Xodimlar oyligi): everyone on the books, the wage in force this month,
    what rode in unpaid from earlier months, what the till has paid out, and what
    carries on to the next month.

    Open to every role. Wages were admin-only, but the sellers are the ones handing the
    cash over and being asked "how much of mine is left?" — and filing the till outflow
    that pays a wage was already theirs to do, so the figure it is measured against was
    the one thing they could not see."""
    year, month = _payroll_month(request)
    employees, q, status = _payroll_employees(request)
    rows = _payroll_rows(employees, year, month)
    # A worker whose account opens later has no figures for this month — only what the
    # till happened to pay them. They are left out of every total but the paid one.
    counted = [r for r in rows if r["tracked"]]
    active = [r for r in counted if r["employee"].is_active]
    totals = {
        "salary": sum((r["salary"] for r in active), Decimal("0")),
        "carried": sum((r["carried"] for r in counted), Decimal("0")),
        "due": sum((r["due"] for r in counted), Decimal("0")),
        "paid": sum((r["paid"] for r in rows), Decimal("0")),
        "remaining": sum((r["remaining"] for r in counted), Decimal("0")),
    }
    # The two ledgers below follow the search and the switch as well: a filtered page
    # whose ledgers still listed everybody would not add up to its own KPI cards.
    tagged = (
        Expense.objects.filter(
            employee__in=employees, date__year=year, date__month=month
        )
        .select_related("employee", "created_by")
        .order_by("-date", "-created_at")
    )
    # Two different things wear the same tag, so they get two tables: money paid TO a
    # worker (wage or advance), and money a worker spent FOR the business. Mixed into
    # one list, "kim qancha oldi" could not be read off the page at a glance.
    payouts = [e for e in tagged if e.counts_against_salary]
    errands = [e for e in tagged if not e.counts_against_salary]
    today = timezone.localdate()
    prev_year, prev_month = _month_shift(year, month, -1)
    next_year, next_month = _month_shift(year, month, 1)
    # Counted over the whole payroll, not the filtered rows, so the switch always says
    # how many people each side holds — including the side you are not looking at.
    everyone = Employee.objects.all()
    active_count = sum(1 for e in everyone if e.is_active)
    segments = [
        _segment(request, "holat", code, label, count, status)
        for code, label, count in (
            ("", "Hammasi", len(everyone)),
            ("faol", "Faol", active_count),
            ("nofaol", "Faol emas", len(everyone) - active_count),
        )
    ]
    export_qs = request.GET.urlencode()
    return render(request, "crm/employee_list.html", {
        "rows": rows,
        "q": q,
        "holat": status,
        "segments": segments,
        "export_url": reverse("employee_export") + (f"?{export_qs}" if export_qs else ""),
        "totals": totals,
        "payouts": payouts,
        "errands": errands,
        "payout_total": sum((e.amount for e in payouts), Decimal("0")),
        "errand_total": sum((e.amount for e in errands), Decimal("0")),
        "month_value": f"{year:04d}-{month:02d}",
        "month_label": uz_month(year, month),
        "prev_label": uz_month(prev_year, prev_month),
        "month_options": _month_options(year, month, today),
        "prev_month": f"{prev_year:04d}-{prev_month:02d}",
        "next_month": f"{next_year:04d}-{next_month:02d}",
        "is_current_month": (year, month) == (today.year, today.month),
    })


def employee_export(request):
    """Excel (.xlsx) of the payroll month as filtered — one row per worker, the same
    figures the page shows."""
    year, month = _payroll_month(request)
    employees, _, _ = _payroll_employees(request)
    rows = _payroll_rows(employees, year, month)
    headers = [
        "Xodim", "Holat", "Hisob boshlangan", "Oyligi (so'm)", "O'tgan oydan (so'm)",
        "Jami olishi kerak (so'm)", "Shu oy berilgan (so'm)", "Qolgan (so'm)", "Izoh",
    ]
    data = [
        [
            r["employee"].name,
            "Faol" if r["employee"].is_active else "Faol emas",
            r["employee"].start_month.strftime("%m.%Y"),
            float(r["salary"] or 0),
            float(r["carried"] or 0),
            float(r["due"] or 0),
            float(r["paid"]),
            float(r["remaining"] or 0),
            r["employee"].note,
        ]
        for r in rows
    ]
    number_formats = {i: "#,##0.00" for i in range(4, 9)}
    return _xlsx_response(
        f"xodimlar-{year:04d}-{month:02d}.xlsx",
        uz_month(year, month), headers, data, number_formats,
    )


def _employee_history(employee, today):
    """Month by month for one worker: the wage that month, what the till handed over,
    and the balance rolling forward. The same accumulation `_payroll_rows` does across
    the whole payroll, told for one person from their opening balance on — so the
    bottom line here and the row on the payroll page can never disagree."""
    drawn_rows = (
        Expense.objects.filter(employee=employee, counts_against_salary=True)
        .annotate(m=TruncMonth("date"))
        .values("m")
        .annotate(total=Sum("amount"))
    )
    drawn = {(r["m"].year, r["m"].month): r["total"] for r in drawn_rows}
    rates = [
        (rate.effective_from, rate.amount)
        for rate in employee.rates.order_by("effective_from")
    ]
    # Runs to today, or further if money was handed over in a later month.
    last = max([today.replace(day=1)] + [date(y, m, 1) for y, m in drawn])
    months = []
    balance = employee.opening_balance
    for y, m in month_span(employee.start_month, last):
        accrues = employee.accrues_in(y, m)
        wage = _rate_at(rates, date(y, m, 1), employee.salary) if accrues else Decimal("0")
        paid = drawn.get((y, m), Decimal("0"))
        balance += wage - paid
        months.append({
            "value": f"{y:04d}-{m:02d}",
            "label": uz_month(y, m),
            "salary": wage,
            "accrues": accrues,
            "paid": paid,
            "balance": balance,
        })
    months.reverse()                      # newest first, like every other ledger
    return months, balance


def employee_detail(request, pk):
    """One worker's whole file: what they are owed now, the month-by-month history
    behind that figure, every payout, and everything they spent for the business.

    The payroll page answers "who gets what this month"; this one answers the question
    that follows it — "and what has happened with this person all along" — which no
    amount of stepping through months one at a time could show."""
    employee = get_object_or_404(Employee, pk=pk)
    today = timezone.localdate()
    months, balance = _employee_history(employee, today)
    payouts = (
        employee.expenses.filter(counts_against_salary=True)
        .select_related("created_by")
        .order_by("-date", "-created_at")
    )
    errands = (
        employee.expenses.filter(counts_against_salary=False)
        .select_related("created_by")
        .order_by("-date", "-created_at")
    )
    this_month = next(
        (m for m in months if m["value"] == f"{today.year:04d}-{today.month:02d}"), None
    )
    return render(request, "crm/employee_detail.html", {
        "employee": employee,
        "months": months,
        "balance": balance,
        "this_month": this_month,
        "payouts": payouts,
        "errands": errands,
        "payout_total": sum((e.amount for e in payouts), Decimal("0")),
        "errand_total": sum((e.amount for e in errands), Decimal("0")),
        "month_value": f"{today.year:04d}-{today.month:02d}",
    })


def _set_salary_rate(employee, amount, effective_from, user):
    """Record what the wage is from `effective_from` on. Re-setting the wage for a
    month that already has a rate overwrites it rather than adding a second one — the
    month can only have had one wage, and a correction made minutes later is a fix, not
    a raise."""
    SalaryRate.objects.update_or_create(
        employee=employee,
        effective_from=effective_from.replace(day=1),
        defaults={"amount": amount, "created_by": user},
    )


@transaction.atomic
def employee_create(request):
    """Add someone to the payroll. Open to every role, like the rest of this page."""
    form = EmployeeForm(request.POST or None)
    title = "Yangi xodim"
    if request.method == "POST":
        if form.is_valid():
            employee = form.save()
            # The opening wage runs from the month the account opens, so the very
            # first month already has a dated rate to read.
            _set_salary_rate(
                employee, employee.salary, employee.start_month, request.user
            )
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Xodim", employee.pk,
                f"{employee.name} qo'shildi — oyligi {employee.salary:,.0f} so'm "
                f"({uz_month(employee.start_month.year, employee.start_month.month)}dan)",
            )
            messages.success(request, f"“{employee.name}” qo'shildi.")
            return form_success(request, reverse("employee_list"))
        return form_response(request, form, title, invalid=True)
    return form_response(request, form, title)


@transaction.atomic
def employee_edit(request, pk):
    """Edit a payroll worker. A changed wage is dated rather than swapped in: it takes
    effect from the month the form asks for (this month by default), leaving every
    month already settled priced as it was agreed."""
    employee = get_object_or_404(Employee, pk=pk)
    was = employee.salary
    form = EmployeeForm(request.POST or None, instance=employee)
    title = "Xodimni tahrirlash"
    if request.method == "POST":
        if form.is_valid():
            form.save()
            # Back on the payroll: the leaving month goes, so the wage accrues again.
            if employee.is_active and employee.end_month:
                employee.end_month = None
                employee.save(update_fields=["end_month"])
            since = form.cleaned_data.get("salary_from") or employee.start_month
            changed = employee.salary != was
            if changed:
                _set_salary_rate(employee, employee.salary, since, request.user)
            elif not employee.rates.exists():
                # Nothing changed, but an older record may carry no rate at all —
                # give it one so its months stop falling back to today's figure.
                _set_salary_rate(
                    employee, employee.salary, employee.start_month, request.user
                )
            summary = f"{employee.name} yangilandi — oyligi {employee.salary:,.0f} so'm"
            if changed:
                summary += (
                    f" (oldin {was:,.0f}; {uz_month(since.year, since.month)}dan)"
                )
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Xodim", employee.pk, summary
            )
            messages.success(request, "Xodim yangilandi.")
            return form_success(request, reverse("employee_list"))
        return form_response(request, form, title, invalid=True)
    return form_response(request, form, title)


def employee_delete(request, pk):
    """Remove someone from the payroll. Refused once money has been paid to them —
    erasing the worker would orphan a real till outflow, so they are deactivated
    instead (they keep their history and stop appearing in the pickers)."""
    employee = get_object_or_404(Employee, pk=pk)
    paid_ever = employee.expenses.exists()
    if request.method == "POST":
        if paid_ever:
            employee.is_active = False
            # Stop the wage accruing from next month on: someone who left in March
            # must not keep earning through December. This month still counts — they
            # worked it, and what is still owed for it stays owed.
            employee.end_month = timezone.localdate().replace(day=1)
            employee.save(update_fields=["is_active", "end_month"])
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Xodim", employee.pk,
                f"{employee.name} faol emas deb belgilandi",
            )
            messages.success(request, f"“{employee.name}” faol emas deb belgilandi.")
        else:
            name = employee.name
            employee.delete()
            AuditLog.record(
                request.user, AuditLog.Action.DELETE, "Xodim", pk, f"{name} o'chirildi"
            )
            messages.success(request, f"“{name}” o'chirildi.")
        return form_reload(request, reverse("employee_list"))
    if paid_ever:
        body = (
            f"“{employee.name}”ga kassadan pul berilgan, shuning uchun butunlay "
            f"o'chirib bo'lmaydi — chiqim yozuvlari egasiz qolardi. Uning o'rniga "
            f"faol emas deb belgilanadi: tarixi saqlanadi, ro'yxatlarda chiqmaydi. "
            f"Davom etasizmi?"
        )
        button = "Ha, faol emas qilish"
    else:
        body = f"“{employee.name}” ro'yxatdan o'chiriladi. Davom etasizmi?"
        button = "Ha, o'chirish"
    return render_confirm(
        request, "Xodimni o'chirish", body, button, confirm_class="btn-danger"
    )


def _remit_summary(remit):
    verb = "ishlab chiqarishdan qaytarib oldi" if remit.is_refund else "topshirdi"
    return (
        f"Sotuvchi {remit.seller} {verb} "
        f"({remit.get_method_display()}) — {remit.abs_amount:,.0f} so'm"
    )


# The three questions shown above the Saqlash button on a handover — in its own file
# so the modal and the full page carry the same warning.
CHECKS = "crm/_remittance_checks.html"


def remittance_create(request):
    """Record cash a seller hands back to production. A seller may only file their
    own; admins/managers may file on behalf of any seller (and can preselect one via
    ?seller= from the per-seller control table)."""
    initial = {}
    seller_pk = request.GET.get("seller", "")
    if request.method == "GET" and request.user.can_see_all_records and seller_pk.isdigit():
        initial["seller"] = seller_pk
    form = ProductionRemittanceForm(request.POST or None, user=request.user, initial=initial)
    title = "Ishlab chiqarishga topshirish"
    if request.method == "POST":
        if form.is_valid():
            remit = form.save(commit=False)
            # A seller cannot spoof the seller field (it's disabled, so absent from
            # POST) — pin it to themselves.
            if not request.user.can_see_all_records:
                remit.seller = request.user
            remit.created_by = request.user
            remit.save()
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Topshiruv", remit.pk,
                _remit_summary(remit),
            )
            messages.success(request, f"Topshirildi: {remit.amount:,.0f} so'm.")
            return form_success(request, reverse("kassa"))
        return form_response(
            request, form, title, invalid=True,
            modal_template="crm/_remittance_modal.html", checks_template=CHECKS,
        )
    return form_response(
        request, form, title,
        modal_template="crm/_remittance_modal.html", checks_template=CHECKS,
    )


def remittance_refund_create(request):
    """Record cash production hands BACK to a seller — a handover in reverse. Same
    permissions as filing one: a seller may only file their own, admins/managers may
    file for anyone (?seller= preselects from the per-seller control table).

    Stored as a negative handover, so the seller's till and their production debt both
    climb back by the amount returned."""
    initial = {}
    seller_pk = request.GET.get("seller", "")
    if request.method == "GET" and request.user.can_see_all_records and seller_pk.isdigit():
        initial["seller"] = seller_pk
    form = ProductionRefundForm(request.POST or None, user=request.user, initial=initial)
    title = "Ishlab chiqarishdan qaytarish"
    if request.method == "POST":
        if form.is_valid():
            remit = form.save(commit=False)
            # A seller cannot spoof the seller field (it's disabled, so absent from
            # POST) — pin it to themselves.
            if not request.user.can_see_all_records:
                remit.seller = request.user
            remit.created_by = request.user
            remit.save()
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Topshiruv", remit.pk,
                _remit_summary(remit),
            )
            messages.success(request, f"Qaytarildi: {remit.abs_amount:,.0f} so'm.")
            return form_success(request, reverse("kassa"))
        return form_response(request, form, title, invalid=True, modal_template="crm/_refund_modal.html")
    return form_response(request, form, title, modal_template="crm/_refund_modal.html")


def remittance_edit(request, pk):
    """Fix a mistaken handover — or a mistaken return, which is the same row with a
    negative amount and so gets the return form. Admins/managers may edit any; a
    seller may edit only their own."""
    qs = ProductionRemittance.objects.all() if request.user.can_see_all_records \
        else ProductionRemittance.objects.filter(seller=request.user)
    remit = get_object_or_404(qs, pk=pk)
    refund = remit.is_refund
    title = "Qaytarishni tahrirlash" if refund else "Topshiruvni tahrirlash"
    modal = "crm/_refund_modal.html" if refund else "crm/_remittance_modal.html"
    form_class = ProductionRefundForm if refund else ProductionRemittanceForm
    form = form_class(request.POST or None, instance=remit, user=request.user)
    if request.method == "POST":
        if form.is_valid():
            remit = form.save(commit=False)
            if not request.user.can_see_all_records:
                remit.seller = request.user
            remit.save()
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Topshiruv", remit.pk,
                _remit_summary(remit),
            )
            messages.success(
                request, "Qaytarish yangilandi." if refund else "Topshiruv yangilandi."
            )
            return form_success(request, reverse("kassa"))
        return form_response(
            request, form, title, invalid=True, modal_template=modal,
            checks_template=None if refund else CHECKS,
        )
    return form_response(
        request, form, title, modal_template=modal,
        checks_template=None if refund else CHECKS,
    )


def remittance_delete(request, pk):
    """Remove a mistaken handover (or return). Admins/managers may erase any; a seller
    may erase only their own."""
    qs = ProductionRemittance.objects.select_related("seller", "created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(seller=request.user)
    remit = get_object_or_404(qs, pk=pk)
    noun = "qaytarish" if remit.is_refund else "topshiruv"
    if request.method == "POST":
        summary = _remit_summary(remit)
        remit.delete()
        AuditLog.record(request.user, AuditLog.Action.DELETE, "Topshiruv", pk, summary)
        messages.success(request, f"{noun.capitalize()} o'chirildi.")
        return form_reload(request, reverse("kassa"))
    return render_confirm(
        request,
        f"{noun.capitalize()}ni o'chirish",
        f"{remit.abs_amount:,.0f} so'm {noun} o'chiriladi. Davom etasizmi?",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


def _profit_summary(payout):
    return (
        f"Sotuvchi {payout.seller} foyda topshirdi "
        f"({payout.get_method_display()}) — {payout.amount:,.0f} so'm"
    )


def profit_payout_create(request):
    """Record realized profit a seller hands up to the boss. A seller may file only
    their own; admins/managers may file for any seller (and can preselect one via
    ?seller= from the per-seller control table)."""
    initial = {}
    seller_pk = request.GET.get("seller", "")
    if request.method == "GET" and request.user.can_see_all_records and seller_pk.isdigit():
        initial["seller"] = seller_pk
    form = ProfitPayoutForm(request.POST or None, user=request.user, initial=initial)
    title = "Foyda topshirish"
    if request.method == "POST":
        if form.is_valid():
            payout = form.save(commit=False)
            if not request.user.can_see_all_records:
                payout.seller = request.user
            payout.created_by = request.user
            payout.save()
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Foyda", payout.pk,
                _profit_summary(payout),
            )
            messages.success(request, f"Foyda topshirildi: {payout.amount:,.0f} so'm.")
            return form_success(request, reverse("kassa"))
        return form_response(request, form, title, invalid=True, modal_template="crm/_profit_payout_modal.html")
    return form_response(request, form, title, modal_template="crm/_profit_payout_modal.html")


def profit_payout_edit(request, pk):
    """Fix a mistaken profit handover. Admins/managers may edit any; a seller may edit
    only their own."""
    qs = ProfitPayout.objects.all() if request.user.can_see_all_records \
        else ProfitPayout.objects.filter(seller=request.user)
    payout = get_object_or_404(qs, pk=pk)
    title = "Foyda topshiruvini tahrirlash"
    form = ProfitPayoutForm(request.POST or None, instance=payout, user=request.user)
    if request.method == "POST":
        if form.is_valid():
            payout = form.save(commit=False)
            if not request.user.can_see_all_records:
                payout.seller = request.user
            payout.save()
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Foyda", payout.pk,
                _profit_summary(payout),
            )
            messages.success(request, "Foyda topshiruvi yangilandi.")
            return form_success(request, reverse("kassa"))
        return form_response(request, form, title, invalid=True, modal_template="crm/_profit_payout_modal.html")
    return form_response(request, form, title, modal_template="crm/_profit_payout_modal.html")


def profit_payout_delete(request, pk):
    """Remove a mistaken profit handover. Admins/managers may erase any; a seller may
    erase only their own."""
    qs = ProfitPayout.objects.select_related("seller", "created_by")
    if not request.user.can_see_all_records:
        qs = qs.filter(seller=request.user)
    payout = get_object_or_404(qs, pk=pk)
    if request.method == "POST":
        summary = _profit_summary(payout)
        payout.delete()
        AuditLog.record(request.user, AuditLog.Action.DELETE, "Foyda", pk, summary)
        messages.success(request, "Foyda topshiruvi o'chirildi.")
        return form_reload(request, reverse("kassa"))
    return render_confirm(
        request,
        "Foyda topshiruvini o'chirish",
        f"{payout.amount:,.0f} so'm foyda topshiruvi o'chiriladi. Davom etasizmi?",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


def _receipt_summary(receipt):
    lines = list(receipt.items.select_related("product"))
    total = sum((it.quantity_kg for it in lines), Decimal("0"))
    return f"Sotuvchi {receipt.seller}: {len(lines)} ta mahsulot, {_kg(total)} kg"


def _pending_zakaz_for_receipt(receipt):
    """Pending zakaz lines the receipt's stock could fulfil: the seller's own
    unfulfilled sale lines for any product on the receipt."""
    product_ids = list(receipt.items.values_list("product_id", flat=True))
    return (
        SaleItem.objects.filter(
            sale__sales_rep=receipt.seller,
            product_id__in=product_ids,
            fulfilled_at__isnull=True,
        )
        .select_related("sale", "sale__client", "product")
        .order_by("sale__date")
    )


def _auto_bind_receipt(receipt):
    """Automatically fulfil the seller's oldest pending zakaz for each received
    product, up to the quantity received. Strict FIFO and whole-order: the oldest
    unfilled order that the remaining stock can't cover stops that product's binding
    (it waits for more stock). Returns the number of orders bound."""
    bound = 0
    for ri in receipt.items.select_related("product"):
        remaining = ri.quantity_kg
        pending = (
            SaleItem.objects.filter(
                sale__sales_rep=receipt.seller,
                product=ri.product,
                fulfilled_at__isnull=True,
            )
            .select_related("sale")
            .order_by("sale__date", "pk")
        )
        for item in pending:
            need = item.weight_kg - item.fulfilled_kg
            if need <= 0:
                continue
            fill = min(need, remaining)
            item.fulfilled_kg += fill
            fields = ["fulfilled_kg"]
            if item.fulfilled_kg >= item.weight_kg:  # this order is now complete
                item.fulfilled_at = receipt.date
                item.fulfilled_by_receipt = receipt
                fields += ["fulfilled_at", "fulfilled_by_receipt"]
            item.save(update_fields=fields)
            remaining -= fill
            bound += 1
            if remaining <= 0:
                break
    return bound


def _render_receipt_form(request, form, formset, title, invalid=False):
    context = {"form": form, "formset": formset, "title": title}
    if is_ajax(request):
        return render(request, "crm/_receipt_modal.html", context, status=422 if invalid else 200)
    return render(request, "crm/receipt_form.html", context)


def receipt_create(request):
    """Log goods a seller received from production into their ombor. A seller files
    only their own; admins/managers may file for any seller (preselect via ?seller=)."""
    initial = {}
    seller_pk = request.GET.get("seller", "")
    if request.method == "GET" and request.user.can_see_all_records and seller_pk.isdigit():
        initial["seller"] = seller_pk
    form = ProductionReceiptForm(request.POST or None, user=request.user, initial=initial)
    formset = ProductionReceiptItemFormSet(
        request.POST or None, instance=ProductionReceipt(), prefix="items"
    )
    title = "Ishlab chiqarishdan qabul"
    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            receipt = form.save(commit=False)
            if not request.user.can_see_all_records:
                receipt.seller = request.user
            receipt.created_by = request.user
            receipt.save()
            formset.instance = receipt
            formset.save()
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Qabul", receipt.pk,
                _receipt_summary(receipt),
            )
            messages.success(request, "Qabul qilingan tovarlar qo'shildi.")
            # Automatically assign the arriving stock to waiting zakaz orders.
            bound = _auto_bind_receipt(receipt)
            if bound:
                AuditLog.record(
                    request.user, AuditLog.Action.UPDATE, "Zakaz", receipt.pk,
                    f"{bound} ta zakaz avtomatik biriktirildi",
                )
                messages.success(request, f"{bound} ta zakaz mijozga avtomatik biriktirildi.")
            return form_success(request, reverse("ombor"))
        return _render_receipt_form(request, form, formset, title, invalid=True)
    return _render_receipt_form(request, form, formset, title)


def receipt_bind(request, pk):
    """Bind a receipt's arriving stock to pending zakaz orders for the same
    products, marking those orders fulfilled (ready to hand over)."""
    qs = ProductionReceipt.objects.all() if request.user.can_see_all_records \
        else ProductionReceipt.objects.filter(seller=request.user)
    receipt = get_object_or_404(qs, pk=pk)
    pending = _pending_zakaz_for_receipt(receipt)
    if request.method == "POST":
        ids = request.POST.getlist("bind")
        n = pending.filter(pk__in=ids).update(
            fulfilled_at=receipt.date, fulfilled_by_receipt=receipt
        )
        if n:
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Zakaz", receipt.pk,
                f"{n} ta zakaz mijozga biriktirildi",
            )
            messages.success(request, f"{n} ta zakaz biriktirildi.")
        return redirect(reverse("ombor"))
    return render(request, "crm/receipt_bind.html", {"receipt": receipt, "pending": pending})


def receipt_edit(request, pk):
    """Fix a receipt. Admins/managers may edit any; a seller only their own."""
    qs = ProductionReceipt.objects.all() if request.user.can_see_all_records \
        else ProductionReceipt.objects.filter(seller=request.user)
    receipt = get_object_or_404(qs, pk=pk)
    form = ProductionReceiptForm(request.POST or None, instance=receipt, user=request.user)
    formset = ProductionReceiptItemFormSet(request.POST or None, instance=receipt, prefix="items")
    title = "Qabulni tahrirlash"
    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            receipt = form.save(commit=False)
            if not request.user.can_see_all_records:
                receipt.seller = request.user
            receipt.save()
            formset.save()
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Qabul", receipt.pk,
                _receipt_summary(receipt),
            )
            messages.success(request, "Qabul yangilandi.")
            return form_reload(request, reverse("ombor"))
        return _render_receipt_form(request, form, formset, title, invalid=True)
    return _render_receipt_form(request, form, formset, title)


def receipt_delete(request, pk):
    """Remove a receipt. Admins/managers may erase any; a seller only their own."""
    qs = ProductionReceipt.objects.select_related("seller")
    if not request.user.can_see_all_records:
        qs = qs.filter(seller=request.user)
    receipt = get_object_or_404(qs, pk=pk)
    if request.method == "POST":
        summary = _receipt_summary(receipt)
        receipt.delete()
        AuditLog.record(request.user, AuditLog.Action.DELETE, "Qabul", pk, summary)
        messages.success(request, "Qabul o'chirildi.")
        return form_reload(request, reverse("ombor"))
    return render_confirm(
        request,
        "Qabulni o'chirish",
        "Bu qabul o'chiriladi va sotuvchi ombori shunga mos kamayadi. Davom etasizmi?",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


# The ombor report opens on the whole current month, not on today like the other
# dated pages: it is read as a monthly sverka, and a single day's slice is usually
# empty. A whole month also lets the toolbar name the window ("Avgust").
OMBOR_DEFAULT_WINDOW = "month"


def _ombor_items(request, date_from, date_to):
    """Sale lines inside the window, scoped to the viewer and narrowed by the ombor
    filters (product search + seller). Shared by the page and its Excel export so
    the download always matches what is on screen.
    Returns (items, filters, reps, rep_obj)."""
    user = request.user
    filters = {"q": request.GET.get("q", "").strip(), "rep": request.GET.get("rep", "")}
    filters["dan"] = date_from.isoformat()
    filters["gacha"] = date_to.isoformat()

    items = SaleItem.objects.filter(sale__date__gte=date_from, sale__date__lte=date_to)
    if not user.can_see_all_records:
        items = items.filter(sale__sales_rep=user)

    reps = rep_obj = None
    if user.can_see_all_records:
        reps = User.objects.filter(is_active=True).order_by(
            "first_name", "last_name", "username"
        )
        if filters["rep"].isdigit():
            rep_obj = reps.filter(pk=filters["rep"]).first()
            if rep_obj:
                items = items.filter(sale__sales_rep=rep_obj)

    if filters["q"]:
        items = items.filter(
            Q(product__name__icontains=filters["q"]) | Q(product__sku__icontains=filters["q"])
        )
    return items, filters, reps, rep_obj


def _ombor_rows(items):
    """One row per product: kg sold, and how many receipts it appeared on."""
    return list(
        items.values("product", "product__name", "product__sku")
        .annotate(total_kg=Sum(ITEM_WEIGHT_KG), sales_count=Count("sale", distinct=True))
        .order_by("-total_kg")
    )


def ombor_view(request):
    """Ombor = sold-goods report, one row per product with the total kg sold in the
    selected date window. A seller sees only their own sales; admins/managers see
    every seller's combined total (and can filter to one seller). Click a product to
    drill into who bought it. Mirrors the debts page's group-then-detail shape."""
    dates = _date_range_context(request, OMBOR_DEFAULT_WINDOW)
    date_from, date_to = dates["date_from"], dates["date_to"]

    items, filters, reps, rep_obj = _ombor_items(request, date_from, date_to)
    rows = _ombor_rows(items)
    total_kg = sum((r["total_kg"] or Decimal("0") for r in rows), Decimal("0"))

    active_filters = _filter_chips(request, [
        {"param": "rep", "label": "Sotuvchi", "value": str(rep_obj) if rep_obj else ""},
    ])
    # Carry the current window and filters into the download link, so the .xlsx is
    # exactly the table on screen (the month a sverka is being done for).
    query = request.GET.urlencode()

    return render(request, "crm/ombor.html", {
        "rows": rows,
        "total_kg": total_kg,
        "product_count": len(rows),
        "q": filters["q"],
        "filters": filters,
        "reps": reps,
        "rep_label": "Sotuvchi",
        "is_admin_view": request.user.can_see_all_records,
        "active_filters": active_filters,
        "filter_count": len(active_filters),
        "has_filters": bool(active_filters),
        "filter_url": reverse("ombor"),
        "catalog_url": reverse("product_list"),
        "export_url": reverse("ombor_export") + (f"?{query}" if query else ""),
        "search_placeholder": "Mahsulot nomi…",
        "show_daterange_picker": True,
        "keep_daterange": True,
        **dates,
    })


def ombor_export(request):
    """The sold-goods report as .xlsx — same window, search and seller filter as the
    page. This is the sheet a monthly production-vs-sold sverka is built from."""
    dates = _date_range_context(request, OMBOR_DEFAULT_WINDOW)
    items, _, _, _ = _ombor_items(request, dates["date_from"], dates["date_to"])
    rows = _ombor_rows(items)

    headers = ["Mahsulot", "SKU", "Sotuvlar soni", "Sotilgan (kg)"]
    data = [
        [
            r["product__name"],
            r["product__sku"],
            r["sales_count"],
            float(r["total_kg"] or 0),
        ]
        for r in rows
    ]
    return _xlsx_response("ombor.xlsx", "Ombor", headers, data, {4: "0.000"})


def _filter_ombor_items(request, items):
    """Narrow one product's sale lines by client / seller / date, so a particular chek
    can be found in a long history.

    Unlike the sales list there is NO default date window: this page IS the product's
    full history, so dates only bite once the user actually sets them. Returns
    (queryset, filters, has_filters)."""
    filters = {key: request.GET.get(key, "") for key in ("client", "rep", "dan", "gacha")}
    filters["q"] = request.GET.get("q", "").strip()
    can_scope_rep = request.user.can_see_all_records

    if filters["q"]:
        items = _client_search(items, filters["q"], "sale__client")
    if filters["client"].isdigit():
        items = items.filter(sale__client_id=filters["client"])
    if filters["rep"].isdigit() and can_scope_rep:
        items = items.filter(sale__sales_rep_id=filters["rep"])

    date_from = _parse_date(filters["dan"])
    date_to = _parse_date(filters["gacha"])
    if date_from and date_to and date_to < date_from:
        date_from, date_to = date_to, date_from
        filters["dan"], filters["gacha"] = date_from.isoformat(), date_to.isoformat()
    if date_from:
        items = items.filter(sale__date__gte=date_from)
    if date_to:
        items = items.filter(sale__date__lte=date_to)

    has_filters = bool(
        filters["q"]
        or filters["client"].isdigit()
        or (filters["rep"].isdigit() and can_scope_rep)
        or date_from
        or date_to
    )
    return items, filters, has_filters


def _ombor_product_items(request, product):
    """One product's sale lines for the current filters, newest first. Shared by the
    drill-down page and its Excel export, so the download matches the screen."""
    qs = SaleItem.objects.filter(product=product).select_related(
        "sale", "sale__client", "sale__sales_rep"
    )
    if not request.user.can_see_all_records:
        qs = qs.filter(sale__sales_rep=request.user)
    scoped = qs
    qs, filters, has_filters = _filter_ombor_items(request, qs)
    items = list(qs)
    items.sort(key=lambda it: (it.sale.date, it.sale.created_at), reverse=True)
    return items, scoped, filters, has_filters


def ombor_product(request, pk):
    """Drill-down for one product: every sale of it, newest first, filterable by
    client / seller / date so one chek can be tracked down. A seller sees only their
    own sales; an admin also gets a per-seller summary (which seller sold how much)
    above the transaction list."""
    product = get_object_or_404(Product, pk=pk)
    user = request.user
    items, scoped, filters, has_filters = _ombor_product_items(request, product)
    # Dropdown options come from this product's own history — offering clients who
    # never bought it would just be noise.
    clients = Client.objects.filter(sales__items__in=scoped).distinct().order_by("name")
    reps = (
        User.objects.filter(sales__items__in=scoped).distinct().order_by("first_name", "username")
        if user.can_see_all_records
        else None
    )
    # The KPI and the per-seller summary count every filtered line, not just the
    # page on screen — paging must not change what "jami sotilgan" means.
    total_kg = sum((it.weight_kg for it in items), Decimal("0"))

    by_seller = None
    if user.can_see_all_records:
        acc = {}
        for it in items:
            rep = it.sale.sales_rep
            row = acc.setdefault(rep.pk, {"seller": rep, "kg": Decimal("0"), "count": 0})
            row["kg"] += it.weight_kg
            row["count"] += 1
        by_seller = sorted(acc.values(), key=lambda r: r["kg"], reverse=True)

    chip_client = clients.filter(pk=filters["client"]).first() if filters["client"].isdigit() else None
    chip_rep = reps.filter(pk=filters["rep"]).first() if reps and filters["rep"].isdigit() else None
    active_filters = _filter_chips(request, [
        {"param": "client", "label": "Mijoz", "value": chip_client.name if chip_client else ""},
        {"param": "rep", "label": "Sotuvchi", "value": str(chip_rep) if chip_rep else ""},
        {"param": "dan", "label": "Sanadan", "value": filters["dan"]},
        {"param": "gacha", "label": "Sanagacha", "value": filters["gacha"]},
    ])
    export_qs = request.GET.urlencode()
    return render(request, "crm/ombor_product.html", {
        "product": product,
        "page": Paginator(items, 25).get_page(request.GET.get("page")),
        "total_count": len(items),
        "total_kg": total_kg,
        "by_seller": by_seller,
        "is_admin_view": user.can_see_all_records,
        "filters": filters,
        "has_filters": has_filters,
        "active_filters": active_filters,
        "filter_count": len(active_filters),
        "clients": clients,
        "reps": reps,
        "filter_url": reverse("ombor_product", args=[product.pk]),
        "export_url": reverse("ombor_product_export", args=[product.pk]) + (
            f"?{export_qs}" if export_qs else ""
        ),
    })


def ombor_product_export(request, pk):
    """Excel (.xlsx) of one product's sales — every filtered line, not just the page."""
    product = get_object_or_404(Product, pk=pk)
    items, _, _, _ = _ombor_product_items(request, product)
    headers = [
        "Sana", "Mijoz", "Sotuvchi", "Razmer / Mikron", "O'lchov",
        "Miqdori", "Miqdori (kg)", "Narxi (1 birlik)", "Umumiy narx",
    ]
    rows = [
        [
            it.sale.date.strftime("%d.%m.%Y"),
            it.sale.client.name,
            str(it.sale.sales_rep),
            it.variant_label,
            it.get_dimension_display(),
            float(it.weight),
            float(it.weight_kg),
            float(it.price),
            float(it.total_price),
        ]
        for it in items
    ]
    number_formats = {
        6: "0.000", 7: "0.000", 8: "#,##0.00", 9: "#,##0.00",
    }
    return _xlsx_response(
        f"mahsulot-{product.pk}-sotuvlar.xlsx", "Sotuvlar", headers, rows, number_formats
    )


def sale_export(request):
    base = (
        Sale.objects.visible_to(request.user)
        .real()  # matches the sales list — opening carry-overs are not sales
        .select_related("client", "sales_rep")
        .with_balance()
    )
    sales, _, _, _, _ = _filter_sales(request, base)
    sales = sales.order_by("-date", "-created_at").prefetch_related("items__product")

    headers = [
        "Sana", "Mijoz", "Mahsulot", "Sotuvchi", "O'lchov", "Og'irligi",
        "Narxi", "Umumiy narx", "Tannarx", "Foyda", "To'lov", "Qarz muddati",
    ]
    # One row per line item, so a multi-product receipt still exports cleanly.
    rows = []
    for s in sales:
        deadline = s.debt_deadline.strftime("%d.%m.%Y") if s.debt_deadline else ""
        status = "Qarz" if s.remaining > 0 else "To'langan"
        for item in s.items.all():
            rows.append([
                s.date.strftime("%d.%m.%Y"),
                s.client.name,
                item.product.name,
                str(s.sales_rep),
                item.get_dimension_display(),
                float(item.weight),
                float(item.price),
                float(item.total_price),
                float(item.total_cost),
                float(item.profit),
                status,
                deadline,
            ])
    number_formats = {6: "0.000", 7: "#,##0.00", 8: "#,##0.00", 9: "#,##0.00", 10: "#,##0.00"}
    return _xlsx_response("sotuvlar.xlsx", "Sotuvlar", headers, rows, number_formats)


def sale_detail(request, pk):
    sale = get_object_or_404(
        Sale.objects.visible_to(request.user)
        .select_related("client", "sales_rep")
        .prefetch_related("items__product"),
        pk=pk,
    )
    rows = sale.payments.select_related("created_by").order_by("-date", "-created_at")
    # Money the client put IN and money handed back to them are two different stories,
    # and mixing them would make the payments table stop adding up to `paid`.
    payments = [p for p in rows if p.kind in PAYING_KINDS]
    settlements = [p for p in rows if p.kind not in PAYING_KINDS]
    returns = sale.returns.select_related("product", "created_by")
    return render(
        request,
        "crm/sale_detail.html",
        {
            "sale": sale,
            "items": sale.items.all(),
            "payments": payments,
            "settlements": settlements,
            "returns": returns,
            "returned": sale.returned_amount,
            "settled": sale.settled_amount,
            "paid": sale.paid_amount,
            "remaining": sale.debt_remaining,
        },
    )


def _render_sale_form(
    request, form, formset, title, invalid=False, zakaz_shortfall=None, overpay=None
):
    context = {
        "form": form,
        "formset": formset,
        "title": title,
        "products_json": _product_price_map(),
        "client_advance_json": _client_advance_map(request.user),
        "zakaz_shortfall": zakaz_shortfall,
        "overpay": overpay,
        "size_suggestions": [c[0] for c in SIZE_CHOICES],
        "micron_suggestions": [c[0] for c in MICRON_CHOICES],
    }
    keep_open = invalid or bool(zakaz_shortfall) or bool(overpay)
    if is_ajax(request):
        return render(request, "crm/_sale_modal.html", context, status=422 if keep_open else 200)
    return render(request, "crm/sale_form.html", context)


def _client_advance_map(user):
    """{client_pk: balance} for clients this user (as seller) is holding an advance
    for — so the sale form can flag "this client has X prepaid" the moment they're
    picked. Scoped to `user` because that's the seller whose advance a new sale would
    actually consume. Only positive balances are included.

    Counts the same kinds as `client_advance_balance`, credit owed back from returns
    and price corrections included — otherwise the hint would contradict the balance
    the sale then actually draws on."""
    rows = (
        Payment.objects.filter(
            created_by=user,
            kind__in=ADVANCE_DEPOSIT_KINDS + ADVANCE_SPENT_KINDS,
        )
        .values("client")
        .annotate(
            deposited=Sum(PAYMENT_CREDIT, filter=Q(kind__in=ADVANCE_DEPOSIT_KINDS)),
            used=Sum(PAYMENT_CREDIT, filter=Q(kind__in=ADVANCE_SPENT_KINDS)),
        )
    )
    result = {}
    for r in rows:
        balance = (r["deposited"] or Decimal("0")) - (r["used"] or Decimal("0"))
        if r["client"] and balance > 0:
            result[str(r["client"])] = float(balance)
    return result


def _product_price_map():
    """Per-kg price/cost for each active product, so the form can auto-fill a row —
    plus whether the product offers the Razmer / Mikron dropdowns, so the JS can show
    or hide them when the product is picked."""
    return {
        str(p.pk): {
            "price": str(p.price),
            "cost": str(p.cost_price),
            "has_size": p.has_size,
            "has_micron": p.has_micron,
        }
        for p in Product.objects.filter(is_active=True)
    }


def sale_create(request):
    form = SaleForm(request.POST or None, user=request.user)
    formset = SaleItemFormSet(request.POST or None, instance=Sale(), prefix="items")
    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            sale = form.save(commit=False)
            sale.sales_rep = request.user
            sale.save()
            formset.instance = sale
            formset.save()
            # No warehouse stock to check against — every line is fulfilled on sale.
            _mark_fulfilment(sale, [])
            # If the client has prepaid this seller, spend that advance on the new
            # receipt (oldest first) — the sale opens already part/fully paid.
            applied = _apply_advance_to_open_sales(sale.client, request.user)
            AuditLog.record(
                request.user, AuditLog.Action.CREATE, "Sotuv", sale.pk,
                f"Mijoz {sale.client.name}, {sale.items.count()} ta mahsulot "
                f"— {sale.total_price:,.0f} so'm",
            )
            if applied > 0:
                messages.success(
                    request,
                    f"Sotuv qo'shildi. Avansdan {applied:,.0f} so'm yechildi.",
                )
            else:
                messages.success(request, "Sotuv qo'shildi (qarz sifatida).")
            return form_success(request, reverse("sale_list"))
        return _render_sale_form(request, form, formset, "Yangi sotuv", invalid=True)
    return _render_sale_form(request, form, formset, "Yangi sotuv")


def _formset_total(formset):
    """Revenue (weight × price) of the formset's surviving (non-deleted) items."""
    total = Decimal("0")
    for f in formset.forms:
        cleaned = getattr(f, "cleaned_data", None)
        if not cleaned or cleaned.get("DELETE"):
            continue
        weight = cleaned.get("weight")
        price = cleaned.get("price")
        if weight is not None and price is not None:
            total += weight * price
    return total


def _return_conflict(formset):
    """Reject edits that would strand a return.

    `Return.sale_item` cascades, so deleting a line that has returns would silently
    take the returns with it while the settlement payment stayed behind — the client's
    credit would survive with nothing backing it. Shrinking a line below what has
    already come back is the same problem in miniature. Returns the error message, or
    None when the edit is safe."""
    for f in formset.forms:
        cleaned = getattr(f, "cleaned_data", None)
        item = cleaned.get("id") if cleaned else None
        if not cleaned or item is None or item.pk is None:
            continue
        returned = sum((r.weight for r in item.returns.all()), Decimal("0"))
        if not returned:
            continue
        name = item.product.name
        if cleaned.get("DELETE"):
            return (
                f"«{name}» qatorini o'chirib bo'lmaydi — undan {returned:g} "
                f"{item.dimension} qaytarilgan. Avval qaytarishni bekor qiling."
            )
        weight = cleaned.get("weight")
        if weight is not None and weight < returned:
            return (
                f"«{name}» miqdorini {weight:g} ga tushirib bo'lmaydi — undan allaqachon "
                f"{returned:g} {item.dimension} qaytarilgan."
            )
    return None


def _weight_dropped(sale, formset):
    """True when the edit takes goods OFF the receipt instead of only re-pricing it.

    Both end in the same over-payment, but they are different events: if the goods
    physically came back this should have been a Return, which restocks them and
    relieves the seller of the tannarx. A price correction deliberately does neither,
    so the form warns rather than guessing which one the seller meant.

    Weights are read fresh from the database: `cleaned_data["id"]` hands back the very
    instance the formset has already overwritten with the new values, so comparing
    against `item.weight` would compare a number with itself."""
    original = dict(sale.items.values_list("pk", "weight"))
    for f in formset.forms:
        cleaned = getattr(f, "cleaned_data", None)
        item = cleaned.get("id") if cleaned else None
        if not cleaned or item is None or item.pk is None:
            continue
        was = original.get(item.pk)
        if was is None:
            continue
        if cleaned.get("DELETE"):
            return True
        weight = cleaned.get("weight")
        if weight is not None and weight < was:
            return True
    return False


def _settle_overpay(sale, excess, refunded, user):
    """Post the money side of a downward price correction — the mirror of
    `_settle_return`, without any goods moving.

    The client paid against a total that has now shrunk, so the difference is theirs:
    parked as advance credit (default — the cash never leaves the drawer, it is only
    re-labelled) or handed back as cash. Must run AFTER the new lines are saved, so the
    credit settles onto a receipt that already carries the corrected total.

    Deliberately always so'm / cash / no fee: this is not a payment the client made but
    a bookkeeping correction of one, and the till only ever knows it in so'm."""
    settlement = Payment.objects.create(
        sale=sale,
        client=sale.client,
        # Today, not the sale's date: the money moves now. A refund backdated into a
        # closed period would silently restate a till that has already been handed over.
        date=timezone.localdate(),
        amount=excess,
        amount_original=excess,
        currency=Payment.Currency.UZS,
        method=Payment.Method.CASH,
        commission=Decimal("0"),
        commission_percent=Decimal("0"),
        kind=Payment.Kind.ADJUST_REFUND if refunded else Payment.Kind.ADJUST_CREDIT,
        note="Narx tuzatildi",
        created_by=user,
    )
    if not refunded:
        # Spend the fresh credit on whatever else the client still owes.
        _apply_advance_to_open_sales(sale.client, user)
    return settlement


def _sale_edit_message(overpay, refunded):
    if overpay <= 0:
        return "Sotuv yangilandi."
    if refunded:
        return (
            f"Sotuv yangilandi. Ortiqcha {overpay:,.0f} so'm mijozga naqd qaytarildi "
            f"— kassadan chiqdi."
        )
    return (
        f"Sotuv yangilandi. Ortiqcha {overpay:,.0f} so'm mijoz avansiga o'tdi — "
        f"uning boshqa qarzlariga ishlatiladi."
    )


@transaction.atomic
def sale_edit(request, pk):
    """Edit a sale — including correcting a price that turns out to have been wrong
    after the client already paid.

    Dropping the total below what has been paid used to be refused outright, which left
    a genuine mistake with no way out. A Return is the wrong instrument here: the goods
    are still with the client, so it would shrink the sold kg, offer to restock stock
    that never came back, and still leave the tannarx wrong. "The goods came back" and
    "the price was wrong" are two different events.

    So the edit asks instead of refusing, and writes nothing until answered: park the
    over-payment as client credit (nothing leaves the till) or hand it back in cash.
    The old hard block survives as the thing that makes the question unskippable — a
    receipt must never be left sitting at a negative balance."""
    sale = get_object_or_404(Sale.objects.visible_to(request.user), pk=pk)
    form = SaleForm(request.POST or None, instance=sale, user=request.user)
    formset = SaleItemFormSet(request.POST or None, instance=sale, prefix="items")
    title = "Sotuvni tahrirlash"
    if request.method != "POST":
        return _render_sale_form(request, form, formset, title)
    if not (form.is_valid() and formset.is_valid()):
        return _render_sale_form(request, form, formset, title, invalid=True)
    conflict = _return_conflict(formset)
    if conflict:
        form.add_error(None, conflict)
        return _render_sale_form(request, form, formset, title, invalid=True)

    # How much the client would be over-paid by once the new total applies. Returned
    # goods, and money already handed back for them, are both out of the comparison.
    net_paid = sale.paid_amount - sale.settled_amount
    returned = sale.returned_amount
    overpay = net_paid - (_formset_total(formset) - returned)
    choice = request.POST.get("overpay_settlement") or ""
    refunded = choice == ReturnForm.SETTLE_REFUND
    if overpay > 0:
        overpay_ctx = {
            "amount": overpay,
            "cash_on_hand": seller_cash_on_hand(request.user),
            "weight_dropped": _weight_dropped(sale, formset),
        }
        if choice not in (ReturnForm.SETTLE_ADVANCE, ReturnForm.SETTLE_REFUND):
            return _render_sale_form(
                request, form, formset, title, overpay=overpay_ctx
            )
        if refunded and overpay > overpay_ctx["cash_on_hand"]:
            form.add_error(
                None,
                f"Naqd qaytarish uchun kassada pul yetarli emas: kerak "
                f"{overpay:,.0f} so'm, kassada "
                f"{overpay_ctx['cash_on_hand']:,.0f} so'm. Avans variantini tanlang "
                f"yoki avval kassaga pul kiriting.",
            )
            return _render_sale_form(
                request, form, formset, title, invalid=True, overpay=overpay_ctx
            )

    sale = form.save()
    formset.save()
    _mark_fulfilment(sale, [], only_unset=True)
    if overpay > 0:
        _settle_overpay(sale, overpay, refunded, request.user)
    summary = (
        f"Mijoz {sale.client.name}, {sale.items.count()} ta mahsulot "
        f"— {sale.total_price:,.0f} so'm"
    )
    if overpay > 0:
        summary += (
            f"; ortiqcha {overpay:,.0f} so'm "
            f"({'naqd qaytarildi' if refunded else 'avansga'})"
        )
    AuditLog.record(request.user, AuditLog.Action.UPDATE, "Sotuv", sale.pk, summary)
    messages.success(request, _sale_edit_message(overpay, refunded))
    return form_reload(request, reverse("sale_list"))


def sale_mark_paid(request, pk):
    """One-click: record a full cash payment so the sale is settled."""
    sale = get_object_or_404(Sale.objects.visible_to(request.user), pk=pk)
    if request.method == "POST":
        remaining = sale.debt_remaining
        if remaining > 0:
            Payment.objects.create(
                sale=sale, amount=remaining, amount_original=remaining,
                method=Payment.Method.CASH,
                kind=Payment.Kind.SALE, date=timezone.localdate(), created_by=request.user,
            )
            AuditLog.record(
                request.user, AuditLog.Action.PAYMENT, "To'lov", sale.pk,
                f"Mijoz {sale.client.name} to'liq to'ladi (Naqd) — {remaining:,.0f} so'm",
            )
            messages.success(request, "Sotuv to'langan deb belgilandi.")
        return form_reload(request, reverse("sale_list"))
    return render_confirm(
        request,
        "To'langan deb belgilash",
        f"“{sale.client.name}” sotuvining qoldig'i "
        f"({sale.debt_remaining:,.0f} so'm) naqd to'langan deb belgilanadimi?",
        "Ha, to'landi",
    )


def _render_debt_pay(request, sale, form, invalid=False):
    context = {
        "form": form,
        "sale": sale,
        "remaining": sale.debt_remaining,
        "title": f"To'lov: {sale.client.name}",
    }
    if is_ajax(request):
        return render(request, "crm/_debt_pay_modal.html", context, status=422 if invalid else 200)
    return render(request, "crm/_debt_pay_page.html", context)


def sale_pay(request, pk):
    """Pay one receipt. Paying MORE than it owes is allowed: the receipt is settled
    and the surplus becomes the client's advance, which then covers their other open
    receipts oldest-first — anything still left stays on their balance."""
    # .with_balance() annotates `remaining`, which _distribute_debt_payment reads.
    sale = get_object_or_404(Sale.objects.visible_to(request.user).with_balance(), pk=pk)
    if sale.is_paid:
        return form_reload(request, reverse("debt_list"))
    remaining = sale.debt_remaining
    if request.method == "POST":
        form = DebtPaymentForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            _, surplus = _distribute_debt_payment(
                [sale], cd["amount"], cd["method"], cd["commission_percent"], cd["note"],
                request.user,
                currency=cd["currency"],
                exchange_rate=cd["exchange_rate"],
                on_date=cd["date"],
                client=sale.client,
                payer=cd["commission_payer"],
            )
            applied = Decimal("0")
            if surplus > 0:
                # The overpayment is the client's credit now — spend it on whatever
                # else they still owe before letting it sit on their balance.
                applied = _apply_advance_to_open_sales(
                    sale.client, request.user, on_date=cd["date"]
                )
            AuditLog.record(
                request.user, AuditLog.Action.PAYMENT, "To'lov", sale.pk,
                f"Mijoz {sale.client.name} to'lovi "
                f"({_method_label(cd['method'])}){_usd_note(cd)} "
                f"— {cd['amount']:,.0f} so'm",
            )
            if sale.debt_remaining <= 0:
                msg = "Qarz to'liq to'landi."
            else:
                msg = f"To'lov qabul qilindi. Qoldiq: {sale.debt_remaining:,.0f} so'm."
            if surplus > 0:
                msg += f" Ortiqcha {surplus:,.0f} so'm avansga o'tdi."
                if applied > 0:
                    msg += f" Shundan {applied:,.0f} so'm boshqa ochiq cheklarga taqsimlandi."
            messages.success(request, msg)
            return form_reload(request, reverse("debt_list"))
        return _render_debt_pay(request, sale, form, invalid=True)
    form = DebtPaymentForm(
        initial={
            "amount": _clean_amount(remaining),
            "method": Payment.Method.CASH,
            "date": timezone.localdate(),
        }
    )
    return _render_debt_pay(request, sale, form)


def _render_return_form(request, sale, form, invalid=False, title=None):
    open_debt = max(Decimal("0"), sale.debt_remaining)
    net_paid = sale.paid_amount - sale.settled_amount
    context = {
        "form": form,
        "sale": sale,
        "title": title or f"Qaytarish: {sale.client.name}",
        # The seller's first question is "has this client paid yet?", because that is
        # what decides whether goods coming back cost the till anything.
        "open_debt": open_debt,
        "net_paid": net_paid,
        "is_unpaid": open_debt > 0 and net_paid <= 0,
        "is_partly_paid": open_debt > 0 and net_paid > 0,
        "is_settled": open_debt <= 0,
        # False when no return, however large, could exceed the debt — the settlement
        # choice is then dropped from the form (see ReturnForm.can_overpay).
        "can_overpay": ReturnForm.can_overpay(sale),
        "cash_on_hand": seller_cash_on_hand(request.user),
    }
    if is_ajax(request):
        return render(request, "crm/_return_modal.html", context, status=422 if invalid else 200)
    return render(request, "crm/_return_page.html", context)


@transaction.atomic
def sale_return(request, pk):
    """Take goods back on a sale and settle the money in one step.

    A return cancels the sale's open debt first. Anything beyond that is value the
    client had already paid for, so it is handed back — either parked as advance
    credit (which then flows onto their other open receipts) or paid out in cash.
    Without that settlement the receipt would sit at a permanent negative balance and
    the money owed to the client would be invisible."""
    sale = get_object_or_404(
        Sale.objects.visible_to(request.user).prefetch_related("items__product", "returns"),
        pk=pk,
    )
    if request.method == "POST":
        form = ReturnForm(request.POST, sale=sale, user=request.user)
        if form.is_valid():
            ret = form.save(commit=False)
            ret.created_by = request.user
            ret.save()
            to_debt, excess, refunded = _settle_return(ret, form, request.user)
            AuditLog.record(
                request.user, AuditLog.Action.RETURN, "Qaytarish", sale.pk,
                f"Mijoz {sale.client.name} qaytardi ({ret.product.name}) — "
                f"{ret.amount:,.0f} so'm; qarzdan {to_debt:,.0f}, "
                f"ortiqcha {excess:,.0f} "
                f"({'naqd berildi' if refunded else 'avansga'})",
            )
            messages.success(request, _return_message(ret.amount, to_debt, excess, refunded))
            return form_reload(request, reverse("sale_detail", args=[sale.pk]))
        return _render_return_form(request, sale, form, invalid=True)
    form = ReturnForm(sale=sale, user=request.user, initial={"restock": True})
    return _render_return_form(request, sale, form)


def _settle_return(ret, form, user):
    """Post the money side of a just-saved return and link it back to the return.

    The return's value cancels open debt first; any excess is money the client had
    already paid and is owed back — parked as advance credit (default) or handed out
    as cash. Kept in one place so `sale_return` and `return_edit` settle identically.
    Returns (to_debt, excess, refunded)."""
    excess = form.excess
    to_debt = form.credited_to_debt
    refunded = form.cleaned_data.get("settlement") == ReturnForm.SETTLE_REFUND
    if excess > 0:
        settlement = Payment.objects.create(
            sale=ret.sale,
            client=ret.sale.client,
            date=ret.date,
            amount=excess,
            method=Payment.Method.CASH,
            kind=(
                Payment.Kind.REFUND_OUT if refunded
                else Payment.Kind.RETURN_CREDIT
            ),
            note=f"Qaytarish: {ret.product.name}",
            created_by=user,
        )
        # Link the payment back to the return so it can be voided/edited as one unit.
        ret.settlement = settlement
        ret.save(update_fields=["settlement"])
        if not refunded:
            # Spend the fresh credit on whatever else the client still owes.
            _apply_advance_to_open_sales(ret.sale.client, user)
    return to_debt, excess, refunded


def _reverse_return(ret):
    """Roll back a return's money side and delete the return itself. The sale's debt,
    the warehouse figures and every till total re-derive to their pre-return state; a
    spent advance credit is peeled back so the pool can't go negative. Shared by
    `return_delete` (final) and `return_edit` (before re-applying the new values)."""
    settlement = ret.settlement
    is_credit = settlement is not None and settlement.kind == Payment.Kind.RETURN_CREDIT
    client = ret.sale.client
    seller = settlement.created_by if settlement else None
    if settlement is not None:
        settlement.delete()
    ret.delete()
    if is_credit and seller is not None:
        _reconcile_client_advance(client, seller)


def return_edit(request, pk):
    """Correct a mistaken return — change the line, quantity, restock flag or how the
    excess was settled. The old return is rolled back in full and the new values are
    applied as a fresh return, so debt, till and advance stay perfectly in sync (a
    return can't be safely edited in place — its settlement is derived from it). A
    seller may edit only their own returns; admins/managers any."""
    qs = Return.objects.select_related(
        "sale", "sale__client", "product", "sale_item", "settlement"
    )
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    ret = get_object_or_404(qs, pk=pk)
    sale = ret.sale
    acceptor, orig_date = ret.created_by, ret.date
    title = "Qaytarishni tahrirlash"
    if request.method == "POST":
        with transaction.atomic():
            _reverse_return(ret)
            # Validate against the restored state, so the quantity cap and the debt
            # split both see the sale as if this return had never happened.
            sale.refresh_from_db()
            form = ReturnForm(request.POST, sale=sale, user=request.user)
            if form.is_valid():
                new = form.save(commit=False)
                new.created_by = acceptor
                new.date = orig_date
                new.save()
                to_debt, excess, refunded = _settle_return(new, form, request.user)
                AuditLog.record(
                    request.user, AuditLog.Action.UPDATE, "Qaytarish", sale.pk,
                    f"Mijoz {sale.client.name} qaytarishi o'zgartirildi "
                    f"({new.product.name}) — {new.amount:,.0f} so'm",
                )
                messages.success(request, "Qaytarish yangilandi.")
                return form_reload(request, reverse("sale_detail", args=[sale.pk]))
            # Invalid: undo the tentative reversal, leaving the return untouched.
            transaction.set_rollback(True)
        return _render_return_form(request, sale, form, invalid=True, title=title)
    settlement_initial = (
        ReturnForm.SETTLE_REFUND
        if ret.settlement and ret.settlement.kind == Payment.Kind.REFUND_OUT
        else ReturnForm.SETTLE_ADVANCE
    )
    form = ReturnForm(
        sale=sale, user=request.user,
        initial={
            "sale_item": ret.sale_item_id,
            "weight": ret.weight,
            "restock": ret.restock,
            "note": ret.note,
            "settlement": settlement_initial,
        },
    )
    return _render_return_form(request, sale, form, title=title)


def return_delete(request, pk):
    """Undo a return in full. Voids the settlement it generated (the cash refund or
    the advance credit) and removes the return itself, so the sale's open debt, the
    warehouse figures and every till total re-derive to exactly their pre-return state.
    A seller may undo only their own returns; admins/managers any."""
    qs = Return.objects.select_related("sale", "sale__client", "product", "settlement")
    if not request.user.can_see_all_records:
        qs = qs.filter(created_by=request.user)
    ret = get_object_or_404(qs, pk=pk)
    settlement = ret.settlement
    is_refund = settlement is not None and settlement.kind == Payment.Kind.REFUND_OUT
    is_credit = settlement is not None and settlement.kind == Payment.Kind.RETURN_CREDIT
    if request.method == "POST":
        sale_pk = ret.sale_id
        client = ret.sale.client
        summary = (
            f"Mijoz {client.name} qaytarishi bekor qilindi "
            f"({ret.product.name}) — {ret.amount:,.0f} so'm"
        )
        with transaction.atomic():
            _reverse_return(ret)
        AuditLog.record(request.user, AuditLog.Action.VOID, "Qaytarish", sale_pk, summary)
        messages.success(request, "Qaytarish bekor qilindi — qarz va kassa qayta hisoblandi.")
        return form_reload(request, reverse("sale_detail", args=[sale_pk]))
    if is_refund:
        extra = " Naqd qaytarilgan pul kassaga qaytadi."
    elif is_credit:
        extra = (
            " Mijoz avansiga o'tgan summa bekor qilinadi — agar u boshqa "
            "sotuvlarga ishlatilgan bo'lsa, o'sha sotuvlar qayta qarzga aylanadi."
        )
    else:
        extra = ""
    return render_confirm(
        request,
        "Qaytarishni bekor qilish",
        f"“{ret.product.name}” — {ret.amount:,.0f} so'm qaytarish bekor qilinadi "
        f"va sotuv qarzi qayta tiklanadi.{extra} Davom etasizmi?",
        "Ha, bekor qilish",
        confirm_class="btn-danger",
    )


def _return_message(total, to_debt, excess, refunded):
    """Spell out where the returned value went — sellers need to see that the money
    side was handled, not just that goods came back."""
    parts = [f"Qaytarish qabul qilindi: {total:,.0f} so'm."]
    if to_debt > 0:
        parts.append(f"Qarzdan {to_debt:,.0f} so'm yopildi.")
    if excess > 0:
        parts.append(
            f"Ortiqcha {excess:,.0f} so'm "
            + ("kassadan naqd qaytarildi." if refunded else "mijoz avansiga o'tdi.")
        )
    return " ".join(parts)


def sale_delete(request, pk):
    """Delete a sale outright. Any payments and returns booked against it are reversed
    with it — they cascade — and the client's advance is reconciled so freed credit
    settles onto other open receipts (or an orphaned credit is peeled back). The till,
    debt and profit all re-derive. Because this removes money records, the confirm
    dialog spells out exactly what will go, mirroring how payment/advance voids work."""
    sale = get_object_or_404(
        Sale.objects.visible_to(request.user).select_related("client"), pk=pk
    )
    client, seller = sale.client, sale.sales_rep
    paid = sale.paid_amount
    return_count = sale.returns.count()
    if request.method == "POST":
        summary = f"Mijoz {sale.client.name} sotuvi — {sale.total_price:,.0f} so'm"
        sale_pk = sale.pk
        with transaction.atomic():
            sale.delete()  # items, payments and returns cascade with it
            # Freed or now-orphaned advance allocations settle back into balance.
            _reconcile_client_advance(client, seller)
        AuditLog.record(request.user, AuditLog.Action.DELETE, "Sotuv", sale_pk, summary)
        messages.success(request, "Sotuv o'chirildi.")
        return form_reload(request, reverse("sale_list"))
    extra = []
    if paid > 0:
        extra.append(f"{paid:,.0f} so'm to'lov")
    if return_count:
        extra.append(f"{return_count} ta qaytarish")
    warn = (
        f" Unga bog'liq {' va '.join(extra)} ham o'chiriladi, kassa va qarz qayta hisoblanadi."
        if extra else ""
    )
    return render_confirm(
        request,
        "Sotuvni o'chirish",
        f"Bu sotuv butunlay o'chiriladi.{warn} Davom etasizmi?",
        "Ha, o'chirish",
        confirm_class="btn-danger",
    )


def _render_production_adjust(request, form, invalid=False):
    """The correction form, as a modal or a full page. Not `form_response`: its
    full-page branch renders the generic form template, which would drop the steering
    warnings — the part that keeps this form from being misused."""
    context = {
        "form": form,
        "title": "Ishlab chiqarish qarzini tuzatish",
        # {reason: message} for the warnings wired up in base.html.
        "steered_json": {str(k): v for k, v in ProductionAdjustForm.STEERED.items()},
    }
    if is_ajax(request):
        return render(
            request, "crm/_production_adjust_modal.html", context,
            status=422 if invalid else 200,
        )
    return render(request, "crm/_production_adjust_page.html", context)


def _adjust_summary(adj):
    sign = "kamaytirildi" if adj.lowers_debt else "oshirildi"
    note = f" — {adj.note}" if adj.note else ""
    return (
        f"Sotuvchi {adj.seller} ishlab chiqarish qarzi {sign}: "
        f"{adj.abs_amount:,.0f} so'm ({adj.get_reason_display()}){note}"
    )


@role_required(User.Role.ADMIN)
def production_adjust_create(request):
    """Correct what a seller owes production, without any money moving.

    Admin only, and deliberately so: this is the one figure in the system with no
    document behind it, and the seller it belongs to must never be able to shrink
    their own liability. Reachable from the per-seller control table on the kassa
    page, which preselects the seller via ?seller=.

    The form warns when the chosen reason names a case whose real fix lives elsewhere
    (a forgotten handover, an unrecorded sale) but does not refuse it — see
    `ProductionAdjustForm`."""
    initial = {}
    seller_pk = request.GET.get("seller", "")
    if request.method == "GET" and seller_pk.isdigit():
        initial["seller"] = seller_pk
    form = ProductionAdjustForm(request.POST or None, user=request.user, initial=initial)
    if request.method == "POST":
        if form.is_valid():
            adj = form.save(commit=False)
            adj.created_by = request.user
            adj.save()
            AuditLog.record(
                request.user, AuditLog.Action.UPDATE, "Tuzatish", adj.pk,
                _adjust_summary(adj),
            )
            messages.success(
                request,
                f"Qarz tuzatildi. {adj.seller} ishlab chiqarishga qarzi endi "
                f"{seller_production_debt(adj.seller):,.0f} so'm.",
            )
            # Reachable from both the kassa control table and the user list, so the
            # modal reloads wherever it was opened from rather than jumping to one.
            return form_reload(request, reverse("kassa"))
        return _render_production_adjust(request, form, invalid=True)
    return _render_production_adjust(request, form)


@role_required(User.Role.ADMIN)
def production_adjust_delete(request, pk):
    """Undo a correction. There is no edit: a wrong figure is voided and re-entered,
    so the audit trail keeps both the mistake and its withdrawal instead of quietly
    rewriting one row."""
    adj = get_object_or_404(ProductionAdjustment.objects.select_related("seller"), pk=pk)
    if request.method == "POST":
        summary = _adjust_summary(adj)
        seller, adj_pk = adj.seller, adj.pk
        adj.delete()
        AuditLog.record(
            request.user, AuditLog.Action.VOID, "Tuzatish", adj_pk,
            f"Bekor qilindi: {summary}",
        )
        messages.success(
            request,
            f"Tuzatish bekor qilindi. {seller} ishlab chiqarishga qarzi endi "
            f"{seller_production_debt(seller):,.0f} so'm.",
        )
        return form_reload(request, reverse("kassa"))
    sign = "kamaytirgan" if adj.lowers_debt else "oshirgan"
    return render_confirm(
        request,
        "Tuzatishni bekor qilish",
        f"{adj.seller} qarzini {adj.abs_amount:,.0f} so'mga {sign} tuzatish "
        f"o'chiriladi va qarz avvalgi holiga qaytadi. Davom etasizmi?",
        "Ha, bekor qilish",
        confirm_class="btn-danger",
    )
