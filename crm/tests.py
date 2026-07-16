from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User

from .forms import SaleForm
from .models import (
    AuditLog,
    Client,
    Expense,
    Payment,
    Product,
    ProductionReceipt,
    ProductionReceiptItem,
    ProductionRemittance,
    Return,
    Sale,
    SaleItem,
    StockEntry,
)
from .views import (
    XLSX_CONTENT_TYPE,
    _kassa_summary,
    _per_employee_kassa,
    _realized_profit_by_seller,
)


def read_xlsx(response):
    """Return an .xlsx export response as a list of row tuples (header first)."""
    return list(load_workbook(BytesIO(response.content)).active.iter_rows(values_only=True))


def make_sale(client, rep, product, weight="10", price="24000", **kwargs):
    """Create a sale (header) with a single line item.

    Every sale is a receivable; pass is_debt=True to leave it unpaid, otherwise
    a full cash payment is recorded so the sale reads as paid (the old default).
    Header kwargs (date, debt_deadline) go to the Sale; cost_price/dimension to
    the item."""
    is_debt = kwargs.pop("is_debt", False)
    deadline = kwargs.pop("debt_deadline", None)
    cost_price = kwargs.pop("cost_price", product.cost_price)
    dimension = kwargs.pop("dimension", "kg")
    header = {"client": client, "sales_rep": rep}
    if "date" in kwargs:
        header["date"] = kwargs.pop("date")
    header["debt_deadline"] = deadline or (timezone.localdate() + timedelta(days=7))
    sale = Sale.objects.create(**header)
    SaleItem.objects.create(
        sale=sale,
        product=product,
        dimension=dimension,
        weight=Decimal(weight),
        price=Decimal(price),
        cost_price=Decimal(cost_price),
        fulfilled_at=sale.date,
    )
    if not is_debt:
        Payment.objects.create(
            sale=sale, amount=sale.total_price, method=Payment.Method.CASH,
            kind=Payment.Kind.SALE, date=sale.date, created_by=rep,
        )
    return sale


def sale_post(client_pk, items, **header):
    """Build POST data for the sale create/edit form (header + item formset)."""
    data = {
        "date": timezone.localdate().isoformat(),
        "client": client_pk,
        "items-TOTAL_FORMS": str(len(items)),
        "items-INITIAL_FORMS": "0",
        "items-MIN_NUM_FORMS": "1",
        "items-MAX_NUM_FORMS": "1000",
    }
    for i, item in enumerate(items):
        for key, value in item.items():
            data[f"items-{i}-{key}"] = value
    data.update(header)
    return data


def one_item(product, weight="5", price="24000", dimension="kg", cost_price=""):
    return {
        "product": product.pk,
        "dimension": dimension,
        "weight": weight,
        "price": price,
        "cost_price": cost_price,
    }


def give_ombor(seller, product, kg="100000", **kwargs):
    """Log an opening production receipt so `seller` holds `kg` of `product`.

    Sales go through the ombor oversell block, so any test that creates a sale via
    the view needs the seller to have received the goods first."""
    receipt = ProductionReceipt.objects.create(seller=seller, created_by=seller, **kwargs)
    ProductionReceiptItem.objects.create(
        receipt=receipt, product=product, quantity_kg=Decimal(kg)
    )
    return receipt


class BaseSetup(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user("t_admin", password="x", role=User.Role.ADMIN)
        cls.manager = User.objects.create_user("t_manager", password="x", role=User.Role.MANAGER)
        cls.sales1 = User.objects.create_user("t_sales1", password="x", role=User.Role.SALES)
        cls.sales2 = User.objects.create_user("t_sales2", password="x", role=User.Role.SALES)
        cls.client1 = Client.objects.create(name="Mijoz A", owner=cls.sales1)
        cls.client2 = Client.objects.create(name="Mijoz B", owner=cls.sales2)
        cls.product = Product.objects.create(
            name="Polietilen paket", sku="PKT-1",
            cost_price=Decimal("18000"), price=Decimal("24000"),
        )
        cls.sale1 = make_sale(cls.client1, cls.sales1, cls.product)
        cls.sale2 = make_sale(cls.client2, cls.sales2, cls.product)
        # Seed each user's ombor so view-based sale tests clear the oversell block.
        for user in (cls.admin, cls.manager, cls.sales1, cls.sales2):
            give_ombor(user, cls.product)


class SaleMathTests(BaseSetup):
    def test_totals_and_profit(self):
        # 10 kg × 24 000 = 240 000; tannarx 10 × 18 000 = 180 000; foyda 60 000
        self.assertEqual(self.sale1.total_price, Decimal("240000"))
        self.assertEqual(self.sale1.total_cost, Decimal("180000"))
        self.assertEqual(self.sale1.profit, Decimal("60000"))

    def test_with_totals_annotation(self):
        sale = Sale.objects.with_totals().get(pk=self.sale1.pk)
        self.assertEqual(sale.total, Decimal("240000"))
        self.assertEqual(sale.profit_total, Decimal("60000"))

    def test_cost_price_for_grams(self):
        self.assertEqual(self.product.cost_price_for(Sale.Dimension.G), Decimal("18"))


class DebtTests(BaseSetup):
    def _sale_data(self, **overrides):
        return sale_post(self.client1.pk, [one_item(self.product)], **overrides)

    def test_new_sale_is_outstanding_with_no_payment(self):
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("sale_create"), self._sale_data())
        self.assertEqual(response.status_code, 302)
        sale = Sale.objects.latest("created_at")
        self.assertTrue(sale.is_outstanding)
        self.assertEqual(sale.payments.count(), 0)

    def test_blank_days_defaults_to_seven_days(self):
        self.client.force_login(self.sales1)
        self.client.post(reverse("sale_create"), self._sale_data(debt_days=""))
        sale = Sale.objects.latest("created_at")
        self.assertEqual(sale.debt_deadline, timezone.localdate() + timedelta(days=7))

    def test_debt_days_sets_deadline(self):
        self.client.force_login(self.sales1)
        self.client.post(reverse("sale_create"), self._sale_data(debt_days="30"))
        sale = Sale.objects.latest("created_at")
        self.assertEqual(sale.debt_deadline, timezone.localdate() + timedelta(days=30))

    def test_overdue_flag(self):
        overdue = make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=timezone.localdate() - timedelta(days=1),
        )
        not_due = make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=timezone.localdate() + timedelta(days=1),
        )
        self.assertTrue(overdue.is_overdue)
        self.assertFalse(not_due.is_overdue)


class CostFallbackTests(BaseSetup):
    def test_empty_cost_price_uses_product_cost(self):
        self.client.force_login(self.sales1)
        data = sale_post(
            self.client1.pk,
            [one_item(self.product, weight="500", price="24", dimension="g")],
        )
        response = self.client.post(reverse("sale_create"), data)
        self.assertEqual(response.status_code, 302)
        sale = Sale.objects.latest("created_at")
        item = sale.items.get()
        # per-gram tannarx = 18000 / 1000
        self.assertEqual(item.cost_price, Decimal("18"))
        self.assertEqual(sale.sales_rep, self.sales1)


class MultiItemSaleTests(BaseSetup):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.product2 = Product.objects.create(
            name="Mayka paket", sku="MYK-1",
            cost_price=Decimal("17000"), price=Decimal("23000"),
        )
        for user in (cls.sales1, cls.sales2):
            give_ombor(user, cls.product2)

    def test_creates_one_sale_with_multiple_items(self):
        self.client.force_login(self.sales1)
        data = sale_post(
            self.client1.pk,
            [
                one_item(self.product, weight="10", price="24000"),   # 240 000
                one_item(self.product2, weight="5", price="23000"),   # 115 000
            ],
        )
        response = self.client.post(reverse("sale_create"), data)
        self.assertEqual(response.status_code, 302)
        sale = Sale.objects.latest("created_at")
        self.assertEqual(sale.items.count(), 2)
        self.assertEqual(sale.total_price, Decimal("355000"))
        # a new receipt is an unpaid receivable — no payment recorded yet
        self.assertEqual(sale.payments.count(), 0)
        self.assertTrue(sale.is_outstanding)

    def test_requires_at_least_one_item(self):
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [])
        response = self.client.post(reverse("sale_create"), data)
        self.assertEqual(response.status_code, 200)  # re-rendered, not saved
        self.assertFalse(Sale.objects.filter(client=self.client1).exclude(pk=self.sale1.pk).exists())


class RoleScopingTests(BaseSetup):
    def test_sales_sees_only_own_sales(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_list"))
        self.assertEqual(list(response.context["page"].object_list), [self.sale1])

    def test_manager_sees_all_sales(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("sale_list"))
        self.assertEqual(len(response.context["page"].object_list), 2)

    def test_sales_cannot_edit_others_sale(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_edit", args=[self.sale2.pk]))
        self.assertEqual(response.status_code, 404)

    def test_sales_sees_only_own_clients(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("client_list"))
        self.assertEqual(list(response.context["page"].object_list), [self.client1])

    def test_sales_can_create_product(self):
        # Products/warehouse are a shared company section: every role, sellers
        # included, may manage them. Only per-seller data (clients, sales, own
        # money actions) is owner-scoped.
        self.client.force_login(self.sales1)
        self.assertEqual(self.client.get(reverse("product_create")).status_code, 200)

    def test_manager_can_create_product(self):
        self.client.force_login(self.manager)
        self.assertEqual(self.client.get(reverse("product_create")).status_code, 200)

    def test_only_admin_can_manage_users(self):
        self.client.force_login(self.manager)
        self.assertEqual(self.client.get(reverse("user_list")).status_code, 403)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse("user_list")).status_code, 200)

    def test_sales_cannot_pick_others_client(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_create"))
        self.assertNotIn(self.client2, response.context["form"].fields["client"].queryset)


class AuthTests(BaseSetup):
    def test_anonymous_redirected_to_login(self):
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response.url)

    def test_dashboard_renders_for_each_role(self):
        for user in (self.admin, self.manager, self.sales1):
            self.client.force_login(user)
            self.assertEqual(self.client.get(reverse("dashboard")).status_code, 200)

    def test_dashboard_provides_chart_data(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("dashboard"))
        ctx = response.context
        self.assertEqual(len(ctx["monthly"]["rows"]), 6)      # 6-month trend
        self.assertIn("revenue", ctx["monthly"]["rows"][0])
        self.assertEqual(len(ctx["donut"]["segments"]), 3)    # cash / card / transfer
        self.assertIsNotNone(ctx["donut"]["grand_short"])
        self.assertTrue(all("pct" in c for c in ctx["top_clients"]))
        # Numbers inside style/SVG attributes must stay unlocalised — a comma
        # decimal separator silently breaks CSS lengths and SVG dash arrays.
        body = response.content.decode()
        self.assertNotRegex(body, r"(?:height|width):\s*\d+,\d")
        self.assertNotRegex(body, r'stroke-dash(?:array|offset)="[^"]*,')

    def test_dashboard_flags_overdue_debt(self):
        make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=timezone.localdate() - timedelta(days=3),
        )
        self.client.force_login(self.manager)
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.context["overdue_count"], 1)
        self.assertEqual(response.context["overdue_total"], Decimal("240000"))
        self.assertContains(response, "overdue-banner")


class DayViewTests(BaseSetup):
    def test_defaults_to_today(self):
        old = make_sale(
            self.client1, self.sales1, self.product,
            date=timezone.localdate() - timedelta(days=3),
        )
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_list"))
        sales = list(response.context["page"].object_list)
        self.assertIn(self.sale1, sales)   # today
        self.assertNotIn(old, sales)       # 3 days ago
        self.assertTrue(response.context["is_today"])

    def test_shows_selected_day(self):
        target = timezone.localdate() - timedelta(days=3)
        old = make_sale(self.client1, self.sales1, self.product, date=target)
        self.client.force_login(self.sales1)
        response = self.client.get(
            reverse("sale_list"), {"dan": target.isoformat(), "gacha": target.isoformat()}
        )
        sales = list(response.context["page"].object_list)
        self.assertIn(old, sales)
        self.assertNotIn(self.sale1, sales)  # today's sale not on the target day
        self.assertFalse(response.context["is_today"])
        self.assertTrue(response.context["is_single_day"])

    def test_date_range_spans_multiple_days(self):
        today = timezone.localdate()
        d5 = make_sale(self.client1, self.sales1, self.product, date=today - timedelta(days=5))
        d2 = make_sale(self.client1, self.sales1, self.product, date=today - timedelta(days=2))
        old = make_sale(self.client1, self.sales1, self.product, date=today - timedelta(days=20))
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_list"), {
            "dan": (today - timedelta(days=7)).isoformat(),
            "gacha": today.isoformat(),
        })
        sales = list(response.context["page"].object_list)
        self.assertIn(d5, sales)
        self.assertIn(d2, sales)
        self.assertIn(self.sale1, sales)     # today, within range
        self.assertNotIn(old, sales)         # 20 days ago, outside range
        self.assertFalse(response.context["is_single_day"])

    def test_filter_ignores_date_range(self):
        # An out-of-window sale must surface once a content filter is applied
        old = make_sale(
            self.client1, self.sales1, self.product,
            date=timezone.localdate() - timedelta(days=30),
        )
        self.client.force_login(self.sales1)
        default = self.client.get(reverse("sale_list"))
        self.assertNotIn(old, list(default.context["page"].object_list))
        self.assertFalse(default.context["has_filters"])

        filtered = self.client.get(reverse("sale_list"), {"client": self.client1.pk})
        self.assertIn(old, list(filtered.context["page"].object_list))  # date window bypassed
        self.assertTrue(filtered.context["has_filters"])
        chips = filtered.context["active_filters"]
        self.assertEqual(len(chips), 1)
        self.assertEqual(chips[0]["label"], "Mijoz")
        self.assertEqual(chips[0]["value"], self.client1.name)

    def test_toolbar_branches_on_has_filters_not_chip_presence(self):
        # A filter id outside the user's visibility: has_filters True but no chip.
        self.client.force_login(self.sales2)
        resp = self.client.get(reverse("sale_list"), {"client": self.client1.pk})
        self.assertTrue(resp.context["has_filters"])
        self.assertEqual(resp.context["active_filters"], [])
        # Toolbar shows the clear-filter control and hides the date-range picker,
        # matching the pre-refactor behavior. (The bare class name also appears
        # in base.html's unconditional JS behavior script, so match the actual
        # rendered element instead of the substring.)
        self.assertContains(resp, "chip-clear")
        self.assertNotContains(resp, 'class="daterange-trigger"')


class StockTests(BaseSetup):
    def setUp(self):
        # BaseSetup already created sale1 (10 kg) + sale2 for self.product
        StockEntry.objects.create(
            product=self.product, quantity_kg=Decimal("100"), created_by=self.admin
        )

    def test_current_stock_is_received_minus_sold(self):
        # 100 kg in; two 10 kg sales out (sale1 + sale2) => 80 kg
        self.product.refresh_from_db()
        self.assertEqual(self.product.total_received, Decimal("100"))
        self.assertEqual(self.product.total_sold, Decimal("20"))
        self.assertEqual(self.product.current_stock, Decimal("80"))

    def test_gram_sale_converts_to_kg(self):
        make_sale(self.client1, self.sales1, self.product, weight="500", dimension="g")
        # 500 g = 0.5 kg extra sold => 20.5 kg out; 100 - 20.5 = 79.5
        self.assertEqual(self.product.total_sold, Decimal("20.500"))
        self.assertEqual(self.product.current_stock, Decimal("79.500"))

    def test_with_stock_annotation_matches_property(self):
        annotated = Product.objects.with_stock().get(pk=self.product.pk)
        self.assertEqual(annotated.stock, self.product.current_stock)

    def test_manager_can_add_kirim(self):
        self.client.force_login(self.manager)
        response = self.client.post(
            reverse("stock_entry_create", args=[self.product.pk]),
            {"date": timezone.localdate().isoformat(), "quantity_kg": "250", "note": "Yangi partiya"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.product.total_received, Decimal("350"))
        entry = StockEntry.objects.latest("created_at")
        self.assertEqual(entry.created_by, self.manager)

    def test_sales_can_add_kirim(self):
        # The warehouse is shared, so a seller may record a stock kirim.
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("stock_entry_create", args=[self.product.pk]))
        self.assertEqual(response.status_code, 200)

    def test_sale_beyond_stock_warns_but_saves(self):
        self.client.force_login(self.sales1)
        data = sale_post(
            self.client1.pk,
            [one_item(self.product, weight="500")],  # far beyond the 80 kg on hand
        )
        response = self.client.post(reverse("sale_create"), data, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(SaleItem.objects.filter(weight=Decimal("500")).exists())
        msgs = [m.message for m in response.context["messages"]]
        self.assertTrue(any("yetarli emas" in m for m in msgs))

    def test_low_stock_flag(self):
        self.product.low_stock_threshold = Decimal("90")
        self.product.save()
        # current stock 80 <= threshold 90 => low
        self.assertTrue(self.product.is_low_stock)

    def test_product_detail_accessible_to_sales(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("product_detail", args=[self.product.pk]))
        self.assertEqual(response.status_code, 200)

    def test_adjust_sets_exact_quantity(self):
        # current stock is 80 (100 in - 20 sold)
        self.client.force_login(self.manager)
        before = StockEntry.objects.count()
        response = self.client.post(
            reverse("stock_adjust", args=[self.product.pk]), {"quantity": "200", "note": ""}
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.product.current_stock, Decimal("200.000"))
        # exactly one movement logged, holding the +120 delta
        self.assertEqual(StockEntry.objects.count(), before + 1)
        self.assertEqual(StockEntry.objects.latest("created_at").quantity_kg, Decimal("120.000"))

    def test_adjust_can_decrease_below_current(self):
        self.client.force_login(self.manager)
        self.client.post(
            reverse("stock_adjust", args=[self.product.pk]), {"quantity": "50"}
        )
        self.assertEqual(self.product.current_stock, Decimal("50.000"))
        self.assertEqual(StockEntry.objects.latest("created_at").quantity_kg, Decimal("-30.000"))

    def test_sales_can_adjust(self):
        # Warehouse is shared — a seller may correct the stock quantity too.
        self.client.force_login(self.sales1)
        self.assertEqual(
            self.client.get(reverse("stock_adjust", args=[self.product.pk])).status_code, 200
        )


class SellerOmborTests(BaseSetup):
    """Each seller has their own ombor = received from production − sold + restocked."""

    def setUp(self):
        # A product this class fully controls — BaseSetup only seeds self.product.
        self.p = Product.objects.create(
            name="Ombor test", sku="OMB-1",
            cost_price=Decimal("10000"), price=Decimal("20000"),
        )

    def _receive(self, seller, kg, **kwargs):
        receipt = ProductionReceipt.objects.create(seller=seller, created_by=seller, **kwargs)
        ProductionReceiptItem.objects.create(
            receipt=receipt, product=self.p, quantity_kg=Decimal(kg)
        )
        return receipt

    def test_on_hand_is_received_minus_sold(self):
        self._receive(self.sales1, "100")
        make_sale(self.client1, self.sales1, self.p, weight="30", price="20000")
        p = Product.objects.with_stock(seller=self.sales1).get(pk=self.p.pk)
        self.assertEqual(p.stock, Decimal("70"))

    def test_restocked_return_adds_back_to_ombor(self):
        self._receive(self.sales1, "100")
        sale = make_sale(self.client1, self.sales1, self.p, weight="30", price="20000")
        Return.objects.create(
            sale=sale, product=self.p, dimension="kg",
            weight=Decimal("4"), price=Decimal("20000"), restock=True,
            created_by=self.sales1,
        )
        p = Product.objects.with_stock(seller=self.sales1).get(pk=self.p.pk)
        self.assertEqual(p.stock, Decimal("74"))  # 100 − 30 + 4

    def test_pre_cutover_receipt_is_excluded(self):
        self._receive(self.sales1, "100", date=date(2019, 1, 1))
        p = Product.objects.with_stock(seller=self.sales1).get(pk=self.p.pk)
        self.assertEqual(p.stock, Decimal("0"))  # pre-cutover receipt ignored, no sales

    def test_ombor_is_isolated_per_seller(self):
        self._receive(self.sales2, "100")
        p1 = Product.objects.with_stock(seller=self.sales1).get(pk=self.p.pk)
        self.assertEqual(p1.stock, Decimal("0"))
        p2 = Product.objects.with_stock(seller=self.sales2).get(pk=self.p.pk)
        self.assertEqual(p2.stock, Decimal("100"))

    def test_product_list_scoped_to_seller_ombor(self):
        # BaseSetup seeds sales1 with 100000 of self.product; sale1 sold 10 => 99990
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("product_list"))
        products = {p.pk: p for p in response.context["page"].object_list}
        self.assertEqual(products[self.product.pk].stock, Decimal("99990"))

    def test_sale_blocked_when_ombor_insufficient(self):
        # sales1 has no receipt for self.p -> ombor 0; selling 3 is blocked
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [one_item(self.p, weight="3")])
        response = self.client.post(reverse("sale_create"), data)
        self.assertEqual(response.status_code, 200)  # re-rendered, not redirected
        self.assertContains(response, "yetarli emas")
        self.assertFalse(SaleItem.objects.filter(product=self.p).exists())

    def test_sale_allowed_within_ombor(self):
        self._receive(self.sales1, "100")
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [one_item(self.p, weight="3")])
        response = self.client.post(reverse("sale_create"), data, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(SaleItem.objects.filter(product=self.p, weight=Decimal("3")).exists())


class ReceiptCrudTests(BaseSetup):
    """Logging goods received from production into a seller's ombor."""

    def _post(self, product, qty="50", **header):
        data = {
            "date": timezone.localdate().isoformat(),
            "note": "",
            "items-TOTAL_FORMS": "1",
            "items-INITIAL_FORMS": "0",
            "items-MIN_NUM_FORMS": "1",
            "items-MAX_NUM_FORMS": "1000",
            "items-0-product": product.pk,
            "items-0-quantity_kg": qty,
        }
        data.update(header)
        return data

    def test_seller_logs_receipt(self):
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("receipt_create"), self._post(self.product, "50"))
        self.assertEqual(response.status_code, 302)
        receipt = ProductionReceipt.objects.latest("created_at")
        self.assertEqual(receipt.seller, self.sales1)
        self.assertEqual(receipt.created_by, self.sales1)
        self.assertEqual(receipt.items.get().quantity_kg, Decimal("50.000"))

    def test_seller_cannot_log_for_another_seller(self):
        self.client.force_login(self.sales1)
        self.client.post(reverse("receipt_create"), self._post(self.product, "50", seller=self.sales2.pk))
        self.assertEqual(ProductionReceipt.objects.latest("created_at").seller, self.sales1)

    def test_admin_logs_receipt_for_a_seller(self):
        self.client.force_login(self.admin)
        self.client.post(reverse("receipt_create"), self._post(self.product, "50", seller=self.sales2.pk))
        receipt = ProductionReceipt.objects.latest("created_at")
        self.assertEqual(receipt.seller, self.sales2)
        self.assertEqual(receipt.created_by, self.admin)

    def test_receipt_raises_seller_ombor(self):
        self.client.force_login(self.sales1)
        before = Product.objects.with_stock(seller=self.sales1).get(pk=self.product.pk).stock
        self.client.post(reverse("receipt_create"), self._post(self.product, "50"))
        after = Product.objects.with_stock(seller=self.sales1).get(pk=self.product.pk).stock
        self.assertEqual(after - before, Decimal("50"))

    def test_ombor_page_renders_for_seller(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("ombor"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ombordagi qoldiq")

    def test_ombor_page_renders_for_admin_with_seller(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("ombor") + f"?seller={self.sales1.pk}")
        self.assertEqual(response.status_code, 200)

    def test_seller_cannot_edit_another_sellers_receipt(self):
        receipt = give_ombor(self.sales2, self.product, "10")
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("receipt_edit", args=[receipt.pk]))
        self.assertEqual(response.status_code, 404)


class ZakazFulfillmentTests(BaseSetup):
    """Selling on zakaz (without stock) and binding arriving stock to orders."""

    def setUp(self):
        self.p = Product.objects.create(
            name="Zakaz test", sku="ZK-1",
            cost_price=Decimal("10000"), price=Decimal("20000"),
        )

    def test_in_stock_sale_line_is_fulfilled(self):
        give_ombor(self.sales1, self.p, "100")
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_create"), sale_post(self.client1.pk, [one_item(self.p, weight="10")])
        )
        item = SaleItem.objects.filter(product=self.p).latest("pk")
        self.assertIsNotNone(item.fulfilled_at)

    def test_oversell_prompts_before_saving(self):
        self.client.force_login(self.sales1)  # no ombor for self.p
        response = self.client.post(
            reverse("sale_create"), sale_post(self.client1.pk, [one_item(self.p, weight="10")])
        )
        self.assertEqual(response.status_code, 200)  # re-prompted, not saved
        self.assertContains(response, "zakaz")
        self.assertFalse(SaleItem.objects.filter(product=self.p).exists())

    def test_zakaz_confirm_saves_pending_line(self):
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [one_item(self.p, weight="10")], allow_zakaz="1")
        response = self.client.post(reverse("sale_create"), data)
        self.assertEqual(response.status_code, 302)  # saved
        item = SaleItem.objects.filter(product=self.p).latest("pk")
        self.assertIsNone(item.fulfilled_at)  # pending zakaz

    def test_oversell_ajax_asks_for_confirmation(self):
        # In the modal (AJAX), an oversell returns a confirm signal, not a re-render
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [one_item(self.p, weight="10")])
        response = self.client.post(
            reverse("sale_create"), data, HTTP_X_REQUESTED_WITH="XMLHttpRequest"
        )
        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.headers.get("X-Zakaz-Confirm"), "1")
        self.assertFalse(SaleItem.objects.filter(product=self.p).exists())

    def test_oversell_ajax_confirmed_saves(self):
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [one_item(self.p, weight="10")], allow_zakaz="1")
        response = self.client.post(
            reverse("sale_create"), data, HTTP_X_REQUESTED_WITH="XMLHttpRequest"
        )
        self.assertEqual(response.status_code, 204)  # modal success (X-Redirect)
        self.assertTrue(SaleItem.objects.filter(product=self.p).exists())

    def _receipt_post(self, product, qty="50"):
        return {
            "date": timezone.localdate().isoformat(), "note": "",
            "items-TOTAL_FORMS": "1", "items-INITIAL_FORMS": "0",
            "items-MIN_NUM_FORMS": "1", "items-MAX_NUM_FORMS": "1000",
            "items-0-product": product.pk, "items-0-quantity_kg": qty,
        }

    def _make_zakaz(self):
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_create"),
            sale_post(self.client1.pk, [one_item(self.p, weight="10")], allow_zakaz="1"),
        )
        return SaleItem.objects.filter(product=self.p).latest("pk")

    def test_receipt_with_pending_zakaz_redirects_to_bind(self):
        self._make_zakaz()
        response = self.client.post(reverse("receipt_create"), self._receipt_post(self.p))
        receipt = ProductionReceipt.objects.latest("created_at")
        self.assertRedirects(
            response, reverse("receipt_bind", args=[receipt.pk]), fetch_redirect_response=False
        )

    def test_receipt_without_pending_goes_to_ombor(self):
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("receipt_create"), self._receipt_post(self.product))
        self.assertRedirects(response, reverse("ombor"), fetch_redirect_response=False)

    def test_bind_marks_order_fulfilled(self):
        item = self._make_zakaz()
        self.assertIsNone(item.fulfilled_at)
        receipt = give_ombor(self.sales1, self.p, "50")
        self.client.post(reverse("receipt_bind", args=[receipt.pk]), {"bind": [item.pk]})
        item.refresh_from_db()
        self.assertIsNotNone(item.fulfilled_at)
        self.assertEqual(item.fulfilled_by_receipt, receipt)


class KassaCombinedCashTests(BaseSetup):
    """Kassadagi pul combines every method AND currency — a dollar payment counts
    at its so'm value (Payment.amount is always stored in so'm)."""

    def _usd_payment(self, seller, som="1270000", usd="100"):
        sale = make_sale(self.client1, seller, self.product, is_debt=True)
        return Payment.objects.create(
            sale=sale, amount=Decimal(som), amount_original=Decimal(usd),
            exchange_rate=Decimal("12700"), currency=Payment.Currency.USD,
            method=Payment.Method.CASH, kind=Payment.Kind.SALE,
            date=timezone.localdate(), created_by=seller,
        )

    def test_summary_cash_includes_dollars(self):
        today = timezone.localdate()
        before = _kassa_summary(today, today)["cash"]
        self._usd_payment(self.sales1)
        after = _kassa_summary(today, today)["cash"]
        self.assertEqual(after - before, Decimal("1270000"))

    def test_per_seller_cash_includes_dollars(self):
        today = timezone.localdate()
        before = _per_employee_kassa(today, today, rep=self.sales1)[0]["cash"]
        self._usd_payment(self.sales1)
        after = _per_employee_kassa(today, today, rep=self.sales1)[0]["cash"]
        self.assertEqual(after - before, Decimal("1270000"))


class ProductDeleteTests(BaseSetup):
    def test_seller_deletes_unsold_product(self):
        # Ombor is a shared section — a seller may delete a product too, as long
        # as it has no sales history tying it down.
        product = Product.objects.create(name="Vaqtinchalik", sku="TMP-1", price=Decimal("1000"))
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("product_delete", args=[product.pk]))
        self.assertRedirects(response, reverse("product_list"))
        self.assertFalse(Product.objects.filter(pk=product.pk).exists())

    def test_product_with_sales_is_protected(self):
        # The base fixture product has sales; deletion is blocked and it survives.
        self.client.force_login(self.sales1)
        self.client.post(reverse("product_delete", args=[self.product.pk]))
        self.assertTrue(Product.objects.filter(pk=self.product.pk).exists())

    def test_delete_cascades_stock_entries(self):
        # A product with only stock entries deletes cleanly (entries cascade away).
        product = Product.objects.create(name="Faqat kirim", sku="TMP-2", price=Decimal("1000"))
        StockEntry.objects.create(
            product=product, quantity_kg=Decimal("5"), created_by=self.sales1
        )
        self.client.force_login(self.sales1)
        self.client.post(reverse("product_delete", args=[product.pk]))
        self.assertFalse(Product.objects.filter(pk=product.pk).exists())
        self.assertFalse(StockEntry.objects.filter(product_id=product.pk).exists())

    def test_delete_writes_audit_log(self):
        product = Product.objects.create(name="Audit paket", sku="TMP-3", price=Decimal("1000"))
        self.client.force_login(self.sales1)
        self.client.post(reverse("product_delete", args=[product.pk]))
        self.assertTrue(
            AuditLog.objects.filter(
                action=AuditLog.Action.DELETE, target_type="Mahsulot", summary="Audit paket"
            ).exists()
        )


class ProductCreateStockTests(BaseSetup):
    def _post(self, **extra):
        data = {
            "name": "Yangi tovar", "sku": "NEW-1", "description": "",
            "cost_price": "10000", "price": "15000",
            "low_stock_threshold": "5", "is_active": "on",
        }
        data.update(extra)
        return self.client.post(reverse("product_create"), data)

    def test_initial_quantity_recorded_as_kirim(self):
        # Creating a product with a starting quantity books it as the first kirim.
        self.client.force_login(self.sales1)
        self._post(initial_quantity="100")
        product = Product.objects.get(sku="NEW-1")
        self.assertEqual(product.current_stock, Decimal("100"))
        entry = product.stock_entries.get()
        self.assertEqual(entry.quantity_kg, Decimal("100"))
        self.assertEqual(entry.created_by, self.sales1)

    def test_no_initial_quantity_leaves_stock_empty(self):
        self.client.force_login(self.sales1)
        self._post()  # no initial_quantity
        product = Product.objects.get(sku="NEW-1")
        self.assertEqual(product.stock_entries.count(), 0)
        self.assertEqual(product.current_stock, Decimal("0"))

    def test_edit_form_omits_initial_quantity(self):
        # Existing stock is changed via Kirim / tuzatish, not the edit form.
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("product_edit", args=[self.product.pk]))
        self.assertNotIn("initial_quantity", response.context["form"].fields)


class SaleFilterExportTests(BaseSetup):
    def setUp(self):
        self.debt_sale = make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=timezone.localdate() + timedelta(days=10),
        )

    def test_filter_by_client(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("sale_list"), {"client": self.client2.pk})
        for sale in response.context["page"].object_list:
            self.assertEqual(sale.client, self.client2)

    def test_debtors_count_is_distinct_clients(self):
        # A second debt sale for the same client must not double-count
        make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=timezone.localdate() + timedelta(days=5),
        )
        self.client.force_login(self.manager)
        response = self.client.get(reverse("sale_list"))
        self.assertEqual(response.context["totals"]["debtors"], 1)

    def test_filter_by_status_debt(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_list"), {"status": "debt"})
        sales = list(response.context["page"].object_list)
        self.assertIn(self.debt_sale, sales)
        self.assertTrue(all(s.is_outstanding for s in sales))

    def test_export_returns_xlsx_scoped_and_filtered(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_export"), {"status": "debt"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])
        rows = read_xlsx(response)
        # header + exactly the one debt sale owned by sales1
        self.assertEqual(len(rows), 2)
        self.assertIn("Mijoz", rows[0])

    def test_sales_export_excludes_other_reps(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_export"))
        values = {str(v) for row in read_xlsx(response) for v in row}
        self.assertNotIn(self.client2.name, values)  # client2 belongs to sales2


class FilterChipTests(BaseSetup):
    def test_sales_chips_resolve_names_and_remove_urls(self):
        self.client.force_login(self.admin)
        url = reverse("sale_list")
        resp = self.client.get(url, {"client": self.client1.pk, "status": "debt"})
        chips = resp.context["active_filters"]
        labels = {c["label"]: c["value"] for c in chips}
        self.assertEqual(labels["Mijoz"], "Mijoz A")
        self.assertEqual(labels["To'lov"], "Qarz")
        # removing the client chip keeps status, drops client + page
        client_chip = next(c for c in chips if c["label"] == "Mijoz")
        self.assertIn("status=debt", client_chip["remove_url"])
        self.assertNotIn("client=", client_chip["remove_url"])


class DebtPageTests(BaseSetup):
    def setUp(self):
        today = timezone.localdate()
        # client1: two open debts (one overdue, one due soon) plus a paid sale
        self.overdue = make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=today - timedelta(days=5),
        )
        self.soon = make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=today + timedelta(days=3),
        )
        self.paid = make_sale(self.client1, self.sales1, self.product)

    def test_debtors_grouped_by_client(self):
        self.client.force_login(self.sales1)
        ctx = self.client.get(reverse("debt_list")).context
        debtors = ctx["debtors"]
        self.assertEqual(len(debtors), 1)  # only client1 owes money
        group = debtors[0]
        self.assertEqual(group["client"], self.client1)
        self.assertEqual(group["count"], 2)               # two open receipts
        self.assertEqual(group["overdue_count"], 1)       # one is overdue
        self.assertEqual(group["earliest"], timezone.localdate() - timedelta(days=5))
        self.assertEqual(group["remaining"], Decimal("480000"))  # 2 × 240000

    def test_paid_client_not_listed(self):
        # client2 only has a paid sale → should not appear as a debtor
        make_sale(self.client2, self.sales2, self.product)
        self.client.force_login(self.manager)
        ctx = self.client.get(reverse("debt_list")).context
        self.assertNotIn(self.client2, {g["client"] for g in ctx["debtors"]})

    def test_scoped_to_sales_rep(self):
        make_sale(
            self.client2, self.sales2, self.product,
            is_debt=True, debt_deadline=timezone.localdate() - timedelta(days=2),
        )
        self.client.force_login(self.sales1)
        ctx = self.client.get(reverse("debt_list")).context
        self.assertEqual({g["client"] for g in ctx["debtors"]}, {self.client1})

    def test_client_debt_detail_lists_open_receipts(self):
        self.client.force_login(self.sales1)
        ctx = self.client.get(reverse("debt_client", args=[self.client1.pk])).context
        self.assertEqual(set(ctx["sales"]), {self.overdue, self.soon})
        self.assertNotIn(self.paid, ctx["sales"])
        self.assertEqual(ctx["total"], Decimal("480000"))

    def test_client_debt_detail_scoped(self):
        # sales1 cannot open another rep's client debt page
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("debt_client", args=[self.client2.pk]))
        self.assertEqual(response.status_code, 404)


class SaleDetailTests(BaseSetup):
    def test_detail_shows_sale_and_payments(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_detail", args=[self.sale1.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["sale"], self.sale1)

    def test_detail_scoped_to_owner(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("sale_detail", args=[self.sale2.pk]))
        self.assertEqual(response.status_code, 404)


class PaymentTests(BaseSetup):
    def _debt_sale(self):
        # 10 kg × 24000 = 240000 total
        return make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=timezone.localdate() + timedelta(days=5),
        )

    def test_new_sale_records_no_payment(self):
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [one_item(self.product, weight="10")])
        self.client.post(reverse("sale_create"), data)
        sale = Sale.objects.latest("created_at")
        self.assertEqual(sale.payments.count(), 0)
        self.assertTrue(sale.is_outstanding)

    def test_mark_paid_settles_the_sale(self):
        # 10 kg × 24000 = 240000; one click records a full cash payment
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        self.client.post(reverse("sale_mark_paid", args=[sale.pk]))
        sale.refresh_from_db()
        self.assertTrue(sale.is_paid)
        self.assertEqual(sale.paid_amount, Decimal("240000"))
        payment = sale.payments.get()
        self.assertEqual(payment.amount, Decimal("240000"))
        self.assertEqual(payment.method, "cash")

    def test_partial_debt_payment_keeps_debt_open(self):
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        self.client.post(reverse("sale_pay", args=[sale.pk]), {"amount": "100000", "method": "cash"})
        sale.refresh_from_db()
        self.assertTrue(sale.is_outstanding)  # still owed
        self.assertEqual(sale.debt_remaining, Decimal("140000"))

    def test_full_debt_payment_closes_debt(self):
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        self.client.post(reverse("sale_pay", args=[sale.pk]), {"amount": "240000", "method": "card"})
        sale.refresh_from_db()
        self.assertTrue(sale.is_paid)
        self.assertEqual(sale.payments.get().kind, "debt")

    def test_transfer_payment_records_commission_from_percent(self):
        sale = self._debt_sale()  # 240000 debt
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_pay", args=[sale.pk]),
            {
                "amount": "200000",
                "method": "transfer",
                "commission_percent": "1.5",
                "note": "Bank o'tkazma",
            },
        )
        sale.refresh_from_db()
        payment = sale.payments.get()
        self.assertEqual(payment.method, "transfer")
        self.assertEqual(payment.commission_percent, Decimal("1.5"))
        self.assertEqual(payment.commission, Decimal("3000.00"))  # 200000 × 1.5%
        self.assertEqual(payment.net_amount, Decimal("197000.00"))  # what hits the till
        self.assertEqual(payment.note, "Bank o'tkazma")
        # Only the net (amount − commission) reduces the debt; the client bears the fee
        self.assertEqual(sale.debt_remaining, Decimal("43000"))  # 240000 − 197000

    def test_commission_ignored_for_non_transfer(self):
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_pay", args=[sale.pk]),
            {"amount": "100000", "method": "cash", "commission_percent": "5"},
        )
        payment = sale.payments.get()
        self.assertEqual(payment.commission, Decimal("0"))
        self.assertEqual(payment.commission_percent, Decimal("0"))

    def test_transfer_grossed_up_settles_debt(self):
        # A transfer can be grossed up over the balance so the net clears it:
        # 240000 / 0.96 = 250000, a 4% fee of 10000 leaves 240000 net.
        sale = self._debt_sale()  # 240000
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_pay", args=[sale.pk]),
            {"amount": "250000", "method": "transfer", "commission_percent": "4"},
        )
        sale.refresh_from_db()
        self.assertTrue(sale.is_paid)
        payment = sale.payments.get()
        self.assertEqual(payment.commission, Decimal("10000.00"))
        self.assertEqual(payment.net_amount, Decimal("240000.00"))
        self.assertEqual(sale.debt_remaining, Decimal("0"))

    def test_client_debt_pay_transfer_credits_net(self):
        # 200000 transfer at 5% → 10000 fee, 190000 net credited to the debt.
        sale = self._debt_sale()  # 240000
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("client_debt_pay", args=[self.client1.pk]),
            {"amount": "200000", "method": "transfer", "commission_percent": "5"},
        )
        sale.refresh_from_db()
        payment = sale.payments.get()
        self.assertEqual(payment.amount, Decimal("200000.00"))
        self.assertEqual(payment.commission, Decimal("10000.00"))
        self.assertEqual(payment.net_amount, Decimal("190000.00"))
        self.assertEqual(sale.debt_remaining, Decimal("50000"))  # 240000 − 190000

    def test_client_debt_pay_distributes_fifo(self):
        today = timezone.localdate()
        older = make_sale(
            self.client1, self.sales1, self.product, is_debt=True, date=today - timedelta(days=10)
        )
        newer = make_sale(
            self.client1, self.sales1, self.product, is_debt=True, date=today - timedelta(days=2)
        )
        self.client.force_login(self.sales1)
        # 300000 clears the older 240000 debt in full, then 60000 of the newer
        self.client.post(
            reverse("client_debt_pay", args=[self.client1.pk]),
            {"amount": "300000", "method": "cash"},
        )
        older.refresh_from_db()
        newer.refresh_from_db()
        self.assertTrue(older.is_paid)  # oldest debt cleared first (FIFO)
        self.assertEqual(older.paid_amount, Decimal("240000"))
        self.assertEqual(newer.debt_remaining, Decimal("180000"))  # partially paid
        self.assertEqual(older.payments.filter(kind="debt").count(), 1)
        self.assertEqual(newer.payments.filter(kind="debt").count(), 1)

    def test_client_debt_pay_capped_at_total(self):
        sale = self._debt_sale()  # 240000
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("client_debt_pay", args=[self.client1.pk]),
            {"amount": "999999999", "method": "cash"},
        )
        self.assertEqual(response.status_code, 200)  # re-rendered with error
        sale.refresh_from_db()
        self.assertTrue(sale.is_outstanding)
        self.assertEqual(sale.payments.count(), 0)

    def test_payment_cannot_exceed_remaining(self):
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("sale_pay", args=[sale.pk]), {"amount": "999999", "method": "cash"}
        )
        self.assertEqual(response.status_code, 200)  # re-rendered with error
        sale.refresh_from_db()
        self.assertTrue(sale.is_outstanding)
        self.assertEqual(sale.payments.count(), 0)

class SaleIntegrityTests(BaseSetup):
    def _paid_sale(self):
        # make_sale with is_debt=False books a full cash payment (240000)
        return make_sale(self.client1, self.sales1, self.product)

    def _edit_post(self, sale, weight):
        item = sale.items.get()
        return {
            "date": sale.date.isoformat(),
            "client": self.client1.pk,
            "debt_days": str((sale.debt_deadline - sale.date).days),
            "items-TOTAL_FORMS": "1",
            "items-INITIAL_FORMS": "1",
            "items-MIN_NUM_FORMS": "1",
            "items-MAX_NUM_FORMS": "1000",
            "items-0-id": item.pk,
            "items-0-product": self.product.pk,
            "items-0-dimension": "kg",
            "items-0-weight": weight,
            "items-0-price": "24000",
            "items-0-cost_price": "18000",
        }

    def test_cannot_delete_sale_with_payments(self):
        sale = self._paid_sale()
        self.client.force_login(self.sales1)
        self.client.post(reverse("sale_delete", args=[sale.pk]))
        self.assertTrue(Sale.objects.filter(pk=sale.pk).exists())

    def test_can_delete_sale_without_payments(self):
        sale = make_sale(self.client1, self.sales1, self.product, is_debt=True)
        self.client.force_login(self.sales1)
        self.client.post(reverse("sale_delete", args=[sale.pk]))
        self.assertFalse(Sale.objects.filter(pk=sale.pk).exists())

    def test_edit_cannot_drop_total_below_paid(self):
        sale = self._paid_sale()  # 240000 paid in full
        self.client.force_login(self.sales1)
        # 5 kg × 24000 = 120000, below the 240000 already paid — must be rejected
        response = self.client.post(
            reverse("sale_edit", args=[sale.pk]), self._edit_post(sale, "5")
        )
        self.assertEqual(response.status_code, 200)  # re-rendered with error
        self.assertEqual(sale.items.get().weight, Decimal("10"))  # unchanged

    def test_edit_allowed_when_total_stays_at_or_above_paid(self):
        sale = self._paid_sale()
        self.client.force_login(self.sales1)
        # 12 kg × 24000 = 288000, above the 240000 paid — allowed
        self.client.post(reverse("sale_edit", args=[sale.pk]), self._edit_post(sale, "12"))
        self.assertEqual(sale.items.get().weight, Decimal("12"))

    def test_sale_rejects_zero_weight(self):
        self.client.force_login(self.sales1)
        before = Sale.objects.count()
        data = sale_post(self.client1.pk, [one_item(self.product, weight="0")])
        self.client.post(reverse("sale_create"), data)
        self.assertEqual(Sale.objects.count(), before)

    def test_sale_rejects_zero_price(self):
        self.client.force_login(self.sales1)
        before = Sale.objects.count()
        data = sale_post(self.client1.pk, [one_item(self.product, weight="5", price="0")])
        self.client.post(reverse("sale_create"), data)
        self.assertEqual(Sale.objects.count(), before)


class PaymentVoidTests(BaseSetup):
    def test_manager_can_void_payment_and_restore_debt(self):
        sale = make_sale(self.client1, self.sales1, self.product)  # paid 240000
        payment = sale.payments.get()
        self.client.force_login(self.manager)
        self.client.post(reverse("payment_delete", args=[payment.pk]))
        sale.refresh_from_db()
        self.assertFalse(Payment.objects.filter(pk=payment.pk).exists())
        self.assertTrue(sale.is_outstanding)
        self.assertEqual(sale.debt_remaining, Decimal("240000"))

    def test_sales_can_void_own_payment(self):
        # A seller manages their own money actions: they may void a payment they
        # recorded themselves (the debt is restored automatically).
        sale = make_sale(self.client1, self.sales1, self.product)
        payment = sale.payments.get()  # created_by == sales1
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("payment_delete", args=[payment.pk]))
        self.assertIn(response.status_code, (200, 302))
        self.assertFalse(Payment.objects.filter(pk=payment.pk).exists())

    def test_sales_cannot_void_others_payment(self):
        # Owner scoping still holds — another seller's payment is out of reach (404).
        sale = make_sale(self.client2, self.sales2, self.product)
        payment = sale.payments.get()  # created_by == sales2
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("payment_delete", args=[payment.pk]))
        self.assertEqual(response.status_code, 404)
        self.assertTrue(Payment.objects.filter(pk=payment.pk).exists())

    def test_void_frees_sale_for_deletion(self):
        # The delete guard blocks a paid sale; voiding its payment unblocks it
        sale = make_sale(self.client1, self.sales1, self.product)
        payment = sale.payments.get()
        self.client.force_login(self.manager)
        self.client.post(reverse("payment_delete", args=[payment.pk]))
        self.client.post(reverse("sale_delete", args=[sale.pk]))
        self.assertFalse(Sale.objects.filter(pk=sale.pk).exists())


class DebtFilterTests(BaseSetup):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        # one overdue debt for client1/sales1, one current debt for client2/sales2
        cls.overdue = make_sale(
            cls.client1, cls.sales1, cls.product, is_debt=True,
            debt_deadline=timezone.localdate() - timedelta(days=3),
        )
        cls.current = make_sale(
            cls.client2, cls.sales2, cls.product, is_debt=True,
            debt_deadline=timezone.localdate() + timedelta(days=10),
        )

    def _debtors(self, **params):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("debt_list"), params)
        return {g["client"].pk for g in resp.context["debtors"]}

    def test_filter_by_client(self):
        self.assertEqual(self._debtors(client=self.client1.pk), {self.client1.pk})

    def test_filter_by_rep(self):
        self.assertEqual(self._debtors(rep=self.sales2.pk), {self.client2.pk})

    def test_overdue_only(self):
        self.assertEqual(self._debtors(overdue="1"), {self.client1.pk})


class QuickAddClientTests(BaseSetup):
    def test_creates_client_owned_by_current_user(self):
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("client_quick_create"), {"name": "Tez Mijoz", "phone": "+998900000000"}
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["text"], "Tez Mijoz")
        self.assertEqual(Client.objects.get(pk=data["id"]).owner, self.sales1)

    def test_requires_name(self):
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("client_quick_create"), {"name": "  "})
        self.assertEqual(response.status_code, 400)

    def test_get_not_allowed(self):
        self.client.force_login(self.sales1)
        self.assertEqual(self.client.get(reverse("client_quick_create")).status_code, 405)


class ClientDuplicateTests(BaseSetup):
    # BaseSetup: client1 "Mijoz A" (owner sales1), client2 "Mijoz B" (owner sales2)

    def test_create_blocks_same_name(self):
        self.client.force_login(self.sales1)
        self.client.post(reverse("client_create"), {"name": "Mijoz A"})
        self.assertEqual(Client.objects.filter(name="Mijoz A").count(), 1)

    def test_create_allows_same_name_with_override(self):
        self.client.force_login(self.sales1)
        self.client.post(reverse("client_create"), {"name": "Mijoz A", "allow_duplicate": "on"})
        self.assertEqual(Client.objects.filter(name="Mijoz A").count(), 2)

    def test_sales_duplicate_scoped_to_own_clients(self):
        # "Mijoz B" belongs to sales2; sales1 doesn't see it, so no clash
        self.client.force_login(self.sales1)
        self.client.post(reverse("client_create"), {"name": "Mijoz B"})
        self.assertEqual(Client.objects.filter(name="Mijoz B").count(), 2)

    def test_manager_duplicate_checks_all_clients(self):
        self.client.force_login(self.manager)
        # Managers pick the responsible employee (required for them on the form).
        self.client.post(
            reverse("client_create"), {"name": "Mijoz A", "owner": self.manager.pk}
        )
        self.assertEqual(Client.objects.filter(name="Mijoz A").count(), 1)  # blocked

    def test_quick_create_reports_duplicate(self):
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("client_quick_create"), {"name": "Mijoz A"})
        self.assertEqual(response.status_code, 409)
        data = response.json()
        self.assertTrue(data["duplicate"])
        self.assertEqual(data["existing"]["id"], self.client1.pk)

    def test_quick_create_override_creates_duplicate(self):
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("client_quick_create"), {"name": "Mijoz A", "allow_duplicate": "1"}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Client.objects.filter(name="Mijoz A").count(), 2)


class ClientSearchTests(BaseSetup):
    def test_search_matches_location(self):
        Client.objects.create(name="Dala Mijoz", address="Sergeli tumani, 5-uy", owner=self.sales1)
        Client.objects.create(name="Chekka Mijoz", address="Chilonzor", owner=self.sales1)
        self.client.force_login(self.admin)  # admin sees every client
        resp = self.client.get(reverse("client_list"), {"q": "sergeli"})
        names = [c.name for c in resp.context["page"].object_list]
        self.assertIn("Dala Mijoz", names)
        self.assertNotIn("Chekka Mijoz", names)


class ClientRepresentativeTests(BaseSetup):
    def test_admin_assigns_representative(self):
        self.client.force_login(self.admin)
        self.client.post(
            reverse("client_create"), {"name": "Vakil Mijoz", "owner": self.sales2.pk}
        )
        self.assertEqual(Client.objects.get(name="Vakil Mijoz").owner, self.sales2)

    def test_seller_client_owned_by_self(self):
        # Sellers don't see the representative field; their client stays theirs.
        self.client.force_login(self.sales1)
        self.client.post(reverse("client_create"), {"name": "O'z Mijozim"})
        self.assertEqual(Client.objects.get(name="O'z Mijozim").owner, self.sales1)


class AuditLogTests(BaseSetup):
    def test_sale_create_is_logged(self):
        self.client.force_login(self.sales1)
        data = sale_post(self.client1.pk, [one_item(self.product, weight="5")])
        self.client.post(reverse("sale_create"), data)
        log = AuditLog.objects.filter(action="create", target_type="Sotuv").latest("created_at")
        self.assertEqual(log.user, self.sales1)

    def test_payment_is_logged(self):
        sale = make_sale(self.client1, self.sales1, self.product, is_debt=True)
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_pay", args=[sale.pk]), {"amount": "100000", "method": "cash"}
        )
        self.assertTrue(
            AuditLog.objects.filter(action="payment", target_id=sale.pk).exists()
        )

    def test_void_is_logged(self):
        sale = make_sale(self.client1, self.sales1, self.product)  # paid
        payment = sale.payments.get()
        self.client.force_login(self.manager)
        self.client.post(reverse("payment_delete", args=[payment.pk]))
        self.assertTrue(AuditLog.objects.filter(action="void", target_id=sale.pk).exists())

    def test_audit_list_open_to_all_but_scoped_for_seller(self):
        # Audit is an "own work" view: a seller may open it but sees only their
        # own actions; admins/managers see everyone's.
        AuditLog.record(self.sales1, AuditLog.Action.CREATE, "Sotuv", 1, "own")
        AuditLog.record(self.manager, AuditLog.Action.CREATE, "Sotuv", 2, "boshqa")
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("audit_list"))
        self.assertEqual(response.status_code, 200)
        user_ids = {log.user_id for log in response.context["page"].object_list}
        self.assertEqual(user_ids, {self.sales1.pk})  # only their own rows
        self.client.force_login(self.manager)
        self.assertEqual(self.client.get(reverse("audit_list")).status_code, 200)


class ReturnTests(BaseSetup):
    def _return(self, sale, weight, price="24000", restock=True):
        data = {
            "product": self.product.pk, "dimension": "kg",
            "weight": weight, "price": price,
        }
        if restock:
            data["restock"] = "on"
        return self.client.post(reverse("sale_return", args=[sale.pk]), data)

    def test_return_reduces_debt(self):
        sale = make_sale(self.client1, self.sales1, self.product, is_debt=True)  # 240000
        self.client.force_login(self.sales1)
        self._return(sale, "4")  # 4 × 24000 = 96000
        sale.refresh_from_db()
        self.assertEqual(sale.returned_amount, Decimal("96000"))
        self.assertEqual(sale.debt_remaining, Decimal("144000"))  # 240000 − 96000

    def _stock(self):
        return Product.objects.get(pk=self.product.pk).current_stock

    def test_restock_return_increases_stock(self):
        sale = make_sale(self.client1, self.sales1, self.product, is_debt=True, weight="10")
        self.client.force_login(self.sales1)
        before = self._stock()
        self._return(sale, "3", restock=True)
        self.assertEqual(self._stock() - before, Decimal("3"))  # 3 kg back in stock

    def test_no_restock_leaves_stock_unchanged(self):
        sale = make_sale(self.client1, self.sales1, self.product, is_debt=True, weight="10")
        self.client.force_login(self.sales1)
        before = self._stock()
        self._return(sale, "3", restock=False)
        self.assertEqual(self._stock() - before, Decimal("0"))  # not restocked

    def test_cannot_return_more_than_sold(self):
        sale = make_sale(self.client1, self.sales1, self.product, is_debt=True, weight="10")
        self.client.force_login(self.sales1)
        self._return(sale, "15")  # more than the 10 kg sold
        self.assertEqual(sale.returns.count(), 0)

    def test_return_is_audited(self):
        sale = make_sale(self.client1, self.sales1, self.product, is_debt=True)
        self.client.force_login(self.sales1)
        self._return(sale, "2")
        self.assertTrue(AuditLog.objects.filter(action="return", target_id=sale.pk).exists())


class ModalFormTests(BaseSetup):
    def _ajax(self):
        return {"HTTP_X_REQUESTED_WITH": "XMLHttpRequest"}

    def test_ajax_get_returns_modal_partial(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("client_create"), **self._ajax())
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "_modal.html")
        self.assertNotContains(response, "<aside")  # no full page chrome

    def test_ajax_success_returns_204_with_redirect(self):
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("client_create"), {"name": "Yangi Mijoz"}, **self._ajax()
        )
        self.assertEqual(response.status_code, 204)
        self.assertEqual(response["X-Redirect"], reverse("client_list"))

    def test_ajax_invalid_returns_422_partial(self):
        self.client.force_login(self.sales1)
        response = self.client.post(reverse("client_create"), {"name": ""}, **self._ajax())
        self.assertEqual(response.status_code, 422)
        self.assertTemplateUsed(response, "_modal.html")

    def test_non_ajax_get_returns_full_page(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("client_create"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "crm/form.html")


class SeedDemoTests(TestCase):
    def test_seed_demo_runs_and_populates(self):
        from django.core.management import call_command

        call_command("seed_demo")
        self.assertTrue(User.objects.filter(username="admin").exists())
        self.assertEqual(Sale.objects.count(), 30)
        # Every seeded sale carries a deadline (there is no is_debt field anymore)
        self.assertFalse(Sale.objects.filter(debt_deadline__isnull=True).exists())


class KassaCurrencyTests(BaseSetup):
    """Dollar payments/expenses and the two-currency kassa (so'm + dollar tills)."""

    def _debt_sale(self):
        # 10 kg × 24000 = 240000 outstanding
        return make_sale(
            self.client1, self.sales1, self.product,
            is_debt=True, debt_deadline=timezone.localdate() + timedelta(days=5),
        )

    def test_dollar_payment_converts_to_som(self):
        # $10 × 12700 = 127000 so'm credited against the debt
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_pay", args=[sale.pk]),
            {"amount": "10", "currency": "usd", "exchange_rate": "12700", "method": "cash"},
        )
        sale.refresh_from_db()
        payment = sale.payments.get()
        self.assertEqual(payment.currency, "usd")
        self.assertEqual(payment.amount, Decimal("127000.00"))        # so'm value
        self.assertEqual(payment.amount_original, Decimal("10.00"))   # the dollars
        self.assertEqual(payment.exchange_rate, Decimal("12700.00"))
        self.assertEqual(sale.debt_remaining, Decimal("113000"))      # 240000 − 127000

    def test_dollar_payment_requires_rate(self):
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("sale_pay", args=[sale.pk]),
            {"amount": "10", "currency": "usd", "method": "cash"},  # no rate
        )
        self.assertEqual(response.status_code, 200)   # re-rendered with errors
        self.assertEqual(sale.payments.count(), 0)    # nothing recorded

    def test_dollar_expense_converts_to_som(self):
        # $20 × 12700 = 254000 so'm
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("expense_create"),
            {
                "date": timezone.localdate().isoformat(),
                "amount": "20", "currency": "usd", "exchange_rate": "12700",
                "category": "purchase", "method": "cash", "note": "dollar rasxot",
            },
        )
        expense = Expense.objects.get(note="dollar rasxot")
        self.assertEqual(expense.currency, "usd")
        self.assertEqual(expense.amount, Decimal("254000.00"))        # so'm value
        self.assertEqual(expense.amount_original, Decimal("20.00"))   # the dollars

    def test_kassa_summary_isolates_dollar_till(self):
        # $10 in and $4 out → a $6 dollar-till balance, independent of the
        # so'm activity in the base fixtures. ($10 × 12700 = 127000 ≤ 240000 debt.)
        sale = self._debt_sale()
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("sale_pay", args=[sale.pk]),
            {"amount": "10", "currency": "usd", "exchange_rate": "12700", "method": "cash"},
        )
        self.client.post(
            reverse("expense_create"),
            {
                "date": timezone.localdate().isoformat(),
                "amount": "4", "currency": "usd", "exchange_rate": "12700",
                "category": "purchase", "method": "cash",
            },
        )
        today = timezone.localdate()
        summary = _kassa_summary(today, today)
        self.assertEqual(summary["usd"]["cash"], Decimal("10.00"))
        self.assertEqual(summary["usd"]["income"], Decimal("10.00"))
        self.assertEqual(summary["usd"]["expense"], Decimal("4.00"))
        self.assertEqual(summary["usd"]["closing"], Decimal("6.00"))

    def test_dollar_expense_edit_prefills_dollars_and_recomputes(self):
        self.client.force_login(self.admin)
        self.client.post(
            reverse("expense_create"),
            {
                "date": timezone.localdate().isoformat(),
                "amount": "20", "currency": "usd", "exchange_rate": "12700",
                "category": "purchase", "method": "cash", "note": "edit-me",
            },
        )
        expense = Expense.objects.get(note="edit-me")
        # The edit form shows the original dollars, not the stored so'm.
        get = self.client.get(reverse("expense_edit", args=[expense.pk]))
        self.assertEqual(get.context["form"].initial["amount"], Decimal("20.00"))
        # Re-saving at $30 reconverts the so'm value.
        self.client.post(
            reverse("expense_edit", args=[expense.pk]),
            {
                "date": timezone.localdate().isoformat(),
                "amount": "30", "currency": "usd", "exchange_rate": "12700",
                "category": "purchase", "method": "cash", "note": "edit-me",
            },
        )
        expense.refresh_from_db()
        self.assertEqual(expense.amount_original, Decimal("30.00"))
        self.assertEqual(expense.amount, Decimal("381000.00"))  # 30 × 12700

    def test_seller_can_edit_own_but_not_others_expense(self):
        # Kassa is shared, but an expense stays owner-scoped: a seller may edit
        # the expense they entered, not one an admin (or another seller) entered.
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("expense_create"),
            {
                "date": timezone.localdate().isoformat(), "amount": "50000",
                "currency": "uzs", "category": "fuel", "method": "cash", "note": "own",
            },
        )
        own = Expense.objects.get(note="own")
        self.assertEqual(
            self.client.get(reverse("expense_edit", args=[own.pk])).status_code, 200
        )
        # An admin-entered expense is outside the seller's scope → 404.
        self.client.force_login(self.admin)
        self.client.post(
            reverse("expense_create"),
            {
                "date": timezone.localdate().isoformat(), "amount": "70000",
                "currency": "uzs", "category": "fuel", "method": "cash", "note": "admins",
            },
        )
        others = Expense.objects.get(note="admins")
        self.client.force_login(self.sales1)
        self.assertEqual(
            self.client.get(reverse("expense_edit", args=[others.pk])).status_code, 404
        )

    def test_expense_xlsx_export(self):
        self.client.force_login(self.admin)
        self.client.post(
            reverse("expense_create"),
            {
                "date": timezone.localdate().isoformat(),
                "amount": "20", "currency": "usd", "exchange_rate": "12700",
                "category": "purchase", "method": "cash", "note": "xlsx-row",
            },
        )
        response = self.client.get(reverse("expense_export"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        cells = {v for row in read_xlsx(response) for v in row}
        self.assertIn("xlsx-row", cells)
        self.assertIn(254000.0, cells)  # so'm value (numeric cell)
        self.assertIn("Dollar", cells)  # currency label

    def test_per_employee_net_subtracts_expense_from_profit(self):
        # sales2's only sale earns 60000 profit; a 20000 expense they record
        # nets their performance to 40000 (foyda − rasxot).
        self.client.force_login(self.sales2)
        self.client.post(
            reverse("expense_create"),
            {
                "date": timezone.localdate().isoformat(),
                "amount": "20000", "currency": "uzs",
                "category": "fuel", "method": "cash", "note": "sales2 rasxot",
            },
        )
        today = timezone.localdate()
        rows = {r["employee"]: r for r in _per_employee_kassa(today, today)}
        row = rows[str(self.sales2)]
        self.assertEqual(row["profit"], Decimal("60000"))
        self.assertEqual(row["out_som"], Decimal("20000"))
        self.assertEqual(row["net"], Decimal("40000"))  # 60000 − 20000


class KassaSupplierCostTests(BaseSetup):
    """'Jami tannarx' — supplier cost (asl narx) of goods sold in the window, so
    the middleman knows what it owes suppliers. Base fixtures: two sales today,
    each 10 kg × 18 000 tannarx = 180 000 → 360 000 total."""

    def test_supplier_cost_sums_window_sales(self):
        today = timezone.localdate()
        self.assertEqual(_kassa_summary(today, today)["cost"], Decimal("360000"))

    def test_supplier_cost_excludes_out_of_window_sales(self):
        # A sale dated last week must not count toward today's supplier cost.
        make_sale(
            self.client1, self.sales1, self.product,
            date=timezone.localdate() - timedelta(days=7),
        )
        today = timezone.localdate()
        self.assertEqual(_kassa_summary(today, today)["cost"], Decimal("360000"))

    def test_supplier_cost_respects_employee_filter(self):
        # Scoped to one seller → only that seller's sale cost (180 000).
        today = timezone.localdate()
        summary = _kassa_summary(today, today, rep=self.sales1)
        self.assertEqual(summary["cost"], Decimal("180000"))


class KassaScopingTests(BaseSetup):
    def setUp(self):
        today = timezone.localdate()
        Expense.objects.create(
            amount=Decimal("50000"), category=Expense.Category.OTHER,
            method=Payment.Method.CASH, created_by=self.sales2, date=today,
        )
        Expense.objects.create(
            amount=Decimal("30000"), category=Expense.Category.OTHER,
            method=Payment.Method.CASH, created_by=self.sales1, date=today,
        )

    def test_seller_sees_only_own_expenses(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("kassa"))
        creators = {e.created_by_id for e in response.context["expenses"]}
        self.assertEqual(creators, {self.sales1.pk})

    def test_seller_cannot_widen_scope_via_rep_param(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("kassa"), {"rep": self.sales2.pk})
        creators = {e.created_by_id for e in response.context["expenses"]}
        self.assertEqual(creators, {self.sales1.pk})

    def test_seller_transactions_scoped_to_self(self):
        # The kirim/chiqim ledgers show only the seller's own rows.
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("kassa"))
        income = response.context["income_rows"]
        outflow = response.context["outflow_rows"]
        creators = {t["created_by"].pk for t in income + outflow}
        self.assertEqual(creators, {self.sales1.pk})
        self.assertTrue(any(t["direction"] == "in" for t in income))    # own sale payment
        self.assertTrue(any(t["direction"] == "out" for t in outflow))  # own expense

    def test_seller_has_no_per_employee_table(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("kassa"))
        self.assertIsNone(response.context["per_employee"])
        self.assertIsNone(response.context["reps"])

    def test_admin_sees_all_expenses_and_per_employee(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("kassa"))
        creators = {e.created_by_id for e in response.context["expenses"]}
        self.assertEqual(creators, {self.sales1.pk, self.sales2.pk})
        self.assertIsNotNone(response.context["per_employee"])

    def test_seller_response_hides_employee_table(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("kassa"))
        self.assertNotContains(response, "Sotuvchilar nazorati")

    def test_admin_response_shows_employee_table(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("kassa"))
        self.assertContains(response, "Sotuvchilar nazorati")


class RemittanceTests(BaseSetup):
    """Ishlab chiqarishga topshirish — a seller handing cash back to production.
    It repays the seller→production debt (= tannarx of goods they've sold) and
    leaves their till, without touching client debts."""

    def setUp(self):
        today = timezone.localdate()
        # sales1 sells 10 kg on debt: revenue 240k, tannarx (cost) 180k.
        self.sale = make_sale(
            self.client1, self.sales1, self.product, weight="10",
            date=today, is_debt=True,
        )
        # Client repays 200k so the seller has cash on hand to remit.
        Payment.objects.create(
            sale=self.sale, amount=Decimal("200000"), method=Payment.Method.CASH,
            kind=Payment.Kind.DEBT, date=today, created_by=self.sales1,
        )

    def _remit(self, user, amount, seller=None):
        self.client.force_login(user)
        data = {
            "date": timezone.localdate().isoformat(),
            "amount": amount, "method": "cash",
        }
        if seller is not None:
            data["seller"] = seller.pk
        return self.client.post(reverse("remittance_create"), data)

    def test_seller_remittance_reduces_debt_and_cash(self):
        today = timezone.localdate()
        before = _kassa_summary(today, today, rep=self.sales1)

        self._remit(self.sales1, "150000")
        remit = ProductionRemittance.objects.get()
        self.assertEqual(remit.seller, self.sales1)
        self.assertEqual(remit.amount, Decimal("150000"))

        after = _kassa_summary(today, today, rep=self.sales1)
        # A handover drops BOTH the production debt and the cash on hand by its
        # amount — and nothing else.
        self.assertEqual(before["production_debt"] - after["production_debt"], Decimal("150000"))
        self.assertEqual(before["cash"] - after["cash"], Decimal("150000"))
        self.assertEqual(after["remitted"], Decimal("150000.00"))
        # The client's debt is untouched by a production handover.
        self.assertEqual(self.sale.debt_remaining, Decimal("40000.00"))  # 240k − 200k

    def test_seller_field_is_pinned_to_self(self):
        # A seller cannot file a handover on another seller's behalf, even by
        # POSTing a different seller id — it snaps back to themselves.
        self._remit(self.sales1, "10000", seller=self.sales2)
        remit = ProductionRemittance.objects.get()
        self.assertEqual(remit.seller, self.sales1)

    def test_admin_can_file_for_any_seller(self):
        self._remit(self.admin, "20000", seller=self.sales2)
        remit = ProductionRemittance.objects.get()
        self.assertEqual(remit.seller, self.sales2)
        self.assertEqual(remit.created_by, self.admin)

    def test_seller_sees_only_own_remittance_in_ledger(self):
        ProductionRemittance.objects.create(
            seller=self.sales2, amount=Decimal("9999"), method=Payment.Method.CASH,
            created_by=self.sales2, date=timezone.localdate(),
        )
        self._remit(self.sales1, "5000")
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("kassa"))
        # Handovers land in the chiqim ledger (direction "remit").
        remits = [
            t for t in response.context["outflow_rows"]
            if t["kind"] == "remittance"
        ]
        self.assertTrue(remits)
        self.assertTrue(all(t["created_by"].pk == self.sales1.pk for t in remits))

    def test_remittance_is_audited(self):
        self._remit(self.sales1, "12000")
        log = AuditLog.objects.filter(target_type="Topshiruv").first()
        self.assertIsNotNone(log)
        self.assertEqual(log.action, AuditLog.Action.CREATE)


class RealizedProfitTests(BaseSetup):
    """Kassa profit is realized cost-first: an unpaid debt sale earns nothing until
    its tannarx is collected; only collections above cost count as profit."""

    def _pay(self, sale, amount):
        Payment.objects.create(
            sale=sale, amount=Decimal(amount), method=Payment.Method.CASH,
            kind=Payment.Kind.DEBT, date=timezone.localdate(), created_by=self.manager,
        )

    def test_cost_first_realization(self):
        # revenue 2,000,000 · tannarx 1,500,000 · unpaid (manager has no other sales)
        sale = make_sale(
            self.client1, self.manager, self.product,
            weight="100", price="20000", cost_price="15000", is_debt=True,
        )
        today = timezone.localdate()

        def realized():
            got = _realized_profit_by_seller(today, today, rep=self.manager)
            return got.get(self.manager.pk, Decimal("0"))

        self.assertEqual(realized(), Decimal("0"))         # nothing collected yet
        self._pay(sale, "1000000")
        self.assertEqual(realized(), Decimal("0"))         # 1.0M < 1.5M cost → still 0
        self._pay(sale, "750000")                          # collected 1.75M
        self.assertEqual(realized(), Decimal("250000"))    # 1.75M − 1.5M
        self._pay(sale, "250000")                          # collected 2.0M (full)
        self.assertEqual(realized(), Decimal("500000"))    # capped at the full margin


class ClientTransferTests(BaseSetup):
    def test_seller_transfers_own_client(self):
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("client_transfer", args=[self.client1.pk]),
            {"new_owner": self.sales2.pk},
        )
        self.assertEqual(response.status_code, 302)
        self.client1.refresh_from_db()
        self.assertEqual(self.client1.owner, self.sales2)
        self.assertEqual(
            list(Sale.objects.filter(client=self.client1).values_list("sales_rep", flat=True)),
            [self.sales2.pk],
        )

    def test_transfer_moves_all_of_a_clients_sales(self):
        make_sale(self.client1, self.sales1, self.product)  # a second sale
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("client_transfer", args=[self.client1.pk]),
            {"new_owner": self.sales2.pk},
        )
        reps = set(
            Sale.objects.filter(client=self.client1).values_list("sales_rep", flat=True)
        )
        self.assertEqual(reps, {self.sales2.pk})

    def test_seller_cannot_transfer_another_sellers_client(self):
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("client_transfer", args=[self.client2.pk]),
            {"new_owner": self.sales1.pk},
        )
        self.assertEqual(response.status_code, 404)
        self.client2.refresh_from_db()
        self.assertEqual(self.client2.owner, self.sales2)

    def test_admin_can_transfer_any_client(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("client_transfer", args=[self.client2.pk]),
            {"new_owner": self.sales1.pk},
        )
        self.assertEqual(response.status_code, 302)
        self.client2.refresh_from_db()
        self.assertEqual(self.client2.owner, self.sales1)

    def test_cannot_transfer_to_current_owner(self):
        self.client.force_login(self.sales1)
        response = self.client.post(
            reverse("client_transfer", args=[self.client1.pk]),
            {"new_owner": self.sales1.pk},
        )
        self.assertEqual(response.status_code, 200)  # re-rendered with error
        self.client1.refresh_from_db()
        self.assertEqual(self.client1.owner, self.sales1)

    def test_transfer_writes_audit_log(self):
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("client_transfer", args=[self.client1.pk]),
            {"new_owner": self.sales2.pk},
        )
        log = AuditLog.objects.filter(action=AuditLog.Action.TRANSFER).first()
        self.assertIsNotNone(log)
        self.assertEqual(log.target_id, self.client1.pk)

    def test_new_owner_gains_and_old_owner_loses_visibility(self):
        make_sale(self.client1, self.sales1, self.product, is_debt=True)
        self.client.force_login(self.sales1)
        self.client.post(
            reverse("client_transfer", args=[self.client1.pk]),
            {"new_owner": self.sales2.pk},
        )
        self.client.force_login(self.sales2)
        resp2 = self.client.get(reverse("client_list"))
        self.assertIn(self.client1, list(resp2.context["page"].object_list))
        self.client.force_login(self.sales1)
        resp1 = self.client.get(reverse("client_list"))
        self.assertNotIn(self.client1, list(resp1.context["page"].object_list))

    def test_transfer_modal_renders_for_ajax(self):
        self.client.force_login(self.sales1)
        response = self.client.get(
            reverse("client_transfer", args=[self.client1.pk]),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Yangi sotuvchi")
        self.assertContains(response, "O'tkazish")

    def test_client_list_shows_transfer_action(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("client_list"))
        self.assertContains(response, reverse("client_transfer", args=[self.client1.pk]))

    def test_transfer_non_ajax_get_renders(self):
        # A direct (non-XHR) GET falls back to the full-page form and still
        # renders the target dropdown, so the non-modal path is not broken.
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("client_transfer", args=[self.client1.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="new_owner"')


class ProductDetailScopingTests(BaseSetup):
    def test_seller_sees_only_own_recent_sales(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("product_detail", args=[self.product.pk]))
        sale_ids = {i.sale_id for i in response.context["recent_items"]}
        self.assertIn(self.sale1.pk, sale_ids)
        self.assertNotIn(self.sale2.pk, sale_ids)

    def test_seller_sees_stock_entries(self):
        # Warehouse is shared — a seller sees the stock-movement log context too.
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("product_detail", args=[self.product.pk]))
        self.assertIsNotNone(response.context["entries"])

    def test_admin_sees_all_recent_sales_and_entries(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("product_detail", args=[self.product.pk]))
        sale_ids = {i.sale_id for i in response.context["recent_items"]}
        self.assertIn(self.sale1.pk, sale_ids)
        self.assertIn(self.sale2.pk, sale_ids)
        self.assertIsNotNone(response.context["entries"])

    def test_seller_response_shows_stock_log(self):
        # The shared warehouse log renders for sellers as well as admins.
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("product_detail", args=[self.product.pk]))
        self.assertContains(response, "Ombor harakatlari")

    def test_admin_response_shows_stock_log(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("product_detail", args=[self.product.pk]))
        self.assertContains(response, "Ombor harakatlari")


class SellerLabelTests(BaseSetup):
    def test_seller_sees_personalized_labels(self):
        self.client.force_login(self.sales1)
        response = self.client.get(reverse("client_list"))
        self.assertContains(response, "Mening mijozlarim")
        response = self.client.get(reverse("sale_list"))
        self.assertContains(response, "Mening sotuvlarim")

    def test_admin_sees_neutral_labels(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("client_list"))
        self.assertNotContains(response, "Mening mijozlarim")


class SaleFormClientSearchTests(TestCase):
    """The client picker on the sale form must be searchable by name, phone or
    address. The widget carries that data on each <option> so the front-end
    combobox can filter on it and show a phone · address subtitle."""

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user("cs_admin", password="x", role=User.Role.ADMIN)
        cls.client_obj = Client.objects.create(
            name="Ali Valiyev",
            company="Valiyev Savdo",
            phone="+998 90 123 45 67",
            address="Chilonzor, Toshkent",
            owner=cls.admin,
        )

    def test_option_carries_search_haystack_with_digits_only_phone(self):
        html = str(SaleForm(user=self.admin)["client"])
        # Name, company, phone and address are all searchable, lowercased, plus a
        # digits-only copy of the phone so "998901234567" matches too.
        self.assertIn(
            'data-search="ali valiyev valiyev savdo +998 90 123 45 67 '
            'chilonzor, toshkent 998901234567"',
            html,
        )

    def test_option_carries_phone_address_subtitle(self):
        html = str(SaleForm(user=self.admin)["client"])
        self.assertIn(
            'data-subtitle="+998 90 123 45 67 · Chilonzor, Toshkent"', html
        )

    def test_empty_choice_has_no_search_metadata(self):
        # Only the one real client option is searchable; the "— choose —" blank
        # option carries no metadata (nothing to guard against a missing instance).
        html = str(SaleForm(user=self.admin)["client"])
        self.assertEqual(html.count("data-search="), 1)

    def test_blank_option_has_no_dashes_and_shows_placeholder(self):
        html = str(SaleForm(user=self.admin)["client"])
        # Django's "---------" placeholder row is gone; the combobox shows a real
        # placeholder in the input instead.
        self.assertNotIn("---------", html)
        self.assertIn('data-placeholder="Mijozni qidiring yoki tanlang"', html)
        # The blank option itself stays (so a required pick is still enforced),
        # just with an empty label.
        self.assertIn('<option value=""', html)


class SaleItemSummaryTests(BaseSetup):
    """The sales-list product summary shows the first product's name and SKU,
    plus a "+N" when the sale has more than one product."""

    def test_summary_shows_first_product_name_and_sku(self):
        self.assertEqual(self.sale1.item_summary, "Polietilen paket · PKT-1")

    def test_summary_appends_plus_n_for_extra_products(self):
        other = Product.objects.create(name="Qora paket", sku="PKT-2", price=Decimal("10000"))
        SaleItem.objects.create(
            sale=self.sale1, product=other, dimension="kg",
            weight=Decimal("2"), price=Decimal("10000"), cost_price=Decimal("8000"),
        )
        self.assertEqual(self.sale1.item_summary, "Polietilen paket · PKT-1  +1")


class TimeagoUzTests(TestCase):
    """The Uzbek relative-time filter used for a client's last-sale column."""

    def test_relative_phrases(self):
        from crm.templatetags.crm_extras import timeago_uz

        today = timezone.localdate()
        self.assertEqual(timeago_uz(None), "")
        self.assertEqual(timeago_uz(today), "Bugun")
        self.assertEqual(timeago_uz(today - timedelta(days=1)), "Kecha")
        self.assertEqual(timeago_uz(today - timedelta(days=5)), "5 kun oldin")
        self.assertEqual(timeago_uz(today - timedelta(days=60)), "2 oy oldin")
        self.assertEqual(timeago_uz(today - timedelta(days=400)), "1 yil oldin")


class ClientLastSaleTests(BaseSetup):
    """The client list shows how long ago each customer's last sale was, and
    "Hech qachon" for a customer who has never bought."""

    def test_list_shows_time_since_last_sale(self):
        recent = Client.objects.create(name="Yaqin mijoz", owner=self.sales1)
        make_sale(recent, self.sales1, self.product,
                  date=timezone.localdate() - timedelta(days=3))
        Client.objects.create(name="Sotuvsiz mijoz", owner=self.sales1)  # never bought

        self.client.force_login(self.admin)
        response = self.client.get(reverse("client_list"))
        self.assertContains(response, "Oxirgi sotuv")
        self.assertContains(response, "3 kun oldin")
        self.assertContains(response, "Hech qachon")
